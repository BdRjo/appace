"""
إدارة المستخدمين والأدوار والصلاحيات وسجل الدخول
مطابق حرفياً لـ v54: show_users + import_users_from_csv + send_bulk_message + print_selected_users
"""
import hashlib, csv, io
from datetime import datetime
from utils.flash_helper import flash_msg
from flask import (Blueprint, render_template, redirect, url_for,
                   request, flash, abort, jsonify, Response)
from flask_login import login_required, current_user
from models.database import (User, Role, Permission, RolePermission,
                              SystemLog, LoginLog, Location, Venue)
from utils.helpers import get_db, admin_required, syslog, paginate

users_bp = Blueprint('users', __name__, url_prefix='/users')

def _hash(pw): return hashlib.sha256(pw.encode()).hexdigest()
def _log(db, action, desc):
    db.add(SystemLog(action=action, description=desc,
                     user_id=current_user.id, level='info'))
    db.commit()


# ── Users list ────────────────────────────────────────────────────────────────
@users_bp.route('/')
@login_required
@admin_required
def index():
    from sqlalchemy import func
    from models.database import Reservation
    db     = get_db()
    search     = request.args.get('q','')
    role_f     = request.args.get('role_id','')
    dept_f     = request.args.get('department','')
    page       = request.args.get('page',1,type=int)

    q = db.query(User)
    if search:
        q = q.filter((User.username.ilike(f'%{search}%')) |
                     (User.full_name.ilike(f'%{search}%')) |
                     (User.email.ilike(f'%{search}%')) |
                     (User.department.ilike(f'%{search}%')) |
                     (User.job_title.ilike(f'%{search}%')))
    if role_f and role_f.isdigit():
        q = q.filter(User.role_id == int(role_f))
    if dept_f:
        q = q.filter(User.department == dept_f)
    items, total, total_pages = paginate(q.order_by(User.created_at.desc()), page, 25)

    user_ids = [u.id for u in items]
    counts = {}
    if user_ids:
        rows = db.query(Reservation.user_id, func.count(Reservation.id))\
                 .filter(Reservation.user_id.in_(user_ids))\
                 .group_by(Reservation.user_id).all()
        counts = {row[0]: row[1] for row in rows}

    roles = db.query(Role).all()
    # Distinct department values for filter dropdown
    all_depts = [r[0] for r in db.query(User.department).distinct().filter(User.department != None, User.department != '').order_by(User.department).all()]
    return render_template('admin/users_mgmt.html',
        users=items, roles=roles, search=search, role_f=role_f, dept_f=dept_f,
        all_depts=all_depts,
        page=page, total=total, total_pages=total_pages,
        res_counts=counts)


# ── Add user ──────────────────────────────────────────────────────────────────
@users_bp.route('/new', methods=['GET','POST'])
@login_required
@admin_required
def new():
    db    = get_db()
    roles = db.query(Role).all()
    if request.method == 'POST':
        username  = request.form.get('username','').strip()
        full_name = request.form.get('full_name','').strip()
        email     = request.form.get('email','').strip()
        phone     = request.form.get('phone','').strip()
        role_id   = request.form.get('role_id','')
        password  = request.form.get('password','').strip()
        is_active = request.form.get('is_active') == 'on'

        errors = []
        if not username:  errors.append('اسم المستخدم مطلوب')
        if not full_name: errors.append('الاسم الكامل مطلوب')
        if not password:  errors.append('كلمة المرور مطلوبة')
        if len(password) < 6: errors.append('كلمة المرور 6 أحرف على الأقل')
        if db.query(User).filter_by(username=username).first():
            errors.append('اسم المستخدم مستخدم مسبقاً')
        if email and db.query(User).filter_by(email=email).first():
            errors.append('البريد الإلكتروني مستخدم مسبقاً')

        if errors:
            for e in errors: flash(e,'danger')
            return render_template('admin/user_form.html',
                                   user=None, roles=roles, form=request.form)

        job_title = request.form.get('job_title','').strip() or None
        department= request.form.get('department','').strip() or None
        user = User(username=username, full_name=full_name,
                    email=email or None, phone=phone or None,
                    job_title=job_title, department=department,
                    role_id=int(role_id) if role_id else None,
                    password_hash=_hash(password),
                    is_active=is_active, is_verified=True)
        db.add(user)
        db.commit()

        # إرسال بريد ترحيب مثل v54
        try:
            from utils.email_helper import send_welcome
            send_welcome(user, password, request.host_url)
        except: pass

        _log(db,'ADD_USER',f'مستخدم جديد: {username}')
        flash_msg(f'✅ تمت إضافة المستخدم: {username}', 'success')
        return redirect(url_for('users.index'))

    return render_template('admin/user_form.html', user=None, roles=roles, form={})


# ── Edit user ─────────────────────────────────────────────────────────────────
@users_bp.route('/<int:user_id>/edit', methods=['GET','POST'])
@login_required
@admin_required
def edit(user_id):
    db    = get_db()
    user  = db.query(User).get(user_id)
    if not user: abort(404)
    roles = db.query(Role).all()

    if request.method == 'POST':
        user.full_name  = request.form.get('full_name','').strip()
        user.email      = request.form.get('email','').strip() or None
        user.phone      = request.form.get('phone','').strip() or None
        user.job_title  = request.form.get('job_title','').strip() or None
        user.department = request.form.get('department','').strip() or None
        user.is_active  = request.form.get('is_active') == 'on'
        rid = request.form.get('role_id','')
        if rid and rid.isdigit(): user.role_id = int(rid)
        new_pw = request.form.get('password','').strip()
        if new_pw:
            if len(new_pw) < 6:
                flash_msg('كلمة المرور 6 أحرف على الأقل', 'danger')
                return render_template('admin/user_form.html',
                                       user=user, roles=roles, form=request.form)
            user.password_hash = _hash(new_pw)

        # Location access — مطابق لـ v54 LocationWindow access tab
        loc_ids = request.form.getlist('allowed_locations', type=int)
        all_locs = db.query(Location).all()
        user.allowed_locations = [l for l in all_locs if l.id in loc_ids]

        # Venue access — مطابق لـ v54 VenueWindow access tab
        ven_ids = request.form.getlist('allowed_venues', type=int)
        all_venues = db.query(Venue).all()
        user.allowed_venues = [v for v in all_venues if v.id in ven_ids]

        db.commit()
        _log(db,'EDIT_USER',f'تعديل المستخدم: {user.username}')
        flash_msg(f'✅ تم تحديث بيانات: {user.username}', 'success')
        return redirect(url_for('users.index'))

    locs   = db.query(Location).all()
    venues = db.query(Venue).all()
    return render_template('admin/user_form.html',
        user=user, roles=roles, form={},
        all_locations=locs, all_venues=venues,
        assigned_locations={l.id for l in user.allowed_locations},
        assigned_venues={v.id for v in user.allowed_venues})


@users_bp.route('/bulk-delete', methods=['POST'])
@login_required
@admin_required
def bulk_delete():
    from models.database import (Reservation, ChecklistItem, Approval,
                                  BookingContact, Rating, LoginLog,
                                  Checklist, BlockedPeriod, EmailLog,
                                  SystemLog, Contact, Attachment)
    from sqlalchemy import text
    db = get_db(); ids = request.form.getlist('ids'); count = 0
    for uid_str in ids:
        try:
            uid = int(uid_str)
            user = db.query(User).get(uid)
            if not user or user.id == current_user.id: continue
            db.query(Reservation).filter(Reservation.approver_id == uid).update({'approver_id': None})
            db.query(Reservation).filter(Reservation.cancelled_by == uid).update({'cancelled_by': None})
            db.query(Reservation).filter(Reservation.requested_employee_id == uid).update({'requested_employee_id': None})
            db.query(ChecklistItem).filter(ChecklistItem.checked_by_id == uid).update({'checked_by_id': None})
            db.query(Checklist).filter(Checklist.created_by_id == uid).update({'created_by_id': None})
            db.query(BlockedPeriod).filter(BlockedPeriod.created_by_id == uid).update({'created_by_id': None})
            db.query(Approval).filter(Approval.approver_id == uid).update({'approver_id': None})
            db.query(Contact).filter(Contact.created_by == uid).update({'created_by': None})
            db.query(SystemLog).filter(SystemLog.user_id == uid).update({'user_id': None})
            db.query(EmailLog).filter(EmailLog.user_id == uid).update({'user_id': None})
            try: db.query(Attachment).filter(Attachment.uploaded_by == uid).update({'uploaded_by': None})
            except: pass
            db.flush()
            for res in db.query(Reservation).filter(Reservation.user_id == uid).all():
                db.query(ChecklistItem).filter(ChecklistItem.reservation_id == res.id).delete()
                db.query(Approval).filter(Approval.reservation_id == res.id).delete()
                db.query(BookingContact).filter(BookingContact.booking_id == res.id).delete()
                db.query(Rating).filter(Rating.reservation_id == res.id).delete()
                db.delete(res)
            db.query(Rating).filter(Rating.user_id == uid).delete()
            db.flush()
            db.execute(text('DELETE FROM user_locations WHERE user_id=:uid'), {'uid': uid})
            db.execute(text('DELETE FROM user_venues WHERE user_id=:uid'), {'uid': uid})
            db.query(LoginLog).filter(LoginLog.user_id == uid).update({'user_id': None})
            db.flush()
            db.delete(user); count += 1
        except Exception as e:
            db.rollback(); continue
    db.commit()
    _log(db, 'BULK_DELETE_USERS', f'حذف {count} مستخدم')
    flash_msg(f'✅ تم حذف {count} مستخدم', 'success' if count else 'warning')
    return redirect(url_for('users.index'))


# ── Delete user ───────────────────────────────────────────────────────────────
@users_bp.route('/<int:user_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete(user_id):
    from models.database import (Reservation, ChecklistItem, Approval,
                                  BookingContact, Rating, LoginLog,
                                  Checklist, BlockedPeriod, EmailLog,
                                  SystemLog, Contact, Attachment)
    from sqlalchemy import text
    db   = get_db()
    user = db.query(User).get(user_id)
    if not user: abort(404)
    if user.id == current_user.id:
        flash_msg('لا يمكنك حذف حسابك الخاص', 'danger')
        return redirect(url_for('users.index'))
    uname = user.username
    uid = user_id
    try:
        # Null out ALL foreign keys pointing to this user
        db.query(Reservation).filter(Reservation.approver_id == uid).update({'approver_id': None})
        db.query(Reservation).filter(Reservation.cancelled_by == uid).update({'cancelled_by': None})
        db.query(Reservation).filter(Reservation.requested_employee_id == uid).update({'requested_employee_id': None})
        db.query(ChecklistItem).filter(ChecklistItem.checked_by_id == uid).update({'checked_by_id': None})
        db.query(Checklist).filter(Checklist.created_by_id == uid).update({'created_by_id': None})
        db.query(BlockedPeriod).filter(BlockedPeriod.created_by_id == uid).update({'created_by_id': None})
        db.query(Approval).filter(Approval.approver_id == uid).update({'approver_id': None})
        db.query(Contact).filter(Contact.created_by == uid).update({'created_by': None})
        db.query(SystemLog).filter(SystemLog.user_id == uid).update({'user_id': None})
        db.query(EmailLog).filter(EmailLog.user_id == uid).update({'user_id': None})
        try:
            db.query(Attachment).filter(Attachment.uploaded_by == uid).update({'uploaded_by': None})
        except: pass
        db.flush()
        # Delete user's reservations and their children
        for res in db.query(Reservation).filter(Reservation.user_id == uid).all():
            db.query(ChecklistItem).filter(ChecklistItem.reservation_id == res.id).delete()
            db.query(Approval).filter(Approval.reservation_id == res.id).delete()
            db.query(BookingContact).filter(BookingContact.booking_id == res.id).delete()
            db.query(Rating).filter(Rating.reservation_id == res.id).delete()
            db.delete(res)
        db.query(Rating).filter(Rating.user_id == uid).delete()
        db.flush()
        # Clear many-to-many
        db.execute(text('DELETE FROM user_locations WHERE user_id=:uid'), {'uid': uid})
        db.execute(text('DELETE FROM user_venues WHERE user_id=:uid'), {'uid': uid})
        db.query(LoginLog).filter(LoginLog.user_id == uid).update({'user_id': None})
        db.flush()
        db.delete(user)
        db.commit()
        _log(db, 'DELETE_USER', f'حذف المستخدم: {uname}')
        flash_msg(f'✅ تم حذف المستخدم: {uname}', 'success')
    except Exception as e:
        db.rollback()
        flash_msg(f'خطأ أثناء الحذف: {e}', 'danger')
    return redirect(url_for('users.index'))


# ── Change password ───────────────────────────────────────────────────────────
@users_bp.route('/<int:user_id>/change-password', methods=['POST'])
@login_required
@admin_required
def change_password(user_id):
    db   = get_db()
    user = db.query(User).get(user_id)
    if not user: abort(404)
    pw = request.form.get('password','').strip()
    if len(pw) < 6:
        flash_msg('كلمة المرور 6 أحرف على الأقل', 'danger')
        return redirect(url_for('users.index'))
    user.password_hash = _hash(pw); db.commit()
    flash_msg('✅ تم تغيير كلمة المرور', 'success')
    return redirect(url_for('users.index'))


# ── Import users from CSV — مطابق حرفياً لـ v54 import_users_from_csv ────────

# ── CSV Template Download ─────────────────────────────────────────────────────
@users_bp.route('/csv-template')
@login_required
@admin_required
def csv_template():
    from flask import session
    is_en = session.get('lang','ar') == 'en'
    rows = [
        ['username','password','full_name','email','phone','role','job_title','department'],
        ['bader01','Pass@123','بدر الكتبي','bader@example.com','0791234567','user','مدير مشاريع','تقنية المعلومات'],
        ['ahmad01','Pass@123','أحمد محمد','ahmad@example.com','0799876543','manager','مشرف','الموارد البشرية'],
    ] if not is_en else [
        ['username','password','full_name','email','phone','role','job_title','department'],
        ['bader01','Pass@123','Bader Al-Kutbi','bader@example.com','0791234567','user','Project Manager','IT'],
        ['ahmad01','Pass@123','Ahmad Mohammed','ahmad@example.com','0799876543','manager','Supervisor','HR'],
    ]
    import csv, io
    out = io.StringIO()
    csv.writer(out).writerows(rows)
    return Response(b'\xef\xbb\xbf' + out.getvalue().encode('utf-8'),
                    mimetype='text/csv',
                    headers={'Content-Disposition': 'attachment; filename=users_template.csv'})

@users_bp.route('/import-csv', methods=['GET','POST'])
@login_required
@admin_required
def import_csv():
    db = get_db()
    if request.method == 'POST':
        f = request.files.get('file')
        if not f or not f.filename.endswith('.csv'):
            flash_msg('يرجى رفع ملف CSV', 'danger')
            return redirect(url_for('users.import_csv'))

        content = f.stream.read().decode('utf-8-sig')
        reader  = csv.DictReader(io.StringIO(content))
        required_fields = ['username','email','full_name','password']

        if not reader.fieldnames or not all(fld in reader.fieldnames for fld in required_fields):
            flash_msg(f'الملف يجب أن يحتوي على الأعمدة: {", ".join(required_fields)}', 'danger')
            return redirect(url_for('users.import_csv'))

        default_role = db.query(Role).filter_by(name='مستخدم').first()
        if not default_role:
            flash_msg('لم يتم العثور على دور المستخدم الافتراضي', 'danger')
            return redirect(url_for('users.import_csv'))

        success_count = error_count = 0
        errors = []; new_users = []

        for row_num, row in enumerate(reader, start=2):
            username  = row.get('username','').strip()
            email     = row.get('email','').strip()
            full_name = row.get('full_name','').strip()
            password  = row.get('password','').strip()
            phone     = row.get('phone','').strip() or None
            is_active = row.get('is_active','1').strip().lower() in ['1','true','yes','نعم']
            is_verified = row.get('is_verified','0').strip().lower() in ['1','true','yes','نعم']

            if not username or not email or not full_name or not password:
                errors.append(f'سطر {row_num}: بيانات ناقصة')
                error_count += 1; continue

            existing = db.query(User).filter(
                (User.username == username) | (User.email == email)).first()
            if existing:
                errors.append(f'سطر {row_num}: المستخدم موجود مسبقاً')
                error_count += 1; continue

            try:
                job_title  = row.get('job_title','').strip() or None
                department = row.get('department','').strip() or None
                u = User(username=username, email=email, full_name=full_name,
                         password_hash=_hash(password), phone=phone,
                         job_title=job_title, department=department,
                         role_id=default_role.id,
                         is_active=is_active, is_verified=is_verified, language='ar')
                db.add(u); db.flush()
                new_users.append((u, password))
                success_count += 1
            except Exception as e:
                errors.append(f'سطر {row_num}: {e}')
                error_count += 1

        db.commit()

        # إرسال بريد ترحيب لكل مستخدم جديد — مثل v54
        for u, pw in new_users:
            try:
                from utils.email_helper import send_welcome
                send_welcome(u, pw, request.host_url)
            except: pass

        syslog('IMPORT_USERS', f'استيراد {success_count} مستخدم')
        msg = f'✅ تم استيراد {success_count} مستخدم بنجاح'
        if error_count:
            msg += f' — ⚠️ فشل: {error_count}'
        flash(msg, 'success' if not error_count else 'warning')
        if errors:
            for e in errors[:5]: flash(e,'danger')
        return redirect(url_for('users.index'))

    return render_template('admin/import_users.html')


# ── Export users CSV ──────────────────────────────────────────────────────────

def _make_xlsx(headers, rows, sheet_name, is_en):
    import io as _io
    from datetime import datetime as _dt
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = sheet_name[:31]
    ws.sheet_view.rightToLeft = not is_en
    h_align = 'left' if is_en else 'right'
    hfill = PatternFill('solid', fgColor='0C67EC'); hfont = Font(color='FFFFFF', bold=True)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h); c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal='center', vertical='center')
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci, val).alignment = Alignment(horizontal=h_align, vertical='center')
        if ri % 2 == 0:
            for ci in range(1, len(headers)+1):
                ws.cell(ri, ci).fill = PatternFill('solid', fgColor='EEF4F5')
    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, 40)
    ws.freeze_panes = 'A2'
    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

@users_bp.route('/export-csv')
@users_bp.route('/export-excel')
@login_required
@admin_required
def export_csv():
    from flask import session; from datetime import datetime as _dt
    from models.database import Reservation as _Res
    db = get_db(); is_en = session.get('lang','ar') == 'en'
    headers = (['Full Name','Username','Email','Phone','Department','Job Title','Role','Reservations','Status','Registration Date'] if is_en
               else ['الاسم الكامل','اسم المستخدم','البريد الإلكتروني','الهاتف','الدائرة','المسمى الوظيفي','الدور','عدد الحجوزات','الحالة','تاريخ التسجيل'])
    rows = []
    for u in db.query(User).order_by(User.full_name).all():
        res_count = db.query(_Res).filter_by(user_id=u.id).count()
        rows.append([u.full_name, u.username, u.email or '', u.phone or '',
                     u.department or '', u.job_title or '',
                     u.role.name if u.role else '',
                     res_count,
                     ('Active' if u.is_active else 'Inactive') if is_en else ('نشط' if u.is_active else 'غير نشط'),
                     u.created_at.strftime('%Y-%m-%d') if u.created_at else ''])
    try:
        data = _make_xlsx(headers, rows, 'Users' if is_en else 'المستخدمون', is_en)
        fname = f'users_{_dt.now().strftime("%Y%m%d")}.xlsx'
        return Response(data, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition': f'attachment; filename={fname}'})
    except ImportError:
        import csv as _csv; import io as _io
        out = _io.StringIO(); _csv.writer(out).writerows([headers] + rows)
        return Response(b'\xef\xbb\xbf' + out.getvalue().encode('utf-8'), mimetype='text/csv',
                        headers={'Content-Disposition': 'attachment; filename=users.csv'})


# ── Bulk message — مطابق لـ v54 send_bulk_message ────────────────────────────
@users_bp.route('/bulk-message', methods=['GET','POST'])
@login_required
@admin_required
def bulk_message():
    from models.database import ContactGroup
    db = get_db()
    if request.method == 'POST':
        subject     = request.form.get('subject','').strip()
        body_html   = request.form.get('body_html','').strip()
        body_plain  = request.form.get('body_plain','').strip()
        editor_mode = request.form.get('editor_mode','html')
        recipient   = request.form.get('recipient','all')
        user_ids    = request.form.getlist('user_ids', type=int)
        group_id    = request.form.get('group_id', type=int)

        # Use HTML body if available, else plain
        body = body_html if (editor_mode == 'html' and body_html) else body_plain

        if not subject or not body:
            flash_msg('عنوان ومحتوى الرسالة مطلوبان', 'danger')
            return redirect(url_for('users.bulk_message'))

        if recipient == 'all':
            targets = db.query(User).filter_by(is_active=True).all()
        elif recipient == 'group' and group_id:
            group = db.query(ContactGroup).get(group_id)
            targets = group.users if group else []
        else:
            targets = [db.query(User).get(uid) for uid in user_ids if uid]
            targets = [u for u in targets if u]

        # Handle attachments
        attach_files = request.files.getlist('attachments')
        attachments  = []
        for f in attach_files:
            if f and f.filename:
                data = f.read()
                attachments.append((f.filename, f.mimetype or 'application/octet-stream', data))

        try:
            from utils.email_helper import send_bulk
            from flask import session as _sess
            ui_lang = _sess.get('lang', 'ar')  # use admin's UI language
            send_bulk(targets, subject, body,
                      html=editor_mode=='html',
                      attachments=attachments,
                      interface_lang=ui_lang)
        except Exception as e:
            flash_msg(f'خطأ في الإرسال: {e}', 'danger')
            return redirect(url_for('users.bulk_message'))

        # Log each recipient for tracking
        recipients_log = ', '.join([u.email or u.username for u in targets[:10]])
        if len(targets) > 10:
            recipients_log += f' ... +{len(targets)-10}'
        syslog('BULK_MESSAGE', f'رسالة جماعية ({len(targets)} مستخدم) | الموضوع: {subject} | المستلمون: {recipients_log}')
        flash_msg(f'✅ تم إرسال الرسالة إلى {len(targets)} مستخدم', 'success')
        return redirect(url_for('users.index'))

    users  = db.query(User).filter_by(is_active=True).order_by(User.full_name).all()
    groups = db.query(ContactGroup).filter_by(is_active=True).order_by(ContactGroup.name).all()
    return render_template('admin/bulk_message.html', users=users, groups=groups)


# ── Roles ─────────────────────────────────────────────────────────────────────
@users_bp.route('/roles')
@login_required
@admin_required
def roles():
    db    = get_db()
    roles = db.query(Role).all()
    return render_template('admin/roles.html', roles=roles)


@users_bp.route('/roles/new', methods=['GET','POST'])
@login_required
@admin_required
def new_role():
    db = get_db()
    if request.method == 'POST':
        name    = request.form.get('name','').strip()
        name_en = request.form.get('name_en','').strip()
        desc    = request.form.get('description','').strip()
        if not name:
            flash_msg('اسم الدور مطلوب', 'danger')
            return render_template('admin/role_form.html', role=None, form=request.form)
        r = Role(name=name, name_en=name_en, description=desc)
        db.add(r); db.commit()
        _log(db,'ADD_ROLE',f'دور جديد: {name}')
        flash_msg(f'✅ تم إنشاء الدور: {name}', 'success')
        return redirect(url_for('users.roles'))
    return render_template('admin/role_form.html', role=None, form={})


@users_bp.route('/roles/<int:role_id>/edit', methods=['GET','POST'])
@login_required
@admin_required
def edit_role(role_id):
    db   = get_db()
    role = db.query(Role).get(role_id)
    if not role: abort(404)
    if request.method == 'POST':
        role.name    = request.form.get('name','').strip()
        role.name_en = request.form.get('name_en','').strip()
        role.description = request.form.get('description','').strip()
        db.commit()
        _log(db,'EDIT_ROLE',f'تعديل الدور: {role.name}')
        flash_msg('✅ تم تحديث الدور', 'success')
        return redirect(url_for('users.roles'))
    return render_template('admin/role_form.html', role=role, form={})


@users_bp.route('/roles/<int:role_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_role(role_id):
    db   = get_db()
    role = db.query(Role).get(role_id)
    if not role: abort(404)
    if role.users:
        flash_msg('لا يمكن حذف دور مرتبط بمستخدمين', 'danger')
        return redirect(url_for('users.roles'))
    name = role.name; db.delete(role); db.commit()
    _log(db,'DELETE_ROLE',f'حذف الدور: {name}')
    flash_msg(f'تم حذف الدور: {name}', 'success')
    return redirect(url_for('users.roles'))


# ── Permissions — مطابق لـ v54 PermissionsWindow ─────────────────────────────
@users_bp.route('/roles/<int:role_id>/permissions', methods=['GET','POST'])
@login_required
@admin_required
def role_permissions(role_id):
    db    = get_db()
    role  = db.query(Role).get(role_id)
    if not role: abort(404)
    perms = db.query(Permission).order_by(Permission.category, Permission.code).all()

    if request.method == 'POST':
        # حذف القديمة وإعادة بناء — مطابق لـ v54 PermissionsWindow.save
        db.query(RolePermission).filter_by(role_id=role_id).delete()
        for perm in perms:
            view_val    = request.form.get(f'view_{perm.id}')    == 'on'
            add_val     = request.form.get(f'add_{perm.id}')     == 'on'
            edit_val    = request.form.get(f'edit_{perm.id}')    == 'on'
            delete_val  = request.form.get(f'delete_{perm.id}')  == 'on'
            approve_val = request.form.get(f'approve_{perm.id}') == 'on'
            comment_val = request.form.get(f'comment_{perm.id}') == 'on'
            if any([view_val,add_val,edit_val,delete_val,approve_val,comment_val]):
                rp = RolePermission(role_id=role_id, permission_id=perm.id,
                                    can_view=view_val, can_add=add_val,
                                    can_edit=edit_val, can_delete=delete_val,
                                    can_approve=approve_val, can_comment=comment_val)
                db.add(rp)
        db.commit()
        _log(db,'EDIT_PERMISSIONS',f'تعديل صلاحيات دور: {role.name}')
        flash_msg('✅ تم حفظ الصلاحيات', 'success')
        return redirect(url_for('users.roles'))

    # بناء خريطة الصلاحيات الحالية
    current = {rp.permission_id: rp for rp in db.query(RolePermission).filter_by(role_id=role_id).all()}
    # تجميع حسب category
    categories = {}
    for p in perms:
        cat = p.category or 'عام'
        categories.setdefault(cat, []).append(p)

    return render_template('admin/role_permissions.html',
        role=role, modules=categories, current=current)


# ── Login log — مطابق لـ v54 SecurityLogWindow ───────────────────────────────
@users_bp.route('/login-log')
@login_required
@admin_required
def login_log():
    db   = get_db()
    page = request.args.get('page',1,type=int)
    q    = db.query(LoginLog).order_by(LoginLog.login_time.desc())
    user_f = request.args.get('user_id','')
    if user_f and user_f.isdigit():
        q = q.filter(LoginLog.user_id == int(user_f))
    items, total, total_pages = paginate(q, page, 30)
    users = db.query(User).order_by(User.full_name).all()
    return render_template('admin/login_log.html',
        logs=items, total=total, page=page, total_pages=total_pages,
        users=users, user_f=user_f)

# ── Export PDF ────────────────────────────────────────────────────────────────
@users_bp.route('/export/pdf')
@login_required
def export_pdf():
    from flask import Response
    import io
    db    = get_db()
    perms = get_permissions()
    if not perms.can('users_view'): abort(403)
    users = db.query(User).order_by(User.full_name).all()
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
                                rightMargin=1.5*cm, leftMargin=1.5*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        story  = [Paragraph('ARS — Users Report', styles['Heading1']), Spacer(1, 0.5*cm)]
        headers = ['Full Name','Username','Email','Role','Logins','Status']
        data = [headers] + [[
            u.full_name, u.username, u.email or '',
            u.role.name if u.role else '', str(u.login_count or 0),
            'Active' if u.is_active else 'Inactive'
        ] for u in users]
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#0C67EC')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTSIZE',(0,0),(-1,-1),9),
            ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#dee2e6')),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#f8f9fa')]),
            ('PADDING',(0,0),(-1,-1),5),
        ]))
        story.append(t)
        doc.build(story)
        buf.seek(0)
        return Response(buf.read(), mimetype='application/pdf',
                        headers={'Content-Disposition': 'attachment;filename=ARS_Users.pdf'})
    except ImportError:
        flash_msg('يرجى تثبيت reportlab', 'danger')
        return redirect(url_for('users.index'))

