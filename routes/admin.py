"""
Admin — لوحة التحكم + الصيانة + الإعدادات
مطابق لـ v54: show_dashboard + MaintenanceWindow + SettingsWindow
"""
import os, shutil, json
from datetime import datetime, timedelta
from utils.flash_helper import flash_msg
from flask import (Blueprint, render_template, request, redirect,
                   url_for, flash, abort, jsonify, send_file)
from flask_login import login_required, current_user
from sqlalchemy import func, text
from models.database import Reservation, User, Venue, Location, SystemLog, LoginLog
from utils.helpers import get_db, admin_required, syslog, paginate

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')

import os as _os
CONFIG_EMAIL = _os.environ.get('EMAIL_CONFIG_PATH',
    _os.path.join(_os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))), 'email_config.json'))
CONFIG_MAINT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'maintenance_config.json')


# ── Dashboard — مطابق لـ v54 show_dashboard ───────────────────────────────────
@admin_bp.route('/')
@admin_bp.route('/dashboard')
@login_required
@admin_required
def dashboard():
    db  = get_db()
    now = datetime.now()

    stats = {
        'total':      db.query(Reservation).count(),
        'pending':    db.query(Reservation).filter_by(status='pending').count(),
        'approved':   db.query(Reservation).filter_by(status='approved').count(),
        'rejected':   db.query(Reservation).filter_by(status='rejected').count(),
        'cancelled':  db.query(Reservation).filter_by(status='cancelled').count(),
        'completed':  db.query(Reservation).filter_by(status='completed').count(),
        'users':      db.query(User).filter_by(is_active=True).count(),
        'venues':     db.query(Venue).filter_by(is_active=True).count(),
        'locations':  db.query(Location).filter_by(is_active=True).count(),
        'this_month': db.query(Reservation).filter(
            Reservation.created_at >= now.replace(day=1)).count(),
    }

    # Trend شهري (12 شهر) — مثل v54 show_bar_chart
    monthly = []
    for i in range(11, -1, -1):
        d   = (now - timedelta(days=30*i)).replace(day=1,hour=0,minute=0,second=0)
        nd  = (d + timedelta(days=32)).replace(day=1)
        cnt = db.query(Reservation).filter(
            Reservation.created_at >= d, Reservation.created_at < nd).count()
        monthly.append({'month': d.strftime('%m/%Y'), 'count': cnt})

    # حجوزات معلقة — مثل v54 pending list
    pending_list = (db.query(Reservation).filter_by(status='pending')
                    .order_by(Reservation.created_at.asc()).limit(8).all())

    # أعلى 5 قاعات
    top_venues = (db.query(Venue.name, func.count(Reservation.id).label('cnt'))
                  .join(Reservation, Reservation.venue_id == Venue.id, isouter=True)
                  .group_by(Venue.id)
                  .order_by(func.count(Reservation.id).desc())
                  .limit(5).all())

    # XY — أعلى 8 مستخدمين نشاطاً (الإجمالي vs. الموافقات)
    from sqlalchemy import case
    _xy_rows = (db.query(
                    User.full_name,
                    func.count(Reservation.id).label('total'),
                    func.sum(case((Reservation.status == 'approved', 1), else_=0)).label('approved')
                )
                .join(Reservation, Reservation.user_id == User.id, isouter=True)
                .group_by(User.id)
                .order_by(func.count(Reservation.id).desc())
                .limit(8).all())
    top_users_xy = [[r[0], int(r[1] or 0), int(r[2] or 0)] for r in _xy_rows]

    return render_template('admin/dashboard.html',
        stats=stats, monthly=monthly,
        pending_list=pending_list, top_venues=top_venues,
        top_users_xy=top_users_xy)


# ── Dashboard PDF Export ───────────────────────────────────────────────────────
@admin_bp.route('/dashboard/pdf')
@login_required
@admin_required
def dashboard_pdf():
    import io
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from utils.pdf_helper import arabic_font, ar
    db  = get_db()
    now = datetime.now()
    stats = {
        'total':     db.query(Reservation).count(),
        'pending':   db.query(Reservation).filter_by(status='pending').count(),
        'approved':  db.query(Reservation).filter_by(status='approved').count(),
        'rejected':  db.query(Reservation).filter_by(status='rejected').count(),
        'cancelled': db.query(Reservation).filter_by(status='cancelled').count(),
        'completed': db.query(Reservation).filter_by(status='completed').count(),
        'users':     db.query(User).filter_by(is_active=True).count(),
        'venues':    db.query(Venue).filter_by(is_active=True).count(),
        'locations': db.query(Location).filter_by(is_active=True).count(),
        'this_month':db.query(Reservation).filter(Reservation.created_at >= now.replace(day=1)).count(),
    }
    monthly = []
    for i in range(11, -1, -1):
        d   = (now - timedelta(days=30*i)).replace(day=1,hour=0,minute=0,second=0)
        nd  = (d + timedelta(days=32)).replace(day=1)
        cnt = db.query(Reservation).filter(Reservation.created_at>=d, Reservation.created_at<nd).count()
        monthly.append((d.strftime('%m/%Y'), cnt))

    from flask import session
    is_en = session.get('lang','ar') == 'en'
    def _t(ar_text, en_text): return en_text if is_en else ar(ar_text)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    fn  = arabic_font()
    elements = []
    BLUE    = colors.HexColor('#1565C0')
    LBLUE   = colors.HexColor('#EBF5FB')
    LBLUE2  = colors.HexColor('#D6EAF8')
    LBLUE3  = colors.HexColor('#AED6F1')
    TEAL    = colors.HexColor('#00695C')
    AMBER   = colors.HexColor('#F57F17')

    title_style = ParagraphStyle('t', fontName=fn, fontSize=16, alignment=1, spaceAfter=12,
                                  textColor=BLUE)
    sub_style   = ParagraphStyle('s', fontName=fn, fontSize=11, alignment=1, spaceAfter=8,
                                  textColor=colors.grey)

    report_title = _t('تقرير لوحة التحكم', 'Dashboard Report')
    report_date  = _t(f'تاريخ التقرير: {now.strftime("%Y-%m-%d %H:%M")}',
                      f'Report Date: {now.strftime("%Y-%m-%d %H:%M")}')
    elements.append(Paragraph(report_title, title_style))
    elements.append(Paragraph(report_date,  sub_style))
    elements.append(Spacer(1, 0.4*cm))

    # Stats table — 3-column layout: label | value | label | value | label | value
    stat_data = [
        [_t('إجمالي الحجوزات','Total Bookings'), str(stats['total']),
         _t('معلقة','Pending'),                  str(stats['pending']),
         _t('موافق عليها','Approved'),            str(stats['approved'])],
        [_t('مرفوضة','Rejected'),                 str(stats['rejected']),
         _t('ملغاة','Cancelled'),                 str(stats['cancelled']),
         _t('هذا الشهر','This Month'),            str(stats['this_month'])],
        [_t('المستخدمون','Active Users'),         str(stats['users']),
         _t('القاعات','Active Venues'),           str(stats['venues']),
         _t('المواقع','Locations'),               str(stats['locations'])],
    ]
    LABEL_FILL  = colors.HexColor('#1565C0')  # label cells
    VALUE_FILL  = colors.HexColor('#E3F2FD')  # value cells
    st = Table(stat_data, colWidths=[4.5*cm, 2.5*cm, 4.5*cm, 2.5*cm, 4.5*cm, 2.5*cm])
    st.setStyle(TableStyle([
        ('FONTNAME',   (0,0),(-1,-1), fn),
        ('FONTSIZE',   (0,0),(-1,-1), 10),
        ('ALIGN',      (0,0),(-1,-1), 'CENTER'),
        ('VALIGN',     (0,0),(-1,-1), 'MIDDLE'),
        # Label columns (0,2,4) — dark blue with white text
        ('BACKGROUND', (0,0),(0,-1), LABEL_FILL),
        ('TEXTCOLOR',  (0,0),(0,-1), colors.white),
        ('FONTNAME',   (0,0),(0,-1), fn),
        ('BACKGROUND', (2,0),(2,-1), LABEL_FILL),
        ('TEXTCOLOR',  (2,0),(2,-1), colors.white),
        ('BACKGROUND', (4,0),(4,-1), LABEL_FILL),
        ('TEXTCOLOR',  (4,0),(4,-1), colors.white),
        # Value columns (1,3,5) — light blue
        ('BACKGROUND', (1,0),(1,-1), VALUE_FILL),
        ('BACKGROUND', (3,0),(3,-1), VALUE_FILL),
        ('BACKGROUND', (5,0),(5,-1), VALUE_FILL),
        ('FONTSIZE',   (1,0),(1,-1), 14),
        ('FONTSIZE',   (3,0),(3,-1), 14),
        ('FONTSIZE',   (5,0),(5,-1), 14),
        ('GRID',       (0,0),(-1,-1), 0.5, LBLUE3),
        ('ROWHEIGHT',  (0,0),(-1,-1), 30),
    ]))
    elements.append(st)
    elements.append(Spacer(1, 0.6*cm))

    # Monthly table
    elements.append(Paragraph(_t('الحجوزات — آخر 12 شهر','Bookings — Last 12 Months'), title_style))
    mon_label = _t('الشهر','Month')
    mon_count = _t('عدد الحجوزات','Count')
    mon_hdr = [[mon_label, mon_count]]
    mon_rows = [[m[0], str(m[1])] for m in monthly]
    mt = Table(mon_hdr + mon_rows, colWidths=[6*cm, 6*cm])
    mt.setStyle(TableStyle([
        ('FONTNAME', (0,0),(-1,-1), fn),
        ('FONTSIZE', (0,0),(-1,-1), 10),
        ('ALIGN',    (0,0),(-1,-1), 'CENTER'),
        ('BACKGROUND',(0,0),(-1,0), BLUE),
        ('TEXTCOLOR', (0,0),(-1,0), colors.white),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, LBLUE]),
        ('GRID',     (0,0),(-1,-1), 0.5, LBLUE3),
    ]))
    elements.append(mt)

    doc.build(elements)
    buf.seek(0)
    lang_suffix = 'en' if is_en else 'ar'
    fname = f'dashboard_{now.strftime("%Y%m%d_%H%M")}_{lang_suffix}.pdf'
    return send_file(buf, as_attachment=True, download_name=fname, mimetype='application/pdf')


# ── Dashboard Excel Export ─────────────────────────────────────────────────────
@admin_bp.route('/dashboard/excel')
@login_required
@admin_required
def dashboard_excel():
    import io
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return 'openpyxl not installed', 500

    from flask import session as fsession
    lang = fsession.get('lang', 'ar')
    is_en = (lang == 'en')

    db  = get_db()
    now = datetime.now()

    if is_en:
        stats = {
            'Total Reservations': db.query(Reservation).count(),
            'Pending':            db.query(Reservation).filter_by(status='pending').count(),
            'Approved':           db.query(Reservation).filter_by(status='approved').count(),
            'Rejected':           db.query(Reservation).filter_by(status='rejected').count(),
            'Cancelled':          db.query(Reservation).filter_by(status='cancelled').count(),
            'Completed':          db.query(Reservation).filter_by(status='completed').count(),
            'Active Users':       db.query(User).filter_by(is_active=True).count(),
            'Venues':             db.query(Venue).filter_by(is_active=True).count(),
            'Locations':          db.query(Location).filter_by(is_active=True).count(),
            'This Month':         db.query(Reservation).filter(Reservation.created_at >= now.replace(day=1)).count(),
        }
        sheet_title      = 'Dashboard Report'
        date_label       = f'Date: {now.strftime("%Y-%m-%d %H:%M")}'
        col_item         = 'Item'
        col_value        = 'Value'
        monthly_sheet    = 'Monthly Bookings'
        col_month        = 'Month'
        col_count        = 'Count'
        lang_suffix      = 'en'
    else:
        stats = {
            'إجمالي الحجوزات': db.query(Reservation).count(),
            'معلقة':            db.query(Reservation).filter_by(status='pending').count(),
            'موافق عليها':      db.query(Reservation).filter_by(status='approved').count(),
            'مرفوضة':           db.query(Reservation).filter_by(status='rejected').count(),
            'ملغاة':            db.query(Reservation).filter_by(status='cancelled').count(),
            'مكتملة':           db.query(Reservation).filter_by(status='completed').count(),
            'المستخدمون':       db.query(User).filter_by(is_active=True).count(),
            'القاعات':          db.query(Venue).filter_by(is_active=True).count(),
            'المواقع':          db.query(Location).filter_by(is_active=True).count(),
            'هذا الشهر':        db.query(Reservation).filter(Reservation.created_at >= now.replace(day=1)).count(),
        }
        sheet_title      = 'تقرير لوحة التحكم'
        date_label       = f'تاريخ: {now.strftime("%Y-%m-%d %H:%M")}'
        col_item         = 'البيان'
        col_value        = 'القيمة'
        monthly_sheet    = 'الحجوزات الشهرية'
        col_month        = 'الشهر'
        col_count        = 'العدد'
        lang_suffix      = 'ar'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Dashboard'
    ws.sheet_view.rightToLeft = not is_en

    hdr_fill  = PatternFill('solid', fgColor='0C67EC')
    hdr_font  = Font(bold=True, color='FFFFFF', name='Calibri', size=12)
    even_fill = PatternFill('solid', fgColor='EBF5FB')
    align_data = Alignment(horizontal='right' if not is_en else 'left')
    align_num  = Alignment(horizontal='center')

    ws['A1'] = sheet_title
    ws['A2'] = date_label
    ws['A1'].font = Font(bold=True, size=14, name='Calibri')
    row = 4
    ws.cell(row=row, column=1, value=col_item).font  = hdr_font
    ws.cell(row=row, column=1).fill = hdr_fill
    ws.cell(row=row, column=2, value=col_value).font = hdr_font
    ws.cell(row=row, column=2).fill = hdr_fill
    for i, (k, v) in enumerate(stats.items(), start=row+1):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
        if i % 2 == 0:
            ws.cell(row=i, column=1).fill = even_fill
            ws.cell(row=i, column=2).fill = even_fill
        ws.cell(row=i, column=1).alignment = align_data
        ws.cell(row=i, column=2).alignment = align_num

    ws2 = wb.create_sheet(monthly_sheet)
    ws2.sheet_view.rightToLeft = not is_en
    ws2.cell(1,1,col_month).font  = hdr_font; ws2.cell(1,1).fill = hdr_fill
    ws2.cell(1,2,col_count).font  = hdr_font; ws2.cell(1,2).fill = hdr_fill
    for i in range(11, -1, -1):
        d   = (now - timedelta(days=30*i)).replace(day=1,hour=0,minute=0,second=0)
        nd  = (d + timedelta(days=32)).replace(day=1)
        cnt = db.query(Reservation).filter(Reservation.created_at>=d, Reservation.created_at<nd).count()
        r = 12 - i + 2
        ws2.cell(row=r, column=1, value=d.strftime('%m/%Y'))
        ws2.cell(row=r, column=2, value=cnt)

    for ws_ in [ws, ws2]:
        for col in ws_.columns:
            ws_.column_dimensions[get_column_letter(col[0].column)].width = 26

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f'dashboard_{now.strftime("%Y%m%d_%H%M")}_{lang_suffix}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── System log — مطابق لـ v54 SecurityLogWindow ───────────────────────────────
@admin_bp.route('/system-log')
@login_required
@admin_required
def system_log():
    db     = get_db()
    page   = request.args.get('page',1,type=int)
    level  = request.args.get('level','')
    action = request.args.get('action','')
    q      = db.query(SystemLog)
    if level:  q = q.filter(SystemLog.level == level)
    if action: q = q.filter(SystemLog.action.ilike(f'%{action}%'))
    items, total, total_pages = paginate(q.order_by(SystemLog.created_at.desc()), page, 30)
    return render_template('admin/system_log.html',
        logs=items, total=total, page=page, total_pages=total_pages,
        level=level, action=action)


# ── Audit Log ──────────────────────────────────────────────────────────────────
@admin_bp.route('/audit-log')
@login_required
@admin_required
def audit_log():
    db   = get_db()
    page = request.args.get('page', 1, type=int)
    user_filter = request.args.get('user', '')
    action_filter = request.args.get('action', '')
    q = db.query(SystemLog)
    if user_filter:
        try:
            uid = int(user_filter)
            q = q.filter(SystemLog.user_id == uid)
        except ValueError:
            pass
    if action_filter:
        q = q.filter(SystemLog.action.ilike(f'%{action_filter}%'))
    items, total, total_pages = paginate(q.order_by(SystemLog.created_at.desc()), page, 50)
    users = db.query(User).order_by(User.full_name).all()
    return render_template('admin/audit_log.html',
        logs=items, total=total, page=page, total_pages=total_pages,
        users=users, user_filter=user_filter, action_filter=action_filter)


# ── Maintenance — مطابق لـ v54 MaintenanceWindow ─────────────────────────────
CONFIG_TICKER      = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'ticker_config.json')
CONFIG_AUTH_TICKER = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'auth_ticker_config.json')

def _load_ticker():
    try:
        with open(CONFIG_TICKER, encoding='utf-8') as f:
            return json.load(f)
    except:
        return {
            'feeds_ar': ['مرحباً بكم في نظام ARS لإدارة الحجوزات'],
            'feeds_en': ['Welcome to ARS Reservation System'],
            'fg': '#F2C99A', 'bg': '', 'font': 'Tahoma',
            'size': 15, 'speed': 35, 'opacity': 0
        }

def _load_auth_ticker():
    try:
        with open(CONFIG_AUTH_TICKER, encoding='utf-8') as f:
            return json.load(f)
    except:
        return {
            'feeds_ar': ['مرحباً بكم — سجّل دخولك للمتابعة'],
            'feeds_en': ['Welcome — Please sign in to continue'],
            'fg': '#ffffff', 'bg': 'transparent', 'font': 'Tajawal',
            'size': 14, 'speed': 35
        }

def _save_ticker(cfg):
    with open(CONFIG_TICKER, 'w', encoding='utf-8') as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

def _save_auth_ticker(cfg):
    with open(CONFIG_AUTH_TICKER, 'w', encoding='utf-8') as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

def _save_maintenance(cfg):
    with open(CONFIG_MAINT, 'w', encoding='utf-8') as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

@admin_bp.route('/maintenance', methods=['GET','POST'])
@login_required
@admin_required
def maintenance():
    # قراءة حالة الصيانة
    try:
        with open(CONFIG_MAINT) as f:
            mcfg = json.load(f)
    except:
        mcfg = {'system_suspended': False, 'registration_suspended': False}

    db = get_db()
    log_count = db.query(SystemLog).count()
    ticker = _load_ticker()

    if request.method == 'POST':
        action = request.form.get('ticker_action','')
        if action == 'add_ar':
            text = request.form.get('ticker_text_ar','').strip()
            if text:
                ticker.setdefault('feeds_ar', []).append(text)
                _save_ticker(ticker)
        elif action == 'add_en':
            text = request.form.get('ticker_text_en','').strip()
            if text:
                ticker.setdefault('feeds_en', []).append(text)
                _save_ticker(ticker)
        elif action == 'del_ar':
            idx = int(request.form.get('ticker_idx', 0))
            feeds = ticker.get('feeds_ar', [])
            if 0 <= idx < len(feeds): feeds.pop(idx); ticker['feeds_ar'] = feeds; _save_ticker(ticker)
        elif action == 'del_en':
            idx = int(request.form.get('ticker_idx', 0))
            feeds = ticker.get('feeds_en', [])
            if 0 <= idx < len(feeds): feeds.pop(idx); ticker['feeds_en'] = feeds; _save_ticker(ticker)
        elif action == 'save_appearance':
            ticker['fg']      = request.form.get('ticker_fg', '#F2C99A')
            ticker['bg']      = request.form.get('ticker_bg', '')
            ticker['font']    = request.form.get('ticker_font', 'Tajawal')
            ticker['size']    = int(request.form.get('ticker_size', 11))
            ticker['speed']   = int(request.form.get('ticker_speed', 35))
            ticker['opacity'] = int(request.form.get('ticker_opacity', 0) or 0)
            _save_ticker(ticker)
        elif action == 'save_logo':
            import base64
            logo_file = request.files.get('logo_file')
            if logo_file and logo_file.filename:
                data = base64.b64encode(logo_file.read()).decode('utf-8')
                mt   = logo_file.content_type or 'image/png'
                mcfg['logo_b64'] = f'data:{mt};base64,{data}'
                if request.form.get('remove_logo'):
                    mcfg.pop('logo_b64', None)
            elif request.form.get('remove_logo'):
                mcfg.pop('logo_b64', None)
            # Save header image position (left / center / right)
            mcfg['header_img_position'] = request.form.get('header_img_position', 'center')
            # Save optional separate header image for PDF
            hdr_file = request.files.get('header_img_file')
            if hdr_file and hdr_file.filename:
                import base64 as _b64
                hdata = _b64.b64encode(hdr_file.read()).decode('utf-8')
                hmt   = hdr_file.content_type or 'image/png'
                mcfg['header_img_b64'] = f'data:{hmt};base64,{hdata}'
            if request.form.get('remove_header_img'):
                mcfg.pop('header_img_b64', None)
            _save_maintenance(mcfg)
        elif action == 'save_report_header':
            mcfg['report_header_title']    = request.form.get('report_header_title','').strip()
            mcfg['report_header_subtitle'] = request.form.get('report_header_subtitle','').strip()
            mcfg['report_header_extra']    = request.form.get('report_header_extra','').strip()
            mcfg['report_header_footer']   = request.form.get('report_header_footer','').strip()
            _save_maintenance(mcfg)
        elif action == 'save_colors':
            mcfg['color_primary']       = request.form.get('color_primary', '#0C67EC')
            mcfg['color_primary_dark']  = request.form.get('color_primary_dark', '#0847B0')
            mcfg['color_primary_light'] = request.form.get('color_primary_light', '#3D8EF5')
            mcfg['color_accent']        = request.form.get('color_accent', '#5C9BDE')
            mcfg['color_bg']            = request.form.get('color_bg', '#eef6f7')
            _save_maintenance(mcfg)
        flash_msg('تم تحديث شريط الأخبار', 'success')
        return redirect(url_for('admin.maintenance'))

    return render_template('admin/maintenance.html',
        mcfg=mcfg, log_count=log_count,
        ticker=ticker, ticker_cfg=ticker,
        ticker_messages_ar=ticker.get('feeds_ar',[]),
        ticker_messages_en=ticker.get('feeds_en',[]))


# ── Ticker config API ─────────────────────────────────────────────────────────
@admin_bp.route('/ticker-config')
def ticker_config_api():
    """Public endpoint for base.html ticker JS"""
    from flask import jsonify
    ticker = _load_ticker()
    # Default opacity = 0
    if 'opacity' not in ticker:
        ticker['opacity'] = 0
    return jsonify(ticker)


# ── Live users + IP/Browser info ──────────────────────────────────────────────
@admin_bp.route('/live-users')
@login_required
@admin_required
def live_users():
    from flask import jsonify
    from datetime import datetime, timedelta
    from models.database import LoginLog
    db = get_db()
    # Consider users active if logged in within last 15 minutes
    cutoff = datetime.now() - timedelta(minutes=15)
    try:
        recent = db.query(LoginLog).filter(
            LoginLog.created_at >= cutoff,
            LoginLog.success == True
        ).order_by(LoginLog.created_at.desc()).all()
        users = []
        seen = set()
        for log in recent:
            if log.user_id not in seen:
                seen.add(log.user_id)
                users.append({
                    'username': log.user.username if log.user else '?',
                    'full_name': log.user.full_name if log.user else '?',
                    'ip': log.ip_address or '—',
                    'browser': (log.platform or log.hostname or 'Browser')[:40],
                    'time': log.created_at.strftime('%H:%M') if log.created_at else '—',
                })
        return jsonify({'users': users, 'count': len(users)})
    except Exception as e:
        return jsonify({'users': [], 'count': 0, 'error': str(e)})


# ── Backup — مطابق لـ v54 backup_database ────────────────────────────────────
@admin_bp.route('/maintenance/backup', methods=['POST'])
@login_required
@admin_required
def backup():
    db_path = 'acs_venues.db'
    if not os.path.exists(db_path):
        flash_msg('لا توجد قاعدة بيانات SQLite (PostgreSQL لا تدعم هذه العملية في Render)', 'warning')
        return redirect(url_for('admin.maintenance'))
    os.makedirs('backups', exist_ok=True)
    ts  = datetime.now().strftime('%Y%m%d_%H%M%S')
    dst = f'backups/acs_backup_{ts}.db'
    shutil.copy2(db_path, dst)
    syslog('BACKUP', f'نسخة احتياطية: {dst}')
    flash_msg(f'✅ تم إنشاء النسخة الاحتياطية: {dst}', 'success')
    return redirect(url_for('admin.maintenance'))


# ── Clean logs — مطابق لـ v54 clean_logs ─────────────────────────────────────
@admin_bp.route('/maintenance/clean-logs', methods=['POST'])
@login_required
@admin_required
def clean_logs():
    db      = get_db()
    cutoff  = datetime.now() - timedelta(days=30)
    deleted = db.query(SystemLog).filter(SystemLog.created_at < cutoff).delete()
    db.commit()
    syslog('CLEAN_LOGS', f'تم حذف {deleted} سجل قديم')
    flash_msg(f'✅ تم حذف {deleted} سجل قديم (أكثر من 30 يوم)', 'success')
    return redirect(url_for('admin.maintenance'))


# ── Optimize — مطابق لـ v54 optimize_database ────────────────────────────────
@admin_bp.route('/maintenance/optimize', methods=['POST'])
@login_required
@admin_required
def optimize():
    db = get_db()
    try:
        db.execute(text('VACUUM'))
        db.commit()
    except:
        pass  # PostgreSQL لا تدعم VACUUM عبر SQLAlchemy
    syslog('OPTIMIZE', 'تحسين قاعدة البيانات')
    flash_msg('✅ تم تحسين قاعدة البيانات', 'success')
    return redirect(url_for('admin.maintenance'))


# ── Suspend toggle — مطابق لـ v54 _toggle_sys / _toggle_reg ──────────────────
@admin_bp.route('/maintenance/toggle', methods=['POST'])
@login_required
@admin_required
def toggle_maintenance():
    try:
        with open(CONFIG_MAINT) as f:
            mcfg = json.load(f)
    except:
        mcfg = {'system_suspended': False, 'registration_suspended': False}

    toggle = request.form.get('toggle','')
    if toggle == 'system':
        mcfg['system_suspended'] = not mcfg.get('system_suspended', False)
    elif toggle == 'registration':
        mcfg['registration_suspended'] = not mcfg.get('registration_suspended', False)

    with open(CONFIG_MAINT, 'w') as f:
        json.dump(mcfg, f)

    syslog('TOGGLE_MAINTENANCE', f'{toggle}: {mcfg}')
    flash_msg('✅ تم تحديث إعدادات الصيانة', 'success')
    return redirect(url_for('admin.maintenance'))


# ── Settings (Email Config) — مطابق لـ v54 EmailConfigWindow ─────────────────
@admin_bp.route('/settings', methods=['GET','POST'])
@login_required
@admin_required
def settings():
    from utils.i18n import get_lang as _gl
    _is_en = _gl() == 'en'
    PROVIDERS = {
        'brevo_api':  {'label':'Brevo API',  'smtp':'',                        'port':0,
                       'help':'الأفضل على Render.com — لا يحتاج SMTP port. اذهب إلى brevo.com → Settings → API Keys → Create API Key' if not _is_en else 'Best for Render.com — no SMTP port needed. Go to brevo.com → Settings → API Keys → Create API Key'},
        'gmail':     {'label':'Gmail',                   'smtp':'smtp.gmail.com',          'port':587,
                      'help':'Use App Password from Google Account → Security → App passwords' if _is_en else 'استخدم App Password من Google Account → Security → App passwords'},
        'office365': {'label':'Microsoft 365 / Outlook', 'smtp':'smtp.office365.com',       'port':587,
                      'help':'Use Microsoft 365 email and password or App Password' if _is_en else 'استخدم بريد Microsoft 365 وكلمة المرور أو App Password'},
        'outlook':   {'label':'Outlook.com (Hotmail)',   'smtp':'smtp-mail.outlook.com',    'port':587,
                      'help':'Use your Outlook.com email and password' if _is_en else 'استخدم بريد Outlook.com وكلمة المرور'},
        'yahoo':     {'label':'Yahoo Mail',              'smtp':'smtp.mail.yahoo.com',      'port':587,
                      'help':'Enable "Allow apps that use less secure sign in" or use App password' if _is_en else 'فعّل "Allow apps that use less secure sign in" أو استخدم App password'},
        'brevo':     {'label':'Brevo SMTP',               'smtp':'smtp-relay.brevo.com',     'port':587,
                      'help':'SMTP Key from brevo.com → Transactional → SMTP & API' if _is_en else 'SMTP Key من brevo.com → Transactional → SMTP & API'},
        'custom':    {'label':'Custom Server' if _is_en else 'خادم مخصص', 'smtp':'', 'port':587,
                      'help':'Enter your server details manually' if _is_en else 'أدخل بيانات خادمك يدوياً'},
    }

    try:
        with open(CONFIG_EMAIL, encoding='utf-8') as f:
            cfg = json.load(f)
    except:
        cfg = {'smtp_server':'smtp.gmail.com','smtp_port':587,
               'sender_email':'','sender_password':'',
               'sender_name':'ARS Applied Reservation System',
               'use_tls':True,'provider_key':'brevo_api',
               'brevo_api_key':'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'}

    if request.method == 'POST':
        action = request.form.get('action','save')

        if action == 'test':
            from utils.email_helper import test_smtp
            brevo_api_key = request.form.get('brevo_api_key', '').strip()
            provider = request.form.get('provider_key', '')
            smtp_port_raw = request.form.get('smtp_port', '') or '587'
            try:
                smtp_port_val = int(smtp_port_raw)
            except (ValueError, TypeError):
                smtp_port_val = 587
            ok, msg = test_smtp(
                request.form.get('smtp_server', ''),
                smtp_port_val,
                request.form.get('sender_email', ''),
                request.form.get('sender_password', ''),
                request.form.get('use_tls') == 'on',
                brevo_api_key=brevo_api_key if provider == 'brevo_api' else None,
            )
            flash_msg(('✅ ' if ok else '❌ ') + msg, 'success' if ok else 'danger')
            return redirect(url_for('admin.settings'))

        # Save config
        new_cfg = {
            'provider_key':   request.form.get('provider_key','gmail'),
            'smtp_server':    request.form.get('smtp_server',''),
            'smtp_port':      int(request.form.get('smtp_port', 587) or 587),
            'sender_email':   request.form.get('sender_email',''),
            'sender_password':request.form.get('sender_password',''),
            'sender_name':    request.form.get('sender_name','ARS Applied Reservation System'),
            'use_tls':        request.form.get('use_tls') == 'on',
            'brevo_api_key':  request.form.get('brevo_api_key','').strip(),
        }
        with open(CONFIG_EMAIL,'w',encoding='utf-8') as f:
            json.dump(new_cfg, f, ensure_ascii=False, indent=2)
        syslog('SAVE_EMAIL_CONFIG', f"provider: {new_cfg['provider_key']}")
        flash_msg('✅ تم حفظ إعدادات البريد الإلكتروني', 'success')
        return redirect(url_for('admin.settings'))

    return render_template('admin/settings.html',
        cfg=cfg, providers=PROVIDERS)


# ── Ticker API — serves full ticker config to frontend ────────────────────────
@admin_bp.route('/api/ticker')
def ticker_api():
    from flask import jsonify
    ticker = _load_ticker()
    lang  = request.args.get('lang','ar')
    if lang == 'ar':
        msgs = ticker.get('feeds_ar', ['مرحباً بكم في نظام ARS']) or ['مرحباً بكم في نظام ARS']
    else:
        msgs = ticker.get('feeds_en', []) or ['Welcome to ARS Reservation Management System']
    return jsonify({
        'messages': msgs,
        'text':     ' ◆ '.join(msgs),
        'bg':       ticker.get('bg', ''),
        'fg':       ticker.get('fg', '#F2C99A'),
        'font':     ticker.get('font', 'Tahoma'),
        'size':     ticker.get('size', 15),
        'speed':    ticker.get('speed', 35),
        'opacity':  ticker.get('opacity', 0),
    })


@admin_bp.route('/api/auth-ticker')
def auth_ticker_api():
    """Public API - no login required - for auth pages"""
    from flask import jsonify
    ticker = _load_auth_ticker()
    lang   = request.args.get('lang', 'ar')
    if lang == 'ar':
        msgs = ticker.get('feeds_ar', ['مرحباً بكم — سجّل دخولك للمتابعة']) or ['مرحباً بكم']
    else:
        msgs = ticker.get('feeds_en', ['Welcome — Please sign in']) or ['Welcome']
    return jsonify({
        'messages': msgs,
        'text':     ' ◆ '.join(msgs),
        'fg':       ticker.get('fg', '#ffffff'),
        'bg':       ticker.get('bg', 'transparent'),
        'font':     ticker.get('font', 'Tajawal'),
        'size':     ticker.get('size', 14),
        'speed':    ticker.get('speed', 35),
    })


@admin_bp.route('/auth-ticker', methods=['GET', 'POST'])
@login_required
def auth_ticker():
    perms = get_permissions()
    if not perms.is_admin_or_manager(): abort(403)
    ticker = _load_auth_ticker()
    lang   = session.get('lang', 'ar')

    if request.method == 'POST':
        action = request.form.get('action', '')
        if action == 'add_ar':
            txt = request.form.get('text_ar', '').strip()
            if txt:
                ticker.setdefault('feeds_ar', []).append(txt)
        elif action == 'add_en':
            txt = request.form.get('text_en', '').strip()
            if txt:
                ticker.setdefault('feeds_en', []).append(txt)
        elif action == 'del_ar':
            idx = int(request.form.get('idx', 0))
            feeds = ticker.get('feeds_ar', [])
            if 0 <= idx < len(feeds): feeds.pop(idx)
            ticker['feeds_ar'] = feeds
        elif action == 'del_en':
            idx = int(request.form.get('idx', 0))
            feeds = ticker.get('feeds_en', [])
            if 0 <= idx < len(feeds): feeds.pop(idx)
            ticker['feeds_en'] = feeds
        elif action == 'style':
            ticker['fg']    = request.form.get('fg', '#ffffff')
            ticker['font']  = request.form.get('font', 'Tajawal')
            ticker['size']  = int(request.form.get('size', 14))
            ticker['speed'] = int(request.form.get('speed', 35))
        _save_auth_ticker(ticker)
        from utils.flash_helper import flash_msg
        flash_msg('✅ تم الحفظ', 'success')
        return redirect(url_for('admin.auth_ticker'))

    return render_template('admin/auth_ticker.html', ticker=ticker)

