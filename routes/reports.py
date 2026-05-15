"""
التقارير والمقارنة — مطابق حرفياً لـ v54 ReportsWindow + ComparisonWindow
"""
import io, csv
from datetime import datetime, timedelta, date
from utils.flash_helper import flash_msg
from flask import Blueprint, render_template, request, Response, abort, session
from flask_login import login_required, current_user
from sqlalchemy import func
from models.database import (Reservation, User, Venue, Location,
                              BlockedPeriod, Rating)
from utils.helpers import get_db, get_permissions

reports_bp = Blueprint('reports', __name__, url_prefix='/reports')


def _date_filter(q, f, start=None, end_d=None):
    """مطابق لـ v54 ReportsWindow.get_date_filter"""
    today = date.today()
    if f == 'today':
        return q.filter(func.date(Reservation.start_time) == today)
    elif f == 'week':
        ws = today - timedelta(days=today.weekday())
        return q.filter(Reservation.start_time >= datetime.combine(ws, datetime.min.time()))
    elif f == 'month':
        ms = today.replace(day=1)
        return q.filter(Reservation.start_time >= datetime.combine(ms, datetime.min.time()))
    elif f == 'custom' and start and end_d:
        try:
            s = datetime.strptime(start, '%Y-%m-%d')
            e = datetime.strptime(end_d, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            return q.filter(Reservation.start_time >= s, Reservation.start_time <= e)
        except: pass
    return q


# ── Main reports ──────────────────────────────────────────────────────────────
@reports_bp.route('/')
@login_required
def index():
    db    = get_db()
    perms = get_permissions()
    if not perms.can('reports_view'): abort(403)

    rtype = request.args.get('type', 'my_reservations')
    f     = request.args.get('filter', 'all')
    start = request.args.get('start', '')
    end_d = request.args.get('end', '')
    page  = request.args.get('page', 1, type=int)
    per_p = 25
    lang  = session.get('lang', 'ar')
    def _t(ar, en): return en if lang == 'en' else ar

    rows=[]; columns=[]; stats={}; total_count=0
    row_ids=[]; row_url_func=''; row_url_param=''

    if rtype == 'my_reservations':
        q = db.query(Reservation).filter_by(user_id=current_user.id)
        q = _date_filter(q, f, start, end_d)
        stats = {
            'total':     q.count(),
            'approved':  q.filter_by(status='approved').count(),
            'pending':   q.filter_by(status='pending').count(),
            'rejected':  q.filter_by(status='rejected').count(),
            'cancelled': q.filter_by(status='cancelled').count(),
        }
        total_count = q.count()
        items = q.order_by(Reservation.created_at.desc()).offset((page-1)*per_p).limit(per_p).all()
        columns = [_t('رقم الحجز','Booking #'), _t('العنوان','Title'),
                   _t('القاعة','Venue'), _t('تاريخ البدء','Start Date'),
                   _t('الانتهاء','End'), _t('الحالة','Status')]
        rows = [[r.booking_number, r.title,
                 r.venue.name if r.venue else '—',
                 r.start_time.strftime('%Y-%m-%d %H:%M') if r.start_time else '',
                 r.end_time.strftime('%H:%M') if r.end_time else '',
                 r.status] for r in items]
        row_ids = [r.id for r in items]
        row_url_func = 'reservations.detail'; row_url_param = 'res_id'

    elif rtype == 'all_reservations' and perms.is_admin_or_manager():
        q = db.query(Reservation)
        q = _date_filter(q, f, start, end_d)
        stats = {
            'total':     q.count(),
            'approved':  q.filter_by(status='approved').count(),
            'pending':   q.filter_by(status='pending').count(),
            'rejected':  q.filter_by(status='rejected').count(),
            'cancelled': q.filter_by(status='cancelled').count(),
        }
        total_count = q.count()
        items = q.order_by(Reservation.created_at.desc()).offset((page-1)*per_p).limit(per_p).all()
        columns = [_t('رقم الحجز','Booking #'), _t('العنوان','Title'),
                   _t('المستخدم','User'), _t('القاعة','Venue'),
                   _t('تاريخ البدء','Start Date'), _t('الحالة','Status')]
        rows = [[r.booking_number, r.title,
                 r.user.full_name if r.user else '—',
                 r.venue.name if r.venue else '—',
                 r.start_time.strftime('%Y-%m-%d %H:%M') if r.start_time else '',
                 r.status] for r in items]
        row_ids = [r.id for r in items]
        row_url_func = 'reservations.detail'; row_url_param = 'res_id'

    elif rtype == 'venues':
        venues = db.query(Venue).order_by(Venue.name).all()
        stats  = {'total_venues': len(venues),
                  'active_venues': sum(1 for v in venues if v.is_active)}
        total_count = len(venues)
        columns = [_t('القاعة','Venue'), _t('الموقع','Location'),
                   _t('إجمالي الحجوزات','Total Bookings'),
                   _t('حجوزاتي','My Bookings'),
                   _t('فترات محظورة','Blocked Periods'),
                   _t('الحالة','Status')]
        rows = [[
            v.name, v.location.name if v.location else '—',
            db.query(Reservation).filter_by(venue_id=v.id).count(),
            db.query(Reservation).filter_by(venue_id=v.id, user_id=current_user.id).count(),
            db.query(BlockedPeriod).filter_by(venue_id=v.id).count(),
            _t('نشط','Active') if v.is_active else _t('موقوف','Inactive'),
        ] for v in venues]
        row_ids = [v.id for v in venues]
        row_url_func = 'venues.detail'; row_url_param = 'venue_id'

    elif rtype == 'locations':
        locs = db.query(Location).order_by(Location.name).all()
        stats = {'total_locations': len(locs),
                 'active_locations': sum(1 for l in locs if l.is_active)}
        total_count = len(locs)
        columns = [_t('الموقع','Location'), _t('المدينة','City'),
                   _t('المنطقة','Area'), _t('القاعات','Venues'),
                   _t('إجمالي الحجوزات','Total Bookings'), _t('الحالة','Status')]
        rows = [[
            l.name, l.city or '—', l.area or '—',
            len(l.venues),
            sum(db.query(Reservation).filter_by(venue_id=v.id).count() for v in l.venues),
            _t('نشط','Active') if l.is_active else _t('موقوف','Inactive'),
        ] for l in locs]
        row_ids = [l.id for l in locs]
        row_url_func = 'locations.index'; row_url_param = ''

    elif rtype == 'users' and perms.can('users_view'):
        users = db.query(User).order_by(User.full_name).all()
        stats = {'total_users': len(users),
                 'active_users':    sum(1 for u in users if u.is_active),
                 'verified_users':  sum(1 for u in users if u.is_verified)}
        total_count = len(users)
        columns = [_t('الاسم الكامل','Full Name'), _t('اسم المستخدم','Username'),
                   _t('البريد','Email'), _t('الدور','Role'),
                   _t('الحجوزات','Bookings'), _t('الحالة','Status')]
        rows = [[
            u.full_name, u.username, u.email or '—',
            u.role.name if u.role else '—',
            db.query(Reservation).filter_by(user_id=u.id).count(),
            '✓' if u.is_active else '✗',
        ] for u in users]
        row_ids = [u.id for u in users]
        row_url_func = 'users.edit'; row_url_param = 'user_id'

    elif rtype == 'email_stats' and perms.is_admin_or_manager():
        try:
            from models.database import EmailLog
            from sqlalchemy import func as _func2
            # Detailed email stats
            email_detail = db.query(EmailLog).order_by(EmailLog.sent_at.desc()).limit(50).all()
            stats = {}
            total_count = db.query(EmailLog).count()
            columns = [_t('المستلم','Recipient'), _t('الموضوع','Subject'),
                       _t('النوع','Type'), _t('الحالة','Status'),
                       _t('التاريخ','Date')]
            rows = [[
                log.recipient or '—',
                (log.subject or '')[:50],
                log.type or '—',
                log.status or '—',
                log.sent_at.strftime('%Y-%m-%d %H:%M') if log.sent_at else '—'
            ] for log in email_detail]
        except Exception:
            stats = {}; rows = []; columns = []; total_count = 0

    total_pages = max(1, (total_count + per_p - 1) // per_p)

    now = datetime.now()
    monthly = []
    for i in range(11, -1, -1):
        d   = (now - timedelta(days=30*i)).replace(day=1, hour=0, minute=0, second=0)
        nd  = (d + timedelta(days=32)).replace(day=1)
        cnt = db.query(Reservation).filter(
            Reservation.created_at >= d, Reservation.created_at < nd).count()
        monthly.append({'month': d.strftime('%m/%Y'), 'count': cnt})

    top_venues = (db.query(Venue.name, func.count(Reservation.id).label('cnt'))
                  .join(Reservation, Reservation.venue_id == Venue.id, isouter=True)
                  .group_by(Venue.id)
                  .order_by(func.count(Reservation.id).desc())
                  .limit(5).all())

    # Email stats for admin
    email_stats = {}
    if perms.is_admin_or_manager():
        try:
            from models.database import EmailLog
            from sqlalchemy import func as _func
            email_stats = {
                'total':   db.query(EmailLog).count(),
                'sent':    db.query(EmailLog).filter(EmailLog.status=='sent').count(),
                'failed':  db.query(EmailLog).filter(EmailLog.status=='failed').count(),
                'bulk':    db.query(EmailLog).filter(EmailLog.type=='bulk').count(),
                'notif':   db.query(EmailLog).filter(EmailLog.type=='notification').count(),
            }
        except Exception:
            email_stats = {}

    return render_template('reports/index.html',
        rows=rows, columns=columns, stats=stats,
        rtype=rtype, filter=f, start=start, end=end_d,
        page=page, total_pages=total_pages, total_count=total_count,
        monthly=monthly, top_venues=top_venues, perms=perms,
        row_ids=row_ids, row_url_func=row_url_func, row_url_param=row_url_param,
        email_stats=email_stats)



def _make_xlsx(headers, rows, sheet_name, is_en):
    import io as _io
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = sheet_name[:31]
    ws.sheet_view.rightToLeft = not is_en
    h_align = 'left' if is_en else 'right'
    hfill = PatternFill('solid', fgColor='0C67EC'); hfont = Font(color='FFFFFF', bold=True)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h); c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal='center', vertical='center')
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci, val).alignment = Alignment(horizontal=h_align, vertical='center')
        if ri % 2 == 0:
            for ci in range(1, len(headers)+1):
                ws.cell(ri, ci).fill = PatternFill('solid', fgColor='EEF4F5')
    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, 40)
    ws.freeze_panes = 'A2'
    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

# ── export CSV/Excel ─────────────────────────────────────────────────────────
@reports_bp.route('/export-csv')
@reports_bp.route('/export-excel')
@login_required
def export_csv():
    from flask import session; from datetime import datetime as _dt
    db = get_db(); perms = get_permissions()
    if not perms.can('reports_export'): abort(403)
    is_en  = session.get('lang','ar') == 'en'
    rtype  = request.args.get('type','my_reservations')
    f      = request.args.get('filter','all')
    start  = request.args.get('start','')
    end_d  = request.args.get('end','')

    STATUS_EN = {'pending':'Pending','approved':'Approved','rejected':'Rejected','cancelled':'Cancelled','completed':'Completed'}
    STATUS_AR = {'pending':'معلق','approved':'موافق','rejected':'مرفوض','cancelled':'ملغي','completed':'مكتمل'}
    sl = STATUS_EN if is_en else STATUS_AR
    def _t(ar, en): return en if is_en else ar

    headers = []; rows = []; sheet = 'Report'

    if rtype == 'my_reservations':
        sheet = 'My Reservations' if is_en else 'حجوزاتي'
        headers = [_t('رقم الحجز','Booking #'), _t('العنوان','Title'), _t('القاعة','Venue'),
                   _t('تاريخ البدء','Start'), _t('الانتهاء','End'), _t('الحالة','Status')]
        q = _date_filter(db.query(Reservation).filter_by(user_id=current_user.id), f, start, end_d)
        rows = [[r.booking_number, r.title, r.venue.name if r.venue else '',
                 r.start_time.strftime('%Y-%m-%d %H:%M') if r.start_time else '',
                 r.end_time.strftime('%H:%M') if r.end_time else '',
                 sl.get(r.status, r.status)] for r in q.all()]

    elif rtype == 'all_reservations' and perms.is_admin_or_manager():
        sheet = 'All Reservations' if is_en else 'جميع الحجوزات'
        headers = [_t('رقم الحجز','Booking #'), _t('العنوان','Title'), _t('المستخدم','User'),
                   _t('القاعة','Venue'), _t('تاريخ البدء','Start'), _t('الحالة','Status')]
        q = _date_filter(db.query(Reservation), f, start, end_d)
        rows = [[r.booking_number, r.title, r.user.full_name if r.user else '',
                 r.venue.name if r.venue else '',
                 r.start_time.strftime('%Y-%m-%d %H:%M') if r.start_time else '',
                 sl.get(r.status, r.status)] for r in q.all()]

    elif rtype == 'venues':
        sheet = 'Venues' if is_en else 'القاعات'
        headers = [_t('القاعة','Venue'), _t('الموقع','Location'),
                   _t('إجمالي الحجوزات','Total Bookings'), _t('فترات محظورة','Blocked Periods'), _t('الحالة','Status')]
        for v in db.query(Venue).order_by(Venue.name).all():
            rows.append([v.name, v.location.name if v.location else '',
                         db.query(Reservation).filter_by(venue_id=v.id).count(),
                         db.query(BlockedPeriod).filter_by(venue_id=v.id).count(),
                         _t('نشط','Active') if v.is_active else _t('موقوف','Inactive')])

    elif rtype == 'locations':
        sheet = 'Locations' if is_en else 'المواقع'
        headers = [_t('الموقع','Location'), _t('المدينة','City'), _t('المنطقة','Area'),
                   _t('القاعات','Venues'), _t('إجمالي الحجوزات','Total Bookings'), _t('الحالة','Status')]
        for l in db.query(Location).order_by(Location.name).all():
            total = sum(db.query(Reservation).filter_by(venue_id=v.id).count() for v in l.venues)
            rows.append([l.name, l.city or '', l.area or '', len(l.venues), total,
                         _t('نشط','Active') if l.is_active else _t('موقوف','Inactive')])

    elif rtype == 'users' and perms.can('users_view'):
        sheet = 'Users' if is_en else 'المستخدمون'
        headers = [_t('الاسم الكامل','Full Name'), _t('اسم المستخدم','Username'),
                   _t('البريد','Email'), _t('الدور','Role'),
                   _t('الحجوزات','Bookings'), _t('الحالة','Status')]
        for u in db.query(User).order_by(User.full_name).all():
            rows.append([u.full_name, u.username, u.email or '',
                         u.role.name if u.role else '',
                         db.query(Reservation).filter_by(user_id=u.id).count(),
                         _t('✓' if u.is_active else '✗','✓' if u.is_active else '✗')])

    import csv as _csv; import io as _io
    out = _io.StringIO()
    _csv.writer(out).writerows([headers] + rows)
    fname = f'{rtype}_{_dt.now().strftime("%Y%m%d")}.csv'
    return Response(b'\xef\xbb\xbf' + out.getvalue().encode('utf-8'), mimetype='text/csv',
                    headers={'Content-Disposition': f'attachment; filename={fname}'})


# ── Export PDF ────────────────────────────────────────────────────────────────
# ── comparison — مطابق حرفياً لـ v54 ComparisonWindow ───────────────────────
@reports_bp.route('/comparison')
@login_required
def comparison():
    db    = get_db()
    perms = get_permissions()
    if not perms.is_admin_or_manager(): abort(403)

    _lang = session.get('lang', 'ar')
    def _t(ar, en): return en if _lang == 'en' else ar
    mode  = request.args.get('mode', 'venue')

    # Multi-item: support up to 5 items (a, b, c, d, e)
    items_keys = ['a','b','c','d','e']
    items_vals = [request.args.get(k, '').strip() for k in items_keys]
    items_vals = [v for v in items_vals if v]  # remove empty

    # Legacy support
    a  = request.args.get('a', '')
    b  = request.args.get('b', '')
    fa = request.args.get('fa', '')
    ta = request.args.get('ta', '')
    fb = request.args.get('fb', '')
    tb = request.args.get('tb', '')

    venues    = db.query(Venue).filter_by(is_active=True).order_by(Venue.name).all()
    locations = db.query(Location).filter_by(is_active=True).order_by(Location.name).all()
    users_all = db.query(User).filter_by(is_active=True).order_by(User.full_name).all()

    result = {}

    def _venue_stats(name):
        v = db.query(Venue).filter_by(name=name).first()
        if not v: return {}
        q = db.query(Reservation).filter_by(venue_id=v.id)
        return {
            _t('إجمالي الحجوزات','Total Bookings'): q.count(),
            _t('موافقة','Approved'):   q.filter_by(status='approved').count(),
            _t('معلقة','Pending'):    q.filter_by(status='pending').count(),
            _t('مرفوضة','Rejected'):  q.filter_by(status='rejected').count(),
            _t('ملغاة','Cancelled'):    q.filter_by(status='cancelled').count(),
            _t('مكتملة','Completed'):   q.filter_by(status='completed').count(),
            _t('فترات محظورة','Blocked Periods'): db.query(BlockedPeriod).filter_by(venue_id=v.id).count(),
        }

    def _location_stats(name):
        loc = db.query(Location).filter_by(name=name).first()
        if not loc: return {}
        venue_ids = [v.id for v in loc.venues]
        q = db.query(Reservation).filter(Reservation.venue_id.in_(venue_ids)) if venue_ids else db.query(Reservation).filter_by(id=-1)
        return {
            _t('عدد القاعات','Venues'): len(loc.venues),
            _t('إجمالي الحجوزات','Total Bookings'): q.count(),
            _t('موافقة','Approved'):   q.filter_by(status='approved').count(),
            _t('معلقة','Pending'):    q.filter_by(status='pending').count(),
            _t('مرفوضة','Rejected'):  q.filter_by(status='rejected').count(),
            _t('ملغاة','Cancelled'):    q.filter_by(status='cancelled').count(),
        }

    def _user_stats(uname):
        u = db.query(User).filter_by(username=uname).first()
        if not u: return {}
        q  = db.query(Reservation).filter_by(user_id=u.id)
        total     = q.count()
        approved  = q.filter_by(status='approved').count()
        pending   = q.filter_by(status='pending').count()
        rejected  = q.filter_by(status='rejected').count()
        cancelled = q.filter_by(status='cancelled').count()
        completed = q.filter_by(status='completed').count()
        rate = round(approved / total * 100) if total else 0
        return {
            _t('إجمالي الحجوزات','Total Bookings'): total,
            _t('موافقة','Approved'):   approved,
            _t('معلقة','Pending'):    pending,
            _t('مرفوضة','Rejected'):  rejected,
            _t('ملغاة','Cancelled'):    cancelled,
            _t('مكتملة','Completed'):   completed,
            _t('نسبة الموافقة %','Approval Rate %'): rate,
        }

    def _period_stats(d_from, d_to):
        try:
            s = datetime.strptime(d_from, '%Y-%m-%d')
            e = datetime.strptime(d_to, '%Y-%m-%d').replace(hour=23, minute=59)
            q = db.query(Reservation).filter(
                Reservation.start_time >= s, Reservation.start_time <= e)
            return {
                _t('إجمالي','Total'): q.count(),
                _t('موافقة','Approved'): q.filter_by(status='approved').count(),
                _t('معلقة','Pending'):  q.filter_by(status='pending').count(),
                _t('مرفوضة','Rejected'): q.filter_by(status='rejected').count(),
                _t('ملغاة','Cancelled'):  q.filter_by(status='cancelled').count(),
                _t('مكتملة','Completed'): q.filter_by(status='completed').count(),
            }
        except: return {}

    # ── Get stat function per mode ────────────────────────────────────────────
    def _get_stats(mode, val):
        if mode in ('venue','rating_venue'): return _venue_stats(val)
        if mode in ('location','rating_location'): return _location_stats(val)
        if mode in ('user','rating_user'): return _user_stats(val)
        return {}

    def _get_label(mode, val):
        if mode in ('user','rating_user'):
            u = db.query(User).filter_by(username=val).first()
            return u.full_name if u else val
        return val

    # ── Multi-item comparison ─────────────────────────────────────────────────
    multi_modes = ('venue','location','user','rating_venue','rating_location','rating_user')

    if mode in multi_modes and len(items_vals) >= 2:
        datasets = []
        for val in items_vals:
            datasets.append({
                'label': _get_label(mode, val),
                'key': val,
                'data': _get_stats(mode, val)
            })
        # Collect all metrics
        all_metrics = []
        for ds in datasets:
            for m in ds['data']:
                if m not in all_metrics:
                    all_metrics.append(m)
        result = {
            'multi': True,
            'datasets': datasets,
            'metrics': all_metrics,
            'title': _t(
                {'venue':'مقارنة القاعات','location':'مقارنة المواقع','user':'مقارنة المستخدمين',
                 'rating_venue':'تقييم القاعات','rating_location':'تقييم المواقع','rating_user':'تقييم المستخدمين'}.get(mode,'مقارنة'),
                {'venue':'Venues Comparison','location':'Locations Comparison','user':'Users Comparison',
                 'rating_venue':'Venue Ratings','rating_location':'Location Ratings','rating_user':'User Ratings'}.get(mode,'Comparison')
            )
        }

    elif mode == 'period' and fa and ta and fb and tb:
        result = {'label_a': f'{fa} → {ta}', 'label_b': f'{fb} → {tb}',
                  'title': _t('مقارنة الفترات','Periods Comparison'),
                  'data_a': _period_stats(fa, ta), 'data_b': _period_stats(fb, tb)}

    elif mode == 'period_user' and fa and ta and items_vals:
        user_labels = []
        user_counts = []
        user_details = []
        for uname in [v for v in items_vals if v]:
            u = db.query(User).filter_by(username=uname).first()
            if not u: continue
            q = _date_filter(db.query(Reservation).filter_by(user_id=u.id), 'custom', fa, ta)
            res_list = q.all()
            total    = len(res_list)
            approved = sum(1 for r in res_list if r.status=='approved')
            pending  = sum(1 for r in res_list if r.status=='pending')
            rejected = sum(1 for r in res_list if r.status=='rejected')
            cancelled= sum(1 for r in res_list if r.status=='cancelled')
            user_labels.append(u.full_name or u.username)
            user_counts.append(total)
            user_details.append({
                _t('المستخدم','User'): u.full_name or u.username,
                _t('الإجمالي','Total'): total,
                _t('موافقة','Approved'): approved,
                _t('معلقة','Pending'): pending,
                _t('مرفوضة','Rejected'): rejected,
                _t('ملغاة','Cancelled'): cancelled,
            })
        period_label = f'{fa} → {ta}'
        result = {
            'title': _t(f'مقارنة المستخدمين: {period_label}', f'Users Comparison: {period_label}'),
            'datasets': [{'label': _t('الإجمالي','Total Reservations'), 'data': dict(zip(user_labels, user_counts))}],
            'metrics': user_labels,
            'multi': True,
            'user_details': user_details,
            'period': period_label,
        }

    elif mode == 'user_venue' and a and b:
        u = db.query(User).filter_by(username=b).first()
        result = {'label_a': a, 'label_b': u.full_name if u else b,
                  'title': _t('مقارنة: قاعة ↔ مستخدم','Venue ↔ User Comparison'),
                  'data_a': _venue_stats(a), 'data_b': _user_stats(b), 'cross': True}

    elif mode == 'user_location' and a and b:
        u = db.query(User).filter_by(username=b).first()
        result = {'label_a': a, 'label_b': u.full_name if u else b,
                  'title': _t('مقارنة: موقع ↔ مستخدم','Location ↔ User Comparison'),
                  'data_a': _location_stats(a), 'data_b': _user_stats(b), 'cross': True}

    elif mode == 'venue_location' and a and b:
        result = {'label_a': a, 'label_b': b,
                  'title': _t('مقارنة: قاعة ↔ موقع','Venue ↔ Location Comparison'),
                  'data_a': _venue_stats(a), 'data_b': _location_stats(b), 'cross': True}

    elif mode == 'status':
        q = db.query(Reservation)
        data = {_t('موافقة','Approved'): q.filter_by(status='approved').count(),
                _t('معلقة','Pending'):  q.filter_by(status='pending').count(),
                _t('مرفوضة','Rejected'): q.filter_by(status='rejected').count(),
                _t('ملغاة','Cancelled'):  q.filter_by(status='cancelled').count(),
                _t('مكتملة','Completed'): q.filter_by(status='completed').count()}
        result = {'label_a': _t('توزيع الحالات','Status Distribution'), 'label_b': '',
                  'title': _t('توزيع حالات الحجوزات','Reservations Status Distribution'),
                  'data_a': data, 'data_b': {}, 'single': True}

    elif mode == 'rating_all':
        venues_data = {}
        for v in db.query(Venue).filter_by(is_active=True).order_by(Venue.name).all():
            rows = db.query(Rating).filter_by(venue_id=v.id).all()
            stars = [r.rating for r in rows if r.rating]
            venues_data[v.name] = round(sum(stars)/len(stars), 1) if stars else 0
        top = dict(sorted(venues_data.items(), key=lambda x: x[1], reverse=True)[:10])
        result = {'label_a': _t('متوسط التقييم','Average Rating'), 'label_b': '',
                  'title': _t('أعلى القاعات تقييماً','Top Rated Venues'),
                  'data_a': top, 'data_b': {}, 'single': True}

    elif mode == 'date_all':
        # Reservations by date — daily activity for selected period
        date_range = request.args.get('date_range', '30')
        try:
            days_back = int(date_range)
        except:
            days_back = 30
        from datetime import date as _date
        today = _date.today()
        daily_data = {}
        status_daily = {}
        for i in range(days_back - 1, -1, -1):
            d = today - timedelta(days=i)
            d_start = datetime.combine(d, datetime.min.time())
            d_end   = datetime.combine(d, datetime.max.time())
            q_day = db.query(Reservation).filter(
                Reservation.start_time >= d_start,
                Reservation.start_time <= d_end
            )
            label = d.strftime('%m/%d')
            daily_data[label] = q_day.count()
            status_daily[label] = {
                _t('موافقة','Approved'):   q_day.filter_by(status='approved').count(),
                _t('معلقة','Pending'):     q_day.filter_by(status='pending').count(),
                _t('مرفوضة','Rejected'):   q_day.filter_by(status='rejected').count(),
                _t('ملغاة','Cancelled'):   q_day.filter_by(status='cancelled').count(),
                _t('مكتملة','Completed'):  q_day.filter_by(status='completed').count(),
            }
        # Total stats for the period
        period_start = datetime.combine(today - timedelta(days=days_back-1), datetime.min.time())
        q_period = db.query(Reservation).filter(Reservation.start_time >= period_start)
        period_total = q_period.count()
        busiest_day  = max(daily_data, key=daily_data.get) if daily_data else '—'
        result = {
            'title':        _t(f'نشاط الحجوزات — آخر {days_back} يوم',
                               f'Reservation Activity — Last {days_back} days'),
            'date_all':     True,
            'daily_data':   daily_data,
            'status_daily': status_daily,
            'date_range':   days_back,
            'period_total': period_total,
            'busiest_day':  busiest_day,
            'busiest_count':daily_data.get(busiest_day, 0),
            'label_a': '', 'label_b': '', 'data_a': daily_data, 'data_b': {},
        }

    elif mode == 'email_stats':
        try:
            from models.database import EmailLog
            from sqlalchemy import func as _func
            date_range = request.args.get('date_range', '30')
            try:
                days_back = int(date_range)
            except:
                days_back = 30
            from datetime import date as _date
            today        = _date.today()
            period_start = datetime.combine(today - timedelta(days=days_back-1), datetime.min.time())
            q_all  = db.query(EmailLog)
            q_per  = db.query(EmailLog).filter(EmailLog.sent_at >= period_start)

            # Daily activity
            email_daily = {}
            for i in range(days_back - 1, -1, -1):
                d       = today - timedelta(days=i)
                d_start = datetime.combine(d, datetime.min.time())
                d_end   = datetime.combine(d, datetime.max.time())
                cnt = db.query(EmailLog).filter(
                    EmailLog.sent_at >= d_start,
                    EmailLog.sent_at <= d_end
                ).count()
                email_daily[d.strftime('%m/%d')] = cnt

            # By type breakdown
            types_data = {}
            for etype in ['bulk', 'notification', 'resend', 'invitation']:
                cnt = q_per.filter(EmailLog.type == etype).count()
                if cnt > 0:
                    types_data[_t(
                        {'bulk':'جماعي','notification':'إشعار','resend':'إعادة إرسال','invitation':'دعوة'}.get(etype, etype),
                        {'bulk':'Bulk','notification':'Notification','resend':'Resend','invitation':'Invitation'}.get(etype, etype)
                    )] = cnt

            # By user (top 10 recipients)
            from sqlalchemy import func as _func2
            top_recipients = (db.query(EmailLog.recipient, _func2.count(EmailLog.id).label('cnt'))
                              .filter(EmailLog.sent_at >= period_start)
                              .group_by(EmailLog.recipient)
                              .order_by(_func2.count(EmailLog.id).desc())
                              .limit(10).all())
            users_email_data = {r.recipient or '—': r.cnt for r in top_recipients}

            # Success rate over time
            total_period = q_per.count()
            sent_period  = q_per.filter(EmailLog.status == 'sent').count()
            fail_period  = q_per.filter(EmailLog.status == 'failed').count()
            rate         = round(sent_period / total_period * 100) if total_period else 0

            # Per-reservation booking emails (map to venues if user_id available)
            venue_email_map = {}
            try:
                from models.database import User as _UE, Reservation as _RE
                user_emails = (db.query(EmailLog.user_id, _func2.count(EmailLog.id).label('cnt'))
                               .filter(EmailLog.sent_at >= period_start, EmailLog.user_id != None)
                               .group_by(EmailLog.user_id)
                               .order_by(_func2.count(EmailLog.id).desc())
                               .limit(8).all())
                for ue in user_emails:
                    u = db.query(_UE).get(ue.user_id)
                    if u:
                        venue_email_map[u.full_name or u.username] = ue.cnt
            except Exception:
                pass

            result = {
                'title':          _t(f'تحليل البريد الإلكتروني — آخر {days_back} يوم',
                                     f'Email Analytics — Last {days_back} days'),
                'email_mode':     True,
                'date_range':     days_back,
                'total_all':      q_all.count(),
                'sent_all':       q_all.filter(EmailLog.status=='sent').count(),
                'failed_all':     q_all.filter(EmailLog.status=='failed').count(),
                'bulk_all':       q_all.filter(EmailLog.type=='bulk').count(),
                'notif_all':      q_all.filter(EmailLog.type=='notification').count(),
                'period_sent':    sent_period,
                'period_fail':    fail_period,
                'period_total':   total_period,
                'success_rate':   rate,
                'email_daily':    email_daily,
                'types_data':     types_data,
                'users_data':     users_email_data,
                'venue_email_map':venue_email_map,
                'label_a': '', 'label_b': '', 'data_a': {}, 'data_b': {},
            }
        except Exception as e:
            result = {'title': 'Email Stats', 'email_mode': True,
                      'error': str(e), 'label_a':'','label_b':'','data_a':{},'data_b':{}}

    return render_template('reports/comparison.html',
        mode=mode, result=result, a=a, b=b,
        fa=fa, ta=ta, fb=fb, tb=tb,
        items_vals=items_vals, items_keys=items_keys,
        venues=venues, locations=locations, users_all=users_all)
    db    = get_db()
    perms = get_permissions()
    if not perms.is_admin_or_manager(): abort(403)

    _lang = session.get('lang', 'ar')
    def _t(ar, en): return en if _lang == 'en' else ar
    mode  = request.args.get('mode', 'venue')
    a     = request.args.get('a', '')
    b     = request.args.get('b', '')
    fa    = request.args.get('fa', '')
    ta    = request.args.get('ta', '')
    fb    = request.args.get('fb', '')
    tb    = request.args.get('tb', '')

    venues    = db.query(Venue).filter_by(is_active=True).order_by(Venue.name).all()
    locations = db.query(Location).filter_by(is_active=True).order_by(Location.name).all()
    users_all = db.query(User).filter_by(is_active=True).order_by(User.full_name).all()

    result = {}

    def _venue_stats(name):
        v = db.query(Venue).filter_by(name=name).first()
        if not v: return {}
        q = db.query(Reservation).filter_by(venue_id=v.id)
        total = q.count()
        return {
            'إجمالي الحجوزات': total,
            'موافقة':   q.filter_by(status='approved').count(),
            'معلقة':    q.filter_by(status='pending').count(),
            'مرفوضة':  q.filter_by(status='rejected').count(),
            'ملغاة':    q.filter_by(status='cancelled').count(),
            'مكتملة':   q.filter_by(status='completed').count(),
            'فترات محظورة': db.query(BlockedPeriod).filter_by(venue_id=v.id).count(),
        }

    def _location_stats(name):
        loc = db.query(Location).filter_by(name=name).first()
        if not loc: return {}
        venue_ids = [v.id for v in loc.venues]
        q = db.query(Reservation).filter(Reservation.venue_id.in_(venue_ids)) if venue_ids else db.query(Reservation).filter_by(id=-1)
        return {
            'عدد القاعات': len(loc.venues),
            'إجمالي الحجوزات': q.count(),
            'موافقة':   q.filter_by(status='approved').count(),
            'معلقة':    q.filter_by(status='pending').count(),
            'مرفوضة':  q.filter_by(status='rejected').count(),
            'ملغاة':    q.filter_by(status='cancelled').count(),
        }

    def _user_stats(uname):
        u = db.query(User).filter_by(username=uname).first()
        if not u: return {}
        q = db.query(Reservation).filter_by(user_id=u.id)
        return {
            'إجمالي الحجوزات': q.count(),
            'موافقة':   q.filter_by(status='approved').count(),
            'معلقة':    q.filter_by(status='pending').count(),
            'مرفوضة':  q.filter_by(status='rejected').count(),
            'ملغاة':    q.filter_by(status='cancelled').count(),
            'مكتملة':   q.filter_by(status='completed').count(),
        }

    def _period_stats(d_from, d_to):
        try:
            s = datetime.strptime(d_from, '%Y-%m-%d')
            e = datetime.strptime(d_to, '%Y-%m-%d').replace(hour=23, minute=59)
            q = db.query(Reservation).filter(
                Reservation.start_time >= s, Reservation.start_time <= e)
            return {
                'إجمالي': q.count(),
                'موافقة': q.filter_by(status='approved').count(),
                'معلقة':  q.filter_by(status='pending').count(),
                'مرفوضة': q.filter_by(status='rejected').count(),
                'ملغاة':  q.filter_by(status='cancelled').count(),
                'مكتملة': q.filter_by(status='completed').count(),
            }
        except: return {}

    if mode == 'venue' and a and b:
        result = {'label_a': a, 'label_b': b, 'title': _t('مقارنة القاعات', 'Venues Comparison'),
                  'data_a': _venue_stats(a), 'data_b': _venue_stats(b)}

    elif mode == 'location' and a and b:
        result = {'label_a': a, 'label_b': b, 'title': _t('مقارنة المواقع', 'Locations Comparison'),
                  'data_a': _location_stats(a), 'data_b': _location_stats(b)}

    elif mode == 'period' and fa and ta and fb and tb:
        result = {'label_a': f'{fa} → {ta}', 'label_b': f'{fb} → {tb}',
                  'title': _t('مقارنة الفترات', 'Periods Comparison'),
                  'data_a': _period_stats(fa, ta), 'data_b': _period_stats(fb, tb)}

    elif mode == 'user' and a and b:
        ua = db.query(User).filter_by(username=a).first()
        ub = db.query(User).filter_by(username=b).first()
        result = {'label_a': ua.full_name if ua else a,
                  'label_b': ub.full_name if ub else b,
                  'title': _t('مقارنة المستخدمين', 'Users Comparison'),
                  'data_a': _user_stats(a), 'data_b': _user_stats(b)}

    # Cross: User ↔ Venue
    elif mode == 'user_venue' and a and b:
        u = db.query(User).filter_by(username=b).first()
        result = {'label_a': a, 'label_b': u.full_name if u else b,
                  'title': _t('مقارنة: قاعة ↔ مستخدم', 'Venue ↔ User Comparison'),
                  'data_a': _venue_stats(a), 'data_b': _user_stats(b),
                  'cross': True}

    # Cross: User ↔ Location
    elif mode == 'user_location' and a and b:
        u = db.query(User).filter_by(username=b).first()
        result = {'label_a': a, 'label_b': u.full_name if u else b,
                  'title': _t('مقارنة: موقع ↔ مستخدم', 'Location ↔ User Comparison'),
                  'data_a': _location_stats(a), 'data_b': _user_stats(b),
                  'cross': True}

    # Cross: Venue ↔ Location
    elif mode == 'venue_location' and a and b:
        result = {'label_a': a, 'label_b': b,
                  'title': _t('مقارنة: قاعة ↔ موقع', 'Venue ↔ Location Comparison'),
                  'data_a': _venue_stats(a), 'data_b': _location_stats(b),
                  'cross': True}

    elif mode == 'status':
        q = db.query(Reservation)
        data = {'موافقة': q.filter_by(status='approved').count(),
                'معلقة':  q.filter_by(status='pending').count(),
                'مرفوضة': q.filter_by(status='rejected').count(),
                'ملغاة':  q.filter_by(status='cancelled').count(),
                'مكتملة': q.filter_by(status='completed').count()}
        result = {'label_a': 'توزيع الحالات', 'label_b': '',
                  'title': _t('توزيع حالات الحجوزات', 'Status Distribution'),
                  'data_a': data, 'data_b': {}, 'single': True}

    # ── Rating: Venue ↔ Venue ────────────────────────────────────────────────
    elif mode == 'rating_venue' and a and b:
        def _rating_venue(name):
            v = db.query(Venue).filter_by(name=name).first()
            if not v: return {}
            rows = db.query(Rating).filter_by(venue_id=v.id).all()
            if not rows: return {'عدد التقييمات': 0, 'متوسط النجوم': 0, '5 نجوم': 0, '4 نجوم': 0, '3 نجوم': 0, '2 نجوم': 0, '1 نجمة': 0}
            stars = [r.rating for r in rows if r.rating]
            return {
                'عدد التقييمات': len(rows),
                'متوسط النجوم':  round(sum(stars)/len(stars), 1) if stars else 0,
                '5 نجوم': sum(1 for s in stars if s == 5),
                '4 نجوم': sum(1 for s in stars if s == 4),
                '3 نجوم': sum(1 for s in stars if s == 3),
                '2 نجوم': sum(1 for s in stars if s == 2),
                '1 نجمة': sum(1 for s in stars if s == 1),
            }
        result = {'label_a': a, 'label_b': b,
                  'title': _t('مقارنة تقييمات القاعات', 'Venue Ratings Comparison'),
                  'data_a': _rating_venue(a), 'data_b': _rating_venue(b)}

    # ── Rating: Location ↔ Location ──────────────────────────────────────────
    elif mode == 'rating_location' and a and b:
        def _rating_location(name):
            loc = db.query(Location).filter_by(name=name).first()
            if not loc: return {}
            venue_ids = [v.id for v in loc.venues]
            if not venue_ids: return {'عدد التقييمات': 0, 'متوسط النجوم': 0}
            rows = db.query(Rating).filter(Rating.venue_id.in_(venue_ids)).all()
            stars = [r.rating for r in rows if r.rating]
            return {
                'عدد القاعات': len(loc.venues),
                'عدد التقييمات': len(rows),
                'متوسط النجوم':  round(sum(stars)/len(stars), 1) if stars else 0,
                '5 نجوم': sum(1 for s in stars if s == 5),
                '4 نجوم': sum(1 for s in stars if s == 4),
                '3 نجوم': sum(1 for s in stars if s == 3),
                'أقل من 3': sum(1 for s in stars if s < 3),
            }
        result = {'label_a': a, 'label_b': b,
                  'title': _t('مقارنة تقييمات المواقع', 'Location Ratings Comparison'),
                  'data_a': _rating_location(a), 'data_b': _rating_location(b)}

    # ── Rating: User ↔ User ───────────────────────────────────────────────────
    elif mode == 'rating_user' and a and b:
        def _rating_user(uname):
            u = db.query(User).filter_by(username=uname).first()
            if not u: return {}
            rows = db.query(Rating).filter_by(user_id=u.id).all()
            stars = [r.rating for r in rows if r.rating]
            return {
                'عدد تقييماته': len(rows),
                'متوسط تقييماته': round(sum(stars)/len(stars), 1) if stars else 0,
                '5 نجوم': sum(1 for s in stars if s == 5),
                '4 نجوم': sum(1 for s in stars if s == 4),
                '3 نجوم': sum(1 for s in stars if s == 3),
                'أقل من 3': sum(1 for s in stars if s < 3),
            }
        ua = db.query(User).filter_by(username=a).first()
        ub = db.query(User).filter_by(username=b).first()
        result = {'label_a': ua.full_name if ua else a,
                  'label_b': ub.full_name if ub else b,
                  'title': _t('مقارنة تقييمات المستخدمين', 'User Ratings Comparison'),
                  'data_a': _rating_user(a), 'data_b': _rating_user(b)}

    # ── Rating: Period ↔ Period ───────────────────────────────────────────────
    elif mode == 'rating_period' and fa and ta and fb and tb:
        def _rating_period(d_from, d_to):
            try:
                s = datetime.strptime(d_from, '%Y-%m-%d')
                e = datetime.strptime(d_to, '%Y-%m-%d').replace(hour=23, minute=59)
                rows = db.query(Rating).filter(Rating.created_at >= s, Rating.created_at <= e).all()
                stars = [r.rating for r in rows if r.rating]
                return {
                    'عدد التقييمات': len(rows),
                    'متوسط النجوم':  round(sum(stars)/len(stars), 1) if stars else 0,
                    '5 نجوم': sum(1 for s in stars if s == 5),
                    '4 نجوم': sum(1 for s in stars if s == 4),
                    '3 نجوم': sum(1 for s in stars if s == 3),
                    '2 نجوم': sum(1 for s in stars if s == 2),
                    '1 نجمة': sum(1 for s in stars if s == 1),
                }
            except: return {}
        result = {'label_a': f'{fa} → {ta}', 'label_b': f'{fb} → {tb}',
                  'title': _t('مقارنة التقييمات بالفترات', 'Period Ratings Comparison'),
                  'data_a': _rating_period(fa, ta), 'data_b': _rating_period(fb, tb)}

    # ── Rating: All venues distribution (single) ──────────────────────────────
    elif mode == 'rating_all':
        venues_data = {}
        for v in db.query(Venue).filter_by(is_active=True).order_by(Venue.name).all():
            rows = db.query(Rating).filter_by(venue_id=v.id).all()
            stars = [r.rating for r in rows if r.rating]
            venues_data[v.name] = round(sum(stars)/len(stars), 1) if stars else 0
        # Sort by rating desc, top 10
        top = dict(sorted(venues_data.items(), key=lambda x: x[1], reverse=True)[:10])
        result = {'label_a': 'متوسط التقييم', 'label_b': '',
                  'title': _t('أعلى القاعات تقييماً', 'Top Rated Venues'),
                  'data_a': top, 'data_b': {}, 'single': True}

    return render_template('reports/comparison.html',
        mode=mode, result=result, a=a, b=b,
        fa=fa, ta=ta, fb=fb, tb=tb,
        venues=venues, locations=locations, users_all=users_all)


# ── Comparison PDF Export ─────────────────────────────────────────────────────
@reports_bp.route('/comparison/pdf')
@login_required
def comparison_pdf():
    from flask import Response
    import io
    db = get_db()
    perms = get_permissions()
    if not perms.is_admin_or_manager(): abort(403)

    # Re-use the comparison view logic by redirecting internally
    # Instead: generate a clean PDF from the print page using weasyprint-style
    # Simple approach: use browser print (redirect to comparison with print param)
    # OR: generate PDF with reportlab using plain text (no Arabic paragraph)
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from utils.pdf_helper import register_arabic_font
        register_arabic_font()
        from utils.pdf_helper import arabic_font as _af, arabic_font_bold as _afb

        # Rebuild result
        from flask import request as req
        from routes.reports import comparison as _comp_view
        # Call comparison logic directly
        mode = req.args.get('mode', 'venue')
        a = req.args.get('a', ''); b = req.args.get('b', '')
        fa = req.args.get('fa', ''); ta = req.args.get('ta', '')
        fb = req.args.get('fb', ''); tb = req.args.get('tb', '')

        # We'll generate PDF from current comparison result
        # Use print-friendly redirect instead
        from flask import redirect, url_for
        # Best approach: use browser print via JS — redirect back with print=1
        return redirect(url_for('reports.comparison', **req.args) + '#print')

    except Exception as e:
        from utils.flash_helper import flash_msg
        flash_msg(f'PDF: يرجى استخدام زر الطباعة في المتصفح', 'info')
        from flask import redirect, request as req
        return redirect(req.referrer or url_for('reports.comparison'))


# ── Comparison CSV Export ─────────────────────────────────────────────────────
@reports_bp.route('/comparison/csv')
@login_required
def comparison_csv():
    from flask import Response
    import csv, io
    db = get_db()
    perms = get_permissions()
    if not perms.is_admin_or_manager(): abort(403)
    # Get result from session or rebuild - use simple CSV of URL params
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Metric', 'A', 'B', 'Delta%'])
    writer.writerow(['mode', request.args.get('mode',''), '', ''])
    writer.writerow(['a', request.args.get('a',''), '', ''])
    writer.writerow(['b', request.args.get('b',''), '', ''])
    output.seek(0)
    return Response(output.read(), mimetype='text/csv',
                    headers={'Content-Disposition': 'attachment; filename=comparison.csv'})
def _get_report_data(db, perms, rtype, filter_val, start, end_d):
    """Returns (list_of_dicts, title_string) for PDF/Excel export."""
    rows = []
    title = 'Report'
    if rtype == 'my_reservations':
        title = 'My Reservations'
        q = _date_filter(db.query(Reservation).filter_by(user_id=current_user.id), filter_val, start, end_d)
        for r in q.all():
            rows.append({'Ref': r.booking_number, 'Title': r.title,
                'Venue': r.venue.name if r.venue else '',
                'Start': r.start_time.strftime('%Y-%m-%d %H:%M') if r.start_time else '',
                'End':   r.end_time.strftime('%H:%M') if r.end_time else '',
                'Status': r.status})
    elif rtype == 'all_reservations' and perms.is_admin_or_manager():
        title = 'All Reservations'
        q = _date_filter(db.query(Reservation), filter_val, start, end_d)
        for r in q.all():
            rows.append({'Ref': r.booking_number, 'Title': r.title,
                'User': r.user.full_name if r.user else '',
                'Venue': r.venue.name if r.venue else '',
                'Start': r.start_time.strftime('%Y-%m-%d %H:%M') if r.start_time else '',
                'Status': r.status})
    elif rtype == 'venues':
        title = 'Venues Report'
        for v in db.query(Venue).all():
            rows.append({'Venue': v.name,
                'Location': v.location.name if v.location else '',
                'Total Reservations': len(v.reservations),
                'Blocked': db.query(BlockedPeriod).filter_by(venue_id=v.id).count(),
                'Status': 'Active' if v.is_active else 'Inactive'})
    elif rtype == 'locations':
        title = 'Locations Report'
        for l in db.query(Location).all():
            rows.append({'Location': l.name, 'City': l.city or '', 'Area': l.area or '',
                'Venues': len(l.venues),
                'Total Reservations': sum(len(v.reservations) for v in l.venues)})
    elif rtype == 'users' and perms.can('users_view'):
        title = 'Users Report'
        for u in db.query(User).all():
            rows.append({'Full Name': u.full_name, 'Username': u.username,
                'Email': u.email or '', 'Role': u.role.name if u.role else '',
                'Reservations': len(u.reservations),
                'Status': 'Active' if u.is_active else 'Inactive'})
    return rows, title


# ── Export PDF ────────────────────────────────────────────────────────────────
@reports_bp.route('/export/pdf')
@login_required
def export_pdf():
    from flask import Response
    import io
    db    = get_db()
    perms = get_permissions()
    rtype  = request.args.get('type', 'my_reservations')
    filter = request.args.get('filter', 'all')
    start  = request.args.get('start', '')
    end    = request.args.get('end', '')
    rows, title_txt = _get_report_data(db, perms, rtype, filter, start, end)

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.lib.units import cm
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        import os, json

        buf = io.BytesIO()
        # Use landscape for wide tables
        page_size = landscape(A4) if rows and len(rows[0]) > 5 else A4
        doc = SimpleDocTemplate(buf, pagesize=page_size,
                                rightMargin=1.5*cm, leftMargin=1.5*cm,
                                topMargin=2*cm, bottomMargin=2*cm)

        # Use ARS Arabic font helper
        from utils.pdf_helper import register_arabic_font, ar, arabic_font as _af
        register_arabic_font()
        arabic_font = _af(bold=False)
        arabic_font_bold = _af(bold=True)

        styles = getSampleStyleSheet()
        story  = []

        # Load report header from maintenance config
        mcfg = {}
        try:
            cfg_path = os.path.join(os.path.dirname(__file__), '..', 'maintenance_config.json')
            if os.path.exists(cfg_path):
                mcfg = json.loads(open(cfg_path).read())
        except: pass

        # Header
        if mcfg.get('report_header_title'):
            h_style = ParagraphStyle('h', fontName=arabic_font_bold, fontSize=14,
                                     textColor=colors.HexColor('#0C67EC'), alignment=TA_CENTER, spaceAfter=2)
            story.append(Paragraph(ar(mcfg['report_header_title']), h_style))
        if mcfg.get('report_header_subtitle'):
            s_style = ParagraphStyle('s', fontName=arabic_font, fontSize=11,
                                     textColor=colors.HexColor('#3D8EF5'), alignment=TA_CENTER, spaceAfter=2)
            story.append(Paragraph(ar(mcfg['report_header_subtitle']), s_style))
        if mcfg.get('report_header_title') or mcfg.get('report_header_subtitle'):
            story.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#0C67EC'), spaceAfter=8))

        title_style = ParagraphStyle('title', fontName=arabic_font, fontSize=13,
                                     textColor=colors.HexColor('#0C67EC'), spaceAfter=6)
        story.append(Paragraph(f'STAP — {title_txt}', title_style))
        from datetime import date
        date_style = ParagraphStyle('date', fontName=arabic_font, fontSize=9,
                                    textColor=colors.grey, spaceAfter=12)
        story.append(Paragraph(f'Report Date: {date.today()}', date_style))
        story.append(Spacer(1, 0.3*cm))

        if rows:
            headers = list(rows[0].keys())
            col_w = (page_size[0] - 3*cm) / len(headers)
            def _cell(v):
                s = str(v) if v else ''
                import re
                return ar(s) if re.search(r'[\u0600-\u06ff]', s) else s
            data = [headers] + [[_cell(r.get(h,'')) for h in headers] for r in rows]
            t = Table(data, repeatRows=1, colWidths=[col_w]*len(headers))
            t.setStyle(TableStyle([
                ('BACKGROUND',    (0,0), (-1,0), colors.HexColor('#0C67EC')),
                ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
                ('FONTNAME',      (0,0), (-1,-1), arabic_font),
                ('FONTSIZE',      (0,0), (-1,0), 9),
                ('FONTSIZE',      (0,1), (-1,-1), 8),
                ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
                ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
                ('GRID',          (0,0), (-1,-1), 0.4, colors.HexColor('#dee2e6')),
                ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.white, colors.HexColor('#EEF4F5')]),
                ('PADDING',       (0,0), (-1,-1), 5),
                ('TOPPADDING',    (0,0), (-1,0), 8),
                ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ]))
            story.append(t)
        else:
            story.append(Paragraph('No data available.', styles['Normal']))

        doc.build(story)
        buf.seek(0)
        return Response(buf.read(), mimetype='application/pdf',
                        headers={'Content-Disposition': f'attachment;filename=ARS_Report_{date.today()}.pdf'})
    except ImportError:
        flash_msg('يرجى تثبيت reportlab: pip install reportlab', 'danger')
        return redirect(url_for('reports.index'))


# ── Export Excel ──────────────────────────────────────────────────────────────
@reports_bp.route('/export/excel')
@login_required
def export_excel():
    from flask import Response
    import io
    db    = get_db()
    perms = get_permissions()
    rtype  = request.args.get('type', 'my_reservations')
    filter = request.args.get('filter', 'all')
    start  = request.args.get('start', '')
    end    = request.args.get('end', '')
    rows, title_txt = _get_report_data(db, perms, rtype, filter, start, end)

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.chart import BarChart, Reference
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title_txt[:30]
        header_fill = PatternFill('solid', fgColor='0C67EC')
        header_font = Font(color='FFFFFF', bold=True)
        if rows:
            headers = list(rows[0].keys())
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(1, ci, h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(
                    bottom=Side(style='medium', color='FFFFFF'),
                    right=Side(style='thin', color='2E8B8F'))
            ws.row_dimensions[1].height = 28
            for ri, row in enumerate(rows, 2):
                for ci, h in enumerate(headers, 1):
                    val = row.get(h, '')
                    cell = ws.cell(ri, ci, val)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if ri % 2 == 0:
                        cell.fill = PatternFill('solid', fgColor='EEF4F5')
                ws.row_dimensions[ri].height = 20
            # Auto-fit column widths
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except: pass
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)
            # Freeze top row
            ws.freeze_panes = 'A2' 
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return Response(buf.read(),
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition': 'attachment;filename=ARS_Report.xlsx'})
    except ImportError:
        flash_msg('يرجى تثبيت openpyxl: pip install openpyxl', 'danger')
        return redirect(url_for('reports.index'))




# ── Export Excel ──────────────────────────────────────────────────────────────
