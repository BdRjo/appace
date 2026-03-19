"""
المواقع والقاعات — مطابق لـ v54 مع إدارة وصول المستخدمين
"""
import csv, io
from datetime import datetime
from utils.flash_helper import flash_msg
from flask import (Blueprint, render_template, redirect, url_for,
                   request, flash, abort, Response)
from flask_login import login_required, current_user
from models.database import Location, Venue, User, SystemLog
from utils.helpers import get_db, admin_required, syslog, paginate

locations_bp    = Blueprint('locations',    __name__, url_prefix='/locations')
venues_mgmt_bp  = Blueprint('venues_mgmt', __name__, url_prefix='/venues-mgmt')


# ══════════════════════════════════════════════════════════════
#  LOCATIONS
# ══════════════════════════════════════════════════════════════

@locations_bp.route('/')
@login_required
@admin_required
def index():
    db     = get_db()
    search = request.args.get('q','')
    page   = request.args.get('page',1,type=int)
    q      = db.query(Location)
    if search: q = q.filter(Location.name.ilike(f'%{search}%'))
    items, total, total_pages = paginate(q.order_by(Location.name), page, 20)
    return render_template('admin/locations.html',
        locations=items, total=total, page=page, total_pages=total_pages, search=search)


@locations_bp.route('/new', methods=['GET','POST'])
@login_required
@admin_required
def new():
    db = get_db()
    if request.method == 'POST':
        name    = request.form.get('name','').strip()
        name_en = request.form.get('name_en','').strip()
        city    = request.form.get('city','').strip()
        area    = request.form.get('area','').strip()
        is_active = request.form.get('is_active') == 'on'

        if not name:
            flash_msg('اسم الموقع مطلوب', 'danger')
            return render_template('admin/location_form.html', loc=None, form=request.form)

        loc = Location(name=name, name_en=name_en, city=city, area=area, is_active=is_active)
        db.add(loc); db.commit()
        syslog('ADD_LOCATION', f'موقع جديد: {name}')
        flash_msg(f'✅ تمت إضافة الموقع: {name}', 'success')
        return redirect(url_for('locations.index'))

    return render_template('admin/location_form.html', loc=None, form={},
                           all_users=[], assigned_users=set())


@locations_bp.route('/<int:loc_id>/edit', methods=['GET','POST'])
@login_required
@admin_required
def edit(loc_id):
    db  = get_db()
    loc = db.query(Location).get(loc_id)
    if not loc: abort(404)

    all_users = db.query(User).filter(
        User.is_active == True,
        User.role.has(~User.role.property.mapper.class_.name.in_(['مدير النظام','مشرف']))
    ).all() if False else \
        [u for u in db.query(User).filter_by(is_active=True).all()
         if u.role and u.role.name not in ('مدير النظام','مشرف')]

    if request.method == 'POST':
        loc.name      = request.form.get('name','').strip()
        loc.name_en   = request.form.get('name_en','').strip()
        loc.city      = request.form.get('city','').strip()
        loc.area      = request.form.get('area','').strip()
        loc.is_active = request.form.get('is_active') == 'on'

        # إدارة وصول المستخدمين — مطابق لـ v54 LocationWindow access tab
        user_ids = request.form.getlist('allowed_users', type=int)
        loc.allowed_users = [u for u in all_users if u.id in user_ids]

        db.commit()
        syslog('EDIT_LOCATION', f'تعديل الموقع: {loc.name}')
        flash_msg('✅ تم تحديث الموقع', 'success')
        return redirect(url_for('locations.index'))

    assigned = {u.id for u in loc.allowed_users}
    return render_template('admin/location_form.html',
        loc=loc, form={}, all_users=all_users, assigned_users=assigned)


@locations_bp.route('/bulk-delete', methods=['POST'])
@login_required
@admin_required
def bulk_delete():
    db = get_db(); ids = request.form.getlist('ids')
    count = 0; skipped = 0
    for lid in ids:
        try:
            loc = db.query(Location).get(int(lid))
            if loc:
                if loc.venues: skipped += 1
                else: db.delete(loc); count += 1
        except: pass
    db.commit()
    msg = f'✅ تم حذف {count} موقع'
    if skipped: msg += f' (تجاوز {skipped} تحتوي قاعات)'
    flash_msg(msg, 'success' if count else 'warning')
    return redirect(url_for('locations.index'))


@locations_bp.route('/<int:loc_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete(loc_id):
    db  = get_db()
    loc = db.query(Location).get(loc_id)
    if not loc: abort(404)
    if loc.venues:
        flash_msg('لا يمكن حذف موقع يحتوي على قاعات', 'danger')
        return redirect(url_for('locations.index'))
    name = loc.name; db.delete(loc); db.commit()
    syslog('DELETE_LOCATION', f'حذف الموقع: {name}')
    flash_msg(f'تم حذف الموقع: {name}', 'success')
    return redirect(url_for('locations.index'))



def _make_xlsx(headers, rows, sheet_name, is_en):
    """Build an Excel file with blue header, alternating rows, RTL support."""
    import io as _io
    from datetime import datetime as _dt
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = sheet_name[:31]
    ws.sheet_view.rightToLeft = not is_en
    h_align = 'left' if is_en else 'right'
    hfill = PatternFill('solid', fgColor='0C67EC')
    hfont = Font(color='FFFFFF', bold=True)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h); c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal='center', vertical='center')
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci, val).alignment = Alignment(horizontal=h_align, vertical='center')
        if ri % 2 == 0:
            for ci in range(1, len(headers)+1):
                ws.cell(ri, ci).fill = PatternFill('solid', fgColor='EEF4F5')
    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, 40)
    ws.freeze_panes = 'A2'
    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

# ── CSV export — مطابق لـ v54 export_locations_csv ───────────────────────────
@locations_bp.route('/export-csv')
@locations_bp.route('/export-excel')
@login_required
@admin_required
def export_csv():
    from flask import session; from datetime import datetime as _dt
    db = get_db(); is_en = session.get('lang','ar') == 'en'
    headers = (['Name (AR)','Name (EN)','City','Area','Venues Count','Status'] if is_en
               else ['الاسم بالعربي','الاسم بالإنجليزي','المدينة','المنطقة','عدد القاعات','الحالة'])
    rows = []
    for l in db.query(Location).order_by(Location.name).all():
        rows.append([l.name, l.name_en or '', l.city or '', l.area or '', len(l.venues),
                     ('Active' if l.is_active else 'Inactive') if is_en else ('نشط' if l.is_active else 'غير نشط')])
    try:
        data = _make_xlsx(headers, rows, 'Locations' if is_en else 'المواقع', is_en)
        fname = f'locations_{_dt.now().strftime("%Y%m%d")}.xlsx'
        return Response(data, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition': f'attachment; filename={fname}'})
    except ImportError:
        import csv as _csv; import io as _io
        out = _io.StringIO(); _csv.writer(out).writerows([headers] + rows)
        return Response(b'\xef\xbb\xbf' + out.getvalue().encode('utf-8'), mimetype='text/csv',
                        headers={'Content-Disposition': 'attachment; filename=locations.csv'})


# ── CSV Template Download ─────────────────────────────────────────────────────
@locations_bp.route('/csv-template')
@login_required
def csv_template():
    from flask import session
    is_en = session.get('lang','ar') == 'en'
    rows = [
        ['الاسم بالعربي','الاسم بالإنجليزي','المدينة','المنطقة','نشط (1/0)'],
        ['مبنى الإدارة','Admin Building','عمّان','العبدلي','1'],
        ['قاعة الاجتماعات','Meeting Hall','عمّان','الشميساني','1'],
    ] if not is_en else [
        ['Name (AR)','Name (EN)','City','Area','Active (1/0)'],
        ['مبنى الإدارة','Admin Building','Amman','Abdali','1'],
        ['قاعة الاجتماعات','Meeting Hall','Amman','Shmeisani','1'],
    ]
    import csv, io
    out = io.StringIO()
    csv.writer(out).writerows(rows)
    return Response(b'\xef\xbb\xbf' + out.getvalue().encode('utf-8'),
                    mimetype='text/csv',
                    headers={'Content-Disposition': 'attachment; filename=locations_template.csv'})


@venues_mgmt_bp.route('/csv-template')
@login_required
def csv_template():
    from flask import session
    is_en = session.get('lang','ar') == 'en'
    rows = [
        ['اسم القاعة','الاسم بالإنجليزي','اسم الموقع','الكود','السعة','التجهيزات','ملاحظات','يحتاج موافقة (1/0)','نشط (1/0)'],
        ['قاعة A1','Hall A1','مبنى الإدارة','A1','50','شاشة عرض، كاميرات','','0','1'],
        ['قاعة B2','Hall B2','مبنى الإدارة','B2','30','مكيف','','1','1'],
    ] if not is_en else [
        ['Name (AR)','Name (EN)','Location Name','Code','Capacity','Equipment','Notes','Requires Approval (1/0)','Active (1/0)'],
        ['قاعة A1','Hall A1','Admin Building','A1','50','Projector, Camera','','0','1'],
        ['قاعة B2','Hall B2','Admin Building','B2','30','AC','','1','1'],
    ]
    import csv, io
    out = io.StringIO()
    csv.writer(out).writerows(rows)
    return Response(b'\xef\xbb\xbf' + out.getvalue().encode('utf-8'),
                    mimetype='text/csv',
                    headers={'Content-Disposition': 'attachment; filename=venues_template.csv'})

# ── CSV import — مطابق لـ v54 ImportCSVWindow (mode=locations) ───────────────
@locations_bp.route('/import-csv', methods=['GET','POST'])
@login_required
@admin_required
def import_csv():
    db = get_db()
    if request.method == 'POST':
        f = request.files.get('file')
        if not f: flash_msg('يرجى رفع ملف CSV', 'danger'); return redirect(url_for('locations.import_csv'))

        content  = f.stream.read().decode('utf-8-sig')
        reader   = csv.DictReader(io.StringIO(content))
        skip_dup = request.form.get('skip_dup') == 'on'
        upd_dup  = request.form.get('update_dup') == 'on'

        imported = updated = skipped = errors = 0
        for row in reader:
            name = row.get('name','').strip()
            if not name: errors += 1; continue
            existing = db.query(Location).filter_by(name=name).first()
            active   = row.get('is_active','1').lower() not in ('0','false','no','لا')
            if existing:
                if upd_dup:
                    existing.name_en = row.get('name_en', existing.name_en or '')
                    existing.city    = row.get('city', existing.city or '')
                    existing.area    = row.get('area', existing.area or '')
                    existing.is_active = active
                    updated += 1
                elif skip_dup:
                    skipped += 1
                    continue
            else:
                db.add(Location(name=name, name_en=row.get('name_en',''),
                                city=row.get('city',''), area=row.get('area',''),
                                is_active=active))
                imported += 1
        db.commit()
        syslog('IMPORT_LOCATIONS', f'استيراد: جديد={imported}, محدّث={updated}, متخطى={skipped}, خطأ={errors}')
        flash_msg(f'✅ استيراد مكتمل — جديد: {imported}، محدّث: {updated}، متخطى: {skipped}، أخطاء: {errors}', 'success')
        return redirect(url_for('locations.index'))

    return render_template('admin/import_csv.html', mode='locations',
                           fields=['name','name_en','city','area','is_active'])


# ══════════════════════════════════════════════════════════════
#  VENUES MANAGEMENT (admin CRUD)
# ══════════════════════════════════════════════════════════════

@venues_mgmt_bp.route('/')
@login_required
@admin_required
def index():
    db     = get_db()
    search = request.args.get('q','')
    loc_f  = request.args.get('loc_id','')
    page   = request.args.get('page',1,type=int)
    q      = db.query(Venue)
    if search: q = q.filter(Venue.name.ilike(f'%{search}%'))
    if loc_f and loc_f.isdigit(): q = q.filter(Venue.location_id == int(loc_f))
    items, total, total_pages = paginate(q.order_by(Venue.name), page, 20)
    locs = db.query(Location).order_by(Location.name).all()
    return render_template('admin/venues_mgmt.html',
        venues=items, total=total, page=page, total_pages=total_pages,
        locations=locs, search=search, loc_f=loc_f)


@venues_mgmt_bp.route('/new', methods=['GET','POST'])
@login_required
@admin_required
def new():
    db   = get_db()
    locs = db.query(Location).filter_by(is_active=True).order_by(Location.name).all()

    if request.method == 'POST':
        name     = request.form.get('name','').strip()
        name_en  = request.form.get('name_en','').strip()
        loc_id   = request.form.get('location_id','')
        code     = request.form.get('code','').strip() or None
        capacity = request.form.get('capacity','')
        equipment= request.form.get('equipment','').strip()
        notes    = request.form.get('notes','').strip()
        requires = request.form.get('requires_approval') == 'on'
        is_active= request.form.get('is_active') == 'on'

        if not name: flash_msg('اسم القاعة مطلوب', 'danger'); return render_template('admin/venue_form.html', venue=None, locations=locs, form=request.form, all_users=[], assigned_users=set())

        try: cap = int(capacity) if capacity else None
        except: cap = None

        v = Venue(name=name, name_en=name_en,
                  location_id=int(loc_id) if loc_id and loc_id.isdigit() else None,
                  code=code, capacity=cap, equipment=equipment, notes=notes,
                  requires_approval=requires, is_active=is_active)
        db.add(v); db.commit()
        syslog('ADD_VENUE', f'قاعة جديدة: {name}')
        flash_msg(f'✅ تمت إضافة القاعة: {name}', 'success')
        return redirect(url_for('venues_mgmt.index'))

    return render_template('admin/venue_form.html', venue=None, locations=locs,
                           form={}, all_users=[], assigned_users=set())


@venues_mgmt_bp.route('/<int:venue_id>/edit', methods=['GET','POST'])
@login_required
@admin_required
def edit(venue_id):
    db    = get_db()
    venue = db.query(Venue).get(venue_id)
    if not venue: abort(404)
    locs  = db.query(Location).filter_by(is_active=True).order_by(Location.name).all()
    all_users = [u for u in db.query(User).filter_by(is_active=True).all()
                 if u.role and u.role.name not in ('مدير النظام','مشرف')]

    if request.method == 'POST':
        venue.name     = request.form.get('name','').strip()
        venue.name_en  = request.form.get('name_en','').strip()
        loc_id = request.form.get('location_id','')
        venue.location_id = int(loc_id) if loc_id and loc_id.isdigit() else None
        venue.code     = request.form.get('code','').strip() or None
        cap_str = request.form.get('capacity','')
        try: venue.capacity = int(cap_str) if cap_str else None
        except: pass
        venue.equipment = request.form.get('equipment','').strip()
        venue.notes     = request.form.get('notes','').strip()
        venue.requires_approval = request.form.get('requires_approval') == 'on'
        venue.is_active = request.form.get('is_active') == 'on'

        # وصول المستخدمين — مطابق لـ v54 VenueWindow access tab
        user_ids = request.form.getlist('allowed_users', type=int)
        venue.allowed_users = [u for u in all_users if u.id in user_ids]

        db.commit()
        syslog('EDIT_VENUE', f'تعديل القاعة: {venue.name}')
        flash_msg('✅ تم تحديث القاعة', 'success')
        return redirect(url_for('venues_mgmt.index'))

    assigned = {u.id for u in venue.allowed_users}
    return render_template('admin/venue_form.html',
        venue=venue, locations=locs, form={},
        all_users=all_users, assigned_users=assigned)


@venues_mgmt_bp.route('/bulk-delete', methods=['POST'])
@login_required
@admin_required
def bulk_delete_venues():
    db = get_db(); ids = request.form.getlist('ids')
    count = 0
    for vid in ids:
        try:
            v = db.query(Venue).get(int(vid))
            if v: db.delete(v); count += 1
        except: pass
    db.commit()
    flash_msg(f'✅ تم حذف {count} قاعة', 'success' if count else 'warning')
    return redirect(url_for('venues_mgmt.index'))


@venues_mgmt_bp.route('/<int:venue_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete(venue_id):
    db    = get_db()
    venue = db.query(Venue).get(venue_id)
    if not venue: abort(404)
    name = venue.name; db.delete(venue); db.commit()
    syslog('DELETE_VENUE', f'حذف القاعة: {name}')
    flash_msg(f'تم حذف القاعة: {name}', 'success')
    return redirect(url_for('venues_mgmt.index'))


# ── CSV export — مطابق لـ v54 export_venues_csv ──────────────────────────────
@venues_mgmt_bp.route('/export-csv')
@venues_mgmt_bp.route('/export-excel')
@login_required
@admin_required
def export_csv():
    from flask import session; from datetime import datetime as _dt
    db = get_db(); is_en = session.get('lang','ar') == 'en'
    headers = (['Name (AR)','Name (EN)','Location','Code','Capacity','Equipment','Notes','Requires Approval','Status'] if is_en
               else ['اسم القاعة','الاسم بالإنجليزي','الموقع','الكود','السعة','التجهيزات','ملاحظات','يحتاج موافقة','الحالة'])
    rows = []
    for v in db.query(Venue).order_by(Venue.name).all():
        rows.append([v.name, v.name_en or '', v.location.name if v.location else '',
                     v.code or '', v.capacity or '', v.equipment or '', v.notes or '',
                     ('Yes' if v.requires_approval else 'No') if is_en else ('نعم' if v.requires_approval else 'لا'),
                     ('Active' if v.is_active else 'Inactive') if is_en else ('نشطة' if v.is_active else 'موقوفة')])
    try:
        data = _make_xlsx(headers, rows, 'Venues' if is_en else 'القاعات', is_en)
        fname = f'venues_{_dt.now().strftime("%Y%m%d")}.xlsx'
        return Response(data, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition': f'attachment; filename={fname}'})
    except ImportError:
        import csv as _csv; import io as _io
        out = _io.StringIO(); _csv.writer(out).writerows([headers] + rows)
        return Response(b'\xef\xbb\xbf' + out.getvalue().encode('utf-8'), mimetype='text/csv',
                        headers={'Content-Disposition': 'attachment; filename=venues.csv'})

# ── CSV import — مطابق لـ v54 ImportCSVWindow (mode=venues) ──────────────────
@venues_mgmt_bp.route('/import-csv', methods=['GET','POST'])
@login_required
@admin_required
def import_csv():
    db = get_db()
    if request.method == 'POST':
        f = request.files.get('file')
        if not f: flash_msg('يرجى رفع ملف CSV', 'danger'); return redirect(url_for('venues_mgmt.import_csv'))

        content  = f.stream.read().decode('utf-8-sig')
        reader   = csv.DictReader(io.StringIO(content))
        skip_dup = request.form.get('skip_dup') == 'on'
        upd_dup  = request.form.get('update_dup') == 'on'

        imported = updated = skipped = errors = 0
        for row in reader:
            name     = row.get('name','').strip()
            loc_name = row.get('location_name','').strip()
            if not name: errors += 1; continue
            loc = db.query(Location).filter_by(name=loc_name).first() if loc_name else None
            if loc_name and not loc: errors += 1; continue

            active   = row.get('is_active','1').lower() not in ('0','false','no','لا')
            approval = row.get('requires_approval','1').lower() not in ('0','false','no','لا')
            try: cap = int(row.get('capacity','')) if row.get('capacity') else None
            except: cap = None

            existing = db.query(Venue).filter_by(name=name).first()
            if existing:
                if upd_dup:
                    existing.name_en = row.get('name_en', existing.name_en or '')
                    if loc: existing.location_id = loc.id
                    existing.equipment = row.get('equipment', existing.equipment or '')
                    existing.notes     = row.get('notes', existing.notes or '')
                    existing.requires_approval = approval
                    existing.is_active = active
                    if cap: existing.capacity = cap
                    updated += 1
                elif skip_dup: skipped += 1; continue
            else:
                db.add(Venue(name=name, name_en=row.get('name_en',''),
                             location_id=loc.id if loc else None,
                             code=row.get('code','').strip() or None,
                             capacity=cap, equipment=row.get('equipment',''),
                             notes=row.get('notes',''),
                             requires_approval=approval, is_active=active))
                imported += 1
        db.commit()
        syslog('IMPORT_VENUES', f'استيراد: جديد={imported}, محدّث={updated}')
        flash_msg(f'✅ استيراد مكتمل — جديد: {imported}، محدّث: {updated}، متخطى: {skipped}، أخطاء: {errors}', 'success')
        return redirect(url_for('venues_mgmt.index'))

    return render_template('admin/import_csv.html', mode='venues',
                           fields=['name','name_en','location_name','code','capacity','equipment','notes','requires_approval','is_active'])
