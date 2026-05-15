"""
مسارات الحجوزات — مطابق حرفياً لـ v54
"""
import csv, io, hashlib
from datetime import datetime, timedelta
from utils.flash_helper import flash_msg
from flask import (Blueprint, render_template, redirect, url_for,
                   request, flash, jsonify, abort, Response, session)
from flask_login import login_required, current_user
from sqlalchemy import or_
from models.database import (Reservation, Venue, Location, BlockedPeriod,
                              Approval, SystemLog, Contact, BookingContact,
                              ChecklistItem, Rating, Attachment, User)
from utils.helpers import get_db, get_permissions, syslog, paginate

reservations_bp = Blueprint('reservations', __name__, url_prefix='/reservations')

STATUS_AR = {
    'pending':   'معلق',
    'approved':  'موافق عليه',
    'rejected':  'مرفوض',
    'cancelled': 'ملغي',
    'completed': 'مكتمل',
}
STATUS_EN = {
    'pending':   'Pending',
    'approved':  'Approved',
    'rejected':  'Rejected',
    'cancelled': 'Cancelled',
    'completed': 'Completed',
}
STATUS_CLS = {
    'pending':   'warning',
    'approved':  'success',
    'rejected':  'danger',
    'cancelled': 'secondary',
    'completed': 'info',
}

def _book_num(db):
    c = db.query(Reservation).count() + 1
    return f"STAP-{datetime.now().strftime('%Y%m%d%H%M%S')}-{c:04d}"

def _conflict(db, venue_id, start, end, exclude_id=None):
    q = db.query(Reservation).filter(
        Reservation.venue_id == venue_id,
        Reservation.status.in_(['pending','approved']),
        Reservation.start_time < end,
        Reservation.end_time   > start,
    )
    if exclude_id:
        q = q.filter(Reservation.id != exclude_id)
    c = q.first()
    if c:
        return f"{c.booking_number} ({c.start_time.strftime('%H:%M')}-{c.end_time.strftime('%H:%M')})"
    blk = db.query(BlockedPeriod).filter(
        BlockedPeriod.venue_id == venue_id,
        BlockedPeriod.start_time < end,
        BlockedPeriod.end_time   > start,
    ).first()
    if blk:
        return f"فترة محظورة: {blk.reason or 'صيانة'}"
    return None


# ── index ─────────────────────────────────────────────────────────────────────
@reservations_bp.route('/')
@login_required
def index():
    db    = get_db()
    perms = get_permissions()
    page  = request.args.get('page', 1, type=int)

    q = db.query(Reservation)
    if not perms.is_admin_or_manager():
        q = q.filter_by(user_id=current_user.id)

    status   = request.args.get('status', '')
    venue_id = request.args.get('venue_id', '')
    search   = request.args.get('q', '')

    if status:   q = q.filter(Reservation.status == status)
    if venue_id and venue_id.isdigit():
        q = q.filter(Reservation.venue_id == int(venue_id))
    if search:
        q = q.filter(or_(
            Reservation.title.ilike(f'%{search}%'),
            Reservation.booking_number.ilike(f'%{search}%')
        ))

    q = q.order_by(Reservation.created_at.desc())
    items, total, total_pages = paginate(q, page, 20)

    # إحصاءات مثل v54 update_reservation_stats
    base = db.query(Reservation)
    if not perms.is_admin_or_manager():
        base = base.filter_by(user_id=current_user.id)
    stats = {
        'total':     base.count(),
        'pending':   base.filter_by(status='pending').count(),
        'approved':  base.filter_by(status='approved').count(),
        'rejected':  base.filter_by(status='rejected').count(),
        'cancelled': base.filter_by(status='cancelled').count(),
    }

    venues = db.query(Venue).filter_by(is_active=True).all()

    # إشعارات الحجوزات القريبة (24 ساعة) — مثل check_upcoming_bookings
    now      = datetime.now()
    tomorrow = now + timedelta(days=1)
    upcoming = db.query(Reservation).filter(
        Reservation.user_id   == current_user.id,
        Reservation.start_time >= now,
        Reservation.start_time <= tomorrow,
        Reservation.status     == 'approved',
    ).all()

    return render_template('reservations/index.html',
        reservations=items, total=total, page=page, total_pages=total_pages,
        stats=stats, venues=venues,
        status_ar=STATUS_AR, status_en=STATUS_EN, status_cls=STATUS_CLS,
        upcoming=upcoming,
        filters={'status': status, 'venue_id': venue_id, 'q': search})


# ── new ───────────────────────────────────────────────────────────────────────
@reservations_bp.route('/new', methods=['GET','POST'])
@login_required
def new():
    db    = get_db()
    perms = get_permissions()
    if not perms.can('reservations_add'):
        flash_msg('ليس لديك صلاحية إنشاء حجز', 'danger')
        return redirect(url_for('reservations.index'))

    venues = perms.get_allowed_venues() if hasattr(perms, 'get_allowed_venues') else \
             db.query(Venue).filter_by(is_active=True).all()

    if request.method == 'POST':
        title     = request.form.get('title','').strip()
        venue_id  = request.form.get('venue_id','')
        start_str = request.form.get('start_time','')
        end_str   = request.form.get('end_time','')
        notes     = request.form.get('notes','').strip()

        errors = []
        if not title:    errors.append(_('عنوان الحجز مطلوب') if False else ('Title is required' if session.get('lang')=='en' else 'عنوان الحجز مطلوب'))
        if not venue_id: errors.append('Please select a venue' if session.get('lang')=='en' else 'يرجى اختيار القاعة')

        start_dt = end_dt = None
        try:
            start_dt = datetime.fromisoformat(start_str)
            end_dt   = datetime.fromisoformat(end_str)
            if start_dt < datetime.now():
                errors.append('⚠️ Cannot create a booking in the past' if session.get('lang')=='en' else '⚠️ لا يمكن إنشاء حجز بتاريخ في الماضي')
            if end_dt <= start_dt:
                errors.append('End time must be after start time' if session.get('lang')=='en' else 'وقت الانتهاء يجب أن يكون بعد وقت البداية')
        except:
            errors.append('Invalid time format' if session.get('lang')=='en' else 'صيغة الوقت غير صحيحة')

        if start_dt and end_dt and venue_id and venue_id.isdigit():
            c = _conflict(db, int(venue_id), start_dt, end_dt)
            if c: errors.append(f'تعارض: {c}')

        if errors:
            for e in errors: flash(e, 'danger')
            from models.database import Contact as _Ct2
            _contacts = get_db().query(_Ct2).order_by(_Ct2.first_name).all()
            _users = db.query(User).filter_by(is_active=True).order_by(User.full_name).all()
            return render_template('reservations/new.html',
                                   venues=venues, form=request.form, now=datetime.now(), contacts=_contacts, all_users=_users)

        venue = db.query(Venue).get(int(venue_id))
        req_emp_id = request.form.get('requested_employee_id', '').strip()
        req_emp_id = int(req_emp_id) if req_emp_id and req_emp_id.isdigit() else None
        req_by_email = request.form.get('requested_by_email', '').strip()
        cc_emails    = request.form.get('cc_emails', '').strip()
        full_notes = notes
        if req_by_email:
            full_notes = (full_notes + '\n' if full_notes else '') + f'[on_behalf:{req_by_email}]'
        if cc_emails:
            full_notes = (full_notes + '\n' if full_notes else '') + f'[cc_emails:{cc_emails}]'
        res   = Reservation(
            booking_number       = _book_num(db),
            title                = title,
            user_id              = current_user.id,
            venue_id             = int(venue_id),
            start_time           = start_dt,
            end_time             = end_dt,
            requester_notes      = full_notes,
            status               = 'pending' if (venue and venue.requires_approval) else 'approved',
            requested_employee_id= req_emp_id,
        )
        db.add(res)
        db.commit()
        syslog('NEW_RESERVATION', f'{res.booking_number} — {title}')
        actor = current_user.full_name or current_user.username
        _log_activity(db, res.id, 'created', f'تم إنشاء الحجز بواسطة {actor}')

        # حفظ المرفقات
        import base64
        for f in request.files.getlist('attachments'):
            if f and f.filename:
                try:
                    data = base64.b64encode(f.read()).decode('utf-8')
                    att  = Attachment(reservation_id=res.id, filename=f.filename,
                                      mimetype=f.content_type or 'application/octet-stream',
                                      filedata=data, uploaded_by=current_user.id)
                    db.add(att)
                except: pass
        db.commit()

        # Save selected contacts & send them invitation emails
        contact_ids = request.form.getlist('contact_ids')
        if contact_ids:
            from models.database import BookingContact, Contact as _BkCt
            from utils.email_helper import send_invitation
            personal_msg = request.form.get('personal_message', '').strip()
            for cid in contact_ids:
                if cid.isdigit():
                    bc = BookingContact(booking_id=res.id, contact_id=int(cid))
                    db.add(bc)
                    # إرسال دعوة لكل جهة اتصال محددة
                    try:
                        ct = db.query(_BkCt).get(int(cid))
                        if ct and ct.email:
                            send_invitation(ct, res, personal_msg)
                    except Exception as _e:
                        print(f'Contact invite error: {_e}')
            db.commit()

        # Send notification to selected users (notify_user_ids)
        notify_user_ids = request.form.getlist('notify_user_ids')
        if notify_user_ids:
            from utils.email_helper import send_employee_reservation_notice
            for uid in notify_user_ids:
                if uid.isdigit():
                    try:
                        nu = db.query(User).get(int(uid))
                        if nu and nu.email:
                            send_employee_reservation_notice(nu, res, current_user)
                    except Exception as _e:
                        print(f'User notify error: {_e}')

        # إرسال بريد إشعار — مثل NewReservationWindow.save
        try:
            from utils.email_helper import send_booking_request
            send_booking_request(current_user, res)
        except: pass

        # إشعار الموظف المطلوب بشكل منفصل
        if req_emp_id:
            try:
                from utils.email_helper import send_employee_reservation_notice
                emp = db.query(User).get(req_emp_id)
                if emp and emp.email:
                    send_employee_reservation_notice(emp, res, current_user)
            except: pass

        flash_msg(f'✅ تم تقديم طلب الحجز — رقم الحجز: {res.booking_number}', 'success')
        return redirect(url_for('reservations.detail', res_id=res.id))

    from models.database import Contact as _Ct
    contacts = get_db().query(_Ct).order_by(_Ct.first_name).all()
    all_users = db.query(User).filter_by(is_active=True).order_by(User.full_name).all()
    return render_template('reservations/new.html', venues=venues, form={}, now=datetime.now(), contacts=contacts, all_users=all_users)


# ── detail ────────────────────────────────────────────────────────────────────
@reservations_bp.route('/<int:res_id>')
@login_required
def detail(res_id):
    db    = get_db()
    perms = get_permissions()
    res   = db.query(Reservation).get(res_id)
    if not res: abort(404)
    if res.user_id != current_user.id and not perms.is_admin_or_manager():
        abort(403)

    checklist_items = db.query(ChecklistItem).filter_by(
        reservation_id=res_id).order_by(ChecklistItem.order_index).all()
    rating = db.query(Rating).filter_by(
        reservation_id=res_id, user_id=current_user.id).first()
    contacts = db.query(BookingContact).filter_by(booking_id=res_id).all()

    attachments = db.query(Attachment).filter_by(reservation_id=res_id).all()

    # Load comments and activity log
    from models.database import ReservationComment, ReservationLog
    res_comments = db.query(ReservationComment)\
        .filter_by(reservation_id=res_id)\
        .order_by(ReservationComment.created_at.asc()).all()
    res_logs = db.query(ReservationLog)\
        .filter_by(reservation_id=res_id)\
        .order_by(ReservationLog.created_at.desc()).all()

    return render_template('reservations/detail.html',
        res=res, perms=perms,
        status_ar=STATUS_AR, status_en=STATUS_EN, status_cls=STATUS_CLS,
        checklist_items=checklist_items, rating=rating,
        booking_contacts=contacts,
        attachments=attachments,
        res_comments=res_comments,
        res_logs=res_logs)


# ── edit pending — مطابق لـ v54 edit_pending_reservation ─────────────────────
@reservations_bp.route('/<int:res_id>/edit', methods=['GET','POST'])
@login_required
def edit(res_id):
    db    = get_db()
    perms = get_permissions()
    res   = db.query(Reservation).get(res_id)
    if not res: abort(404)

    # يمكن تعديل pending/approved فقط (مثل v54)
    if res.status not in ('pending','approved'):
        _STATUS_EN = {'pending':'Pending','approved':'Approved','rejected':'Rejected','cancelled':'Cancelled','completed':'Completed'}
        if session.get('lang','ar') == 'en':
            flash('Cannot edit this booking — current status: ' + _STATUS_EN.get(res.status, res.status), 'warning')
        else:
            flash('لا يمكن تعديل هذا الحجز — الحالة الحالية: ' + STATUS_AR.get(res.status,''), 'warning')
        return redirect(url_for('reservations.detail', res_id=res_id))

    if res.user_id != current_user.id and not perms.is_admin_or_manager():
        abort(403)

    venues    = db.query(Venue).filter_by(is_active=True).all()
    employees = db.query(User).filter_by(is_active=True).order_by(User.full_name).all()

    if request.method == 'POST':
        title      = request.form.get('title','').strip()
        venue_id   = request.form.get('venue_id','')
        start_str  = request.form.get('start_time','')
        end_str    = request.form.get('end_time','')
        notes      = request.form.get('notes','').strip()
        req_emp_id = request.form.get('requested_employee_id','').strip()

        errors = []
        start_dt = end_dt = None
        try:
            start_dt = datetime.fromisoformat(start_str)
            end_dt   = datetime.fromisoformat(end_str)
            if end_dt <= start_dt: errors.append('End time must be after start time' if session.get('lang')=='en' else 'وقت الانتهاء يجب أن يكون بعد وقت البداية')
        except:
            errors.append('Invalid time format' if session.get('lang')=='en' else 'صيغة الوقت غير صحيحة')

        if start_dt and end_dt and venue_id and venue_id.isdigit():
            c = _conflict(db, int(venue_id), start_dt, end_dt, exclude_id=res_id)
            if c: errors.append(f'تعارض: {c}')

        if errors:
            for e in errors: flash(e, 'danger')
            return render_template('reservations/edit.html',
                                   res=res, venues=venues)

        # إذا كان موافقاً عليه يرجع لمعلق (مثل v54)
        if res.status == 'approved':
            res.status = 'pending'

        res.title           = title
        res.venue_id        = int(venue_id) if venue_id.isdigit() else res.venue_id
        res.start_time      = start_dt
        res.end_time        = end_dt
        res.requester_notes = notes
        # Update requested employee
        old_emp_id = res.requested_employee_id
        new_emp_id = int(req_emp_id) if req_emp_id.isdigit() else None
        res.requested_employee_id = new_emp_id
        db.commit()
        syslog('EDIT_RESERVATION', f'{res.booking_number} — تم التعديل')
        _log_activity(db, res.id, 'edited', f'تم التعديل بواسطة {current_user.full_name or current_user.username}')
        # Notify new employee if changed
        if new_emp_id and new_emp_id != old_emp_id:
            try:
                from utils.email_helper import send_employee_reservation_notice
                emp = db.query(User).get(new_emp_id)
                if emp and emp.email:
                    send_employee_reservation_notice(emp, res, current_user)
            except Exception as e:
                print(f'Employee notify error: {e}')
        flash_msg('✅ تم حفظ التعديلات' + (' — تم إرجاع الحجز لحالة معلق' if res.status == 'pending' else ''), 'success')
        return redirect(url_for('reservations.detail', res_id=res_id))

    return render_template('reservations/edit.html', res=res, venues=venues, employees=employees)


# ── approve / reject / cancel — مطابق لـ v54 ─────────────────────────────────
@reservations_bp.route('/<int:res_id>/approve', methods=['POST'])
@login_required
def approve(res_id):
    db    = get_db()
    perms = get_permissions()
    if not perms.can_action('reservations_approve','approve'): abort(403)
    res = db.query(Reservation).get(res_id)
    if not res: abort(404)
    if res.status != 'pending':
        flash_msg('يمكن الموافقة فقط على الحجوزات المعلقة', 'warning')
        return redirect(url_for('reservations.detail', res_id=res_id))

    res.status        = 'approved'
    res.approver_id   = current_user.id
    res.approval_date = datetime.now()
    res.approver_notes= request.form.get('note','')
    appr = Approval(reservation_id=res.id, approver_id=current_user.id,
                    status='approved', comments=request.form.get('note',''))
    db.add(appr)
    db.commit()
    syslog('APPROVE_RESERVATION', f'{res.booking_number}')
    actor = current_user.full_name or current_user.username
    _log_activity(db, res.id, 'approved', f'تمت الموافقة بواسطة {actor}')

    # بريد موافقة مثل v54
    try:
        from utils.email_helper import send_booking_approved, push_notification
        send_booking_approved(res)
        # إرسال للضيف الخارجي إذا كان الحجز عاماً
        import re as _re
        notes_raw = getattr(res, 'requester_notes', '') or ''
        guest_email_m = _re.search(r'\[guest_email:([^\]]+)\]', notes_raw)
        guest_name_m  = _re.search(r'\[guest_name:([^\]]+)\]', notes_raw)
        if guest_email_m and not res.user_id:
            from utils.email_helper import _send, _html_wrapper, _info_row
            ge = guest_email_m.group(1)
            gn = guest_name_m.group(1) if guest_name_m else ge
            venue = db.query(Venue).filter_by(id=res.venue_id).first() if res.venue_id else None
            subj = f'✅ تمت الموافقة على طلب الحجز — {res.booking_number}'
            content = f"""
<div style="padding:16px;background:#f0f9f0;border-radius:10px;margin-bottom:16px;text-align:center">
  <div style="font-size:32px">✅</div>
  <h3 style="color:#1B5E20;margin:8px 0">تمت الموافقة على طلب الحجز</h3>
</div>
<p>مرحباً {gn}،<br>يسعدنا إبلاغك بأنه تمت الموافقة على طلب الحجز.</p>
<table width="100%" cellpadding="0" cellspacing="0" style="border:1px solid #C8E6C9;border-radius:10px;overflow:hidden">
  {_info_row('📋 رقم الطلب', f'<strong style="color:#0C67EC">{res.booking_number}</strong>', '#F1F8E9')}
  {_info_row('📌 العنوان', res.title or '', '#FAFFFE')}
  {_info_row('🏢 القاعة', venue.name if venue else '', '#F1F8E9')}
  {_info_row('🕐 البداية', res.start_time.strftime('%Y-%m-%d %H:%M') if res.start_time else '', '#FAFFFE')}
  {_info_row('🕐 الانتهاء', res.end_time.strftime('%Y-%m-%d %H:%M') if res.end_time else '', '#F1F8E9')}
  {_info_row('💬 ملاحظة', request.form.get('note',''), '#FAFFFE') if request.form.get('note') else ''}
</table>"""
            _send(ge, gn, subj, _html_wrapper(content, subj, 'ar'),
                  f'تمت الموافقة على الحجز {res.booking_number}', sync=False, email_type='notification')
        # Local notification
        try:
            from models.database import Notification
            lang = getattr(res.user, 'language', 'ar') if res.user else 'ar'
            push_notification(db, res.user_id,
                '✅ تمت الموافقة على حجزك',
                '✅ Your booking has been approved',
                f'الحجز {res.booking_number} — {res.title}',
                f'Booking {res.booking_number} — {res.title}',
                f'/reservations/{res.id}', lang)
        except: pass
    except: pass

    flash_msg('✅ تمت الموافقة على الحجز', 'success')
    return redirect(url_for('reservations.detail', res_id=res_id))


@reservations_bp.route('/<int:res_id>/reject', methods=['POST'])
@login_required
def reject(res_id):
    db    = get_db()
    perms = get_permissions()
    if not perms.can_action('reservations_approve','approve'): abort(403)
    res = db.query(Reservation).get(res_id)
    if not res: abort(404)
    if res.status != 'pending':
        flash_msg('يمكن رفض الحجوزات المعلقة فقط', 'warning')
        return redirect(url_for('reservations.detail', res_id=res_id))

    reason = request.form.get('reason','')
    res.status         = 'rejected'
    res.approver_id    = current_user.id
    res.approver_notes = reason
    res.approval_date  = datetime.now()
    appr = Approval(reservation_id=res.id, approver_id=current_user.id,
                    status='rejected', comments=reason)
    db.add(appr)
    db.commit()
    syslog('REJECT_RESERVATION', f'{res.booking_number} — {reason}')
    actor = current_user.full_name or current_user.username
    _log_activity(db, res.id, 'rejected', f'تم الرفض بواسطة {actor} — السبب: {reason}')

    try:
        from utils.email_helper import send_booking_rejected, push_notification
        send_booking_rejected(res, reason)
        # إرسال للضيف الخارجي إذا كان الحجز عاماً
        import re as _re
        notes_raw = getattr(res, 'requester_notes', '') or ''
        guest_email_m = _re.search(r'\[guest_email:([^\]]+)\]', notes_raw)
        guest_name_m  = _re.search(r'\[guest_name:([^\]]+)\]', notes_raw)
        if guest_email_m and not res.user_id:
            from utils.email_helper import _send, _html_wrapper, _info_row
            ge = guest_email_m.group(1)
            gn = guest_name_m.group(1) if guest_name_m else ge
            subj = f'❌ اعتذار — طلب الحجز {res.booking_number}'
            content = f"""
<div style="padding:16px;background:#fef2f2;border-radius:10px;margin-bottom:16px;text-align:center">
  <div style="font-size:32px">❌</div>
  <h3 style="color:#991B1B;margin:8px 0">اعتذار — لم تتم الموافقة على الطلب</h3>
</div>
<p>مرحباً {gn}،<br>نأسف لإبلاغك بأنه لم تتم الموافقة على طلب الحجز رقم <strong>{res.booking_number}</strong>.</p>
{f'<div style="background:#fef3c7;border-radius:8px;padding:12px 16px;margin-top:12px;color:#92400e"><strong>السبب:</strong> {reason}</div>' if reason else ''}
<p style="margin-top:16px;color:#64748b;font-size:14px">يمكنك التواصل معنا لمزيد من المعلومات.</p>"""
            _send(ge, gn, subj, _html_wrapper(content, subj, 'ar'),
                  f'اعتذار — طلب الحجز {res.booking_number}', sync=False, email_type='notification')
        # Local notification
        try:
            lang = getattr(res.user, 'language', 'ar') if res.user else 'ar'
            push_notification(db, res.user_id,
                '❌ تم رفض طلب الحجز',
                '❌ Booking request rejected',
                f'الحجز {res.booking_number} — السبب: {reason or "—"}',
                f'Booking {res.booking_number} — Reason: {reason or "—"}',
                f'/reservations/{res.id}', lang)
        except: pass
    except: pass

    flash_msg('تم رفض الحجز', 'warning')
    return redirect(url_for('reservations.detail', res_id=res_id))


@reservations_bp.route('/bulk-cancel', methods=['POST'])
@login_required
def bulk_cancel():
    db = get_db(); perms = get_permissions()
    ids = request.form.getlist('ids'); count = 0
    for rid in ids:
        try:
            res = db.query(Reservation).get(int(rid))
            if not res: continue
            if res.user_id != current_user.id and not perms.is_admin_or_manager(): continue
            db.delete(res); count += 1
        except: pass
    db.commit()
    flash_msg(f'✅ تم حذف {count} حجز', 'success' if count else 'warning')
    return redirect(url_for('reservations.index'))


@reservations_bp.route('/<int:res_id>/cancel', methods=['POST'])
@login_required
def cancel(res_id):
    db    = get_db()
    perms = get_permissions()
    res   = db.query(Reservation).get(res_id)
    if not res: abort(404)
    if res.user_id != current_user.id and not perms.is_admin_or_manager():
        abort(403)
    if res.status in ('cancelled','rejected'):
        flash_msg('الحجز ملغي أو مرفوض بالفعل', 'warning')
        return redirect(url_for('reservations.detail', res_id=res_id))

    res.status = 'cancelled'
    db.commit()
    _log_activity(db, res.id, 'cancelled', f'تم الإلغاء بواسطة {current_user.full_name or current_user.username}')
    syslog('CANCEL_RESERVATION', f'{res.booking_number}')
    # Email notification
    try:
        from utils.email_helper import send_booking_cancelled
        send_booking_cancelled(res, cancelled_by=current_user)
    except Exception as e:
        print(f'Cancel email error: {e}')
    flash_msg('تم إلغاء الحجز', 'success')
    return redirect(url_for('reservations.index'))


# ── export CSV — مطابق لـ v54 export_reservations_excel (CSV للويب) ──────────
@reservations_bp.route('/export-csv')
@reservations_bp.route('/export-excel')
@login_required
def export_csv():
    db    = get_db()
    perms = get_permissions()
    q     = db.query(Reservation)
    if not perms.is_admin_or_manager():
        q = q.filter_by(user_id=current_user.id)

    status   = request.args.get('status','')
    venue_id = request.args.get('venue_id','')
    if status:   q = q.filter(Reservation.status == status)
    if venue_id and venue_id.isdigit():
        q = q.filter(Reservation.venue_id == int(venue_id))

    # Use openpyxl for proper Excel file
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        is_en = session.get('lang', 'ar') == 'en'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Reservations'
        ws.sheet_view.rightToLeft = not is_en
        h_align = 'left' if is_en else 'right'
        if is_en:
            headers = ['Booking No.','Title','User','Venue','Start Time','End Time','Status','Notes']
        else:
            headers = ['رقم الحجز','العنوان','المستخدم','القاعة','تاريخ البدء','تاريخ الانتهاء','الحالة','ملاحظات']
        STATUS_LABEL = {'pending':'Pending','approved':'Approved','rejected':'Rejected','cancelled':'Cancelled','completed':'Completed'} if is_en else STATUS_AR
        hfill = PatternFill('solid', fgColor='0C67EC')
        hfont = Font(color='FFFFFF', bold=True)
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(1, ci, h)
            cell.fill = hfill; cell.font = hfont
            cell.alignment = Alignment(horizontal='center')
        for ri, r in enumerate(q.all(), 2):
            row = [
                r.booking_number, r.title,
                r.user.full_name if r.user else '',
                r.venue.name if r.venue else '',
                r.start_time.strftime('%Y-%m-%d %H:%M') if r.start_time else '',
                r.end_time.strftime('%Y-%m-%d %H:%M')   if r.end_time   else '',
                STATUS_LABEL.get(r.status, r.status),
                r.requester_notes or '',
            ]
            for ci, val in enumerate(row, 1):
                ws.cell(ri, ci, val).alignment = Alignment(horizontal=h_align, vertical='center')
            if ri % 2 == 0:
                for ci in range(1, 9):
                    ws.cell(ri, ci).fill = PatternFill('solid', fgColor='EEF4F5')
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)
        ws.freeze_panes = 'A2'
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        fname = f'reservations_{datetime.now().strftime("%Y%m%d")}.xlsx'
        return Response(buf.read(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={fname}'})
    except ImportError:
        # Fallback to CSV
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Ref','Title','User','Venue','Start','End','Status','Notes'])
        for r in q.all():
            writer.writerow([r.booking_number, r.title,
                r.user.full_name if r.user else '',
                r.venue.name if r.venue else '',
                r.start_time.strftime('%Y-%m-%d %H:%M') if r.start_time else '',
                r.end_time.strftime('%Y-%m-%d %H:%M')   if r.end_time   else '',
                r.status, r.requester_notes or ''])
        content = b'\xef\xbb\xbf' + output.getvalue().encode('utf-8')
        fname   = f'reservations_{datetime.now().strftime("%Y%m%d")}.csv'
        return Response(content, mimetype='text/csv',
                        headers={'Content-Disposition': f'attachment; filename={fname}'})


# ── export single reservation PDF ────────────────────────────────────────────
@reservations_bp.route('/<int:res_id>/pdf')
@login_required
def export_single_pdf(res_id):
    import io, json, os
    db    = get_db()
    perms = get_permissions()
    res   = db.query(Reservation).get(res_id)
    if not res: abort(404)
    if res.user_id != current_user.id and not perms.is_admin_or_manager(): abort(403)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer, HRFlowable, Image)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
        from datetime import date as dt_date
        from utils.pdf_helper import register_arabic_font, ar, arabic_font
        import base64

        # Register Arabic font
        register_arabic_font()
        AF  = arabic_font(bold=False)
        AFB = arabic_font(bold=True)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2.5*cm)

        # Load maintenance config for header/footer
        mcfg = {}
        try:
            p = os.path.join(os.path.dirname(__file__), '..', 'maintenance_config.json')
            if os.path.exists(p): mcfg = json.loads(open(p).read())
        except: pass

        story = []

        # ── Letterhead Header ──
        center_style = ParagraphStyle('c', fontName=AFB, fontSize=14, alignment=TA_CENTER,
                                      textColor=colors.HexColor('#0C67EC'), spaceAfter=3)
        sub_style    = ParagraphStyle('s', fontName=AF,  fontSize=11, alignment=TA_CENTER,
                                      textColor=colors.HexColor('#3D8EF5'), spaceAfter=2)
        small_style  = ParagraphStyle('sm', fontName=AF, fontSize=8, alignment=TA_CENTER,
                                      textColor=colors.grey, spaceAfter=6)

        # Header image (dedicated header_img_b64 or fallback to logo_b64)
        _himg_b64 = mcfg.get('header_img_b64') or mcfg.get('logo_b64')
        _himg_pos = mcfg.get('header_img_position', 'center').upper()
        _halign_map = {'LEFT': 'LEFT', 'CENTER': 'CENTER', 'RIGHT': 'RIGHT',
                       'left': 'LEFT', 'center': 'CENTER', 'right': 'RIGHT'}
        _halign = _halign_map.get(_himg_pos, 'CENTER')
        if _himg_b64:
            try:
                raw = _himg_b64
                if ',' in raw: raw = raw.split(',', 1)[1]
                img_data = base64.b64decode(raw)
                img_buf  = io.BytesIO(img_data)
                logo_img = Image(img_buf, width=3*cm, height=3*cm, kind='proportional')
                logo_img.hAlign = _halign
                story.append(logo_img)
                story.append(Spacer(1, 0.2*cm))
            except Exception: pass

        if mcfg.get('report_header_title'):
            story.append(Paragraph(ar(mcfg['report_header_title']), center_style))
        if mcfg.get('report_header_subtitle'):
            story.append(Paragraph(ar(mcfg['report_header_subtitle']), sub_style))
        if mcfg.get('report_header_extra'):
            story.append(Paragraph(ar(mcfg['report_header_extra']), small_style))
        if mcfg.get('report_header_title') or mcfg.get('report_header_subtitle'):
            story.append(HRFlowable(width='100%', thickness=1.5,
                                    color=colors.HexColor('#0C67EC'), spaceAfter=10))

        # ── Language & Direction ──
        is_en = session.get('lang', 'ar') == 'en'

        # ── Title ──
        title_text = f'Reservation Details — {res.booking_number}' if is_en else f'تفاصيل الحجز — {res.booking_number}'
        date_text  = f'Print Date: {dt_date.today()}' if is_en else f'تاريخ الطباعة: {dt_date.today()}'
        story.append(Paragraph(ar(title_text),
            ParagraphStyle('t', fontName=AFB, fontSize=13, textColor=colors.HexColor('#0C67EC'),
                           spaceAfter=4, alignment=TA_CENTER)))
        story.append(Paragraph(ar(date_text), small_style))
        story.append(Spacer(1, 0.4*cm))

        # ── Data table ──
        teal  = colors.HexColor('#0C67EC')
        light = colors.HexColor('#EEF4F5')
        STATUS_EN_PDF = {'pending':'Pending','approved':'Approved','rejected':'Rejected',
                         'cancelled':'Cancelled','completed':'Completed'}
        STATUS_AR_PDF = {'pending':'معلق','approved':'موافق عليه','rejected':'مرفوض',
                         'cancelled':'ملغي','completed':'مكتمل'}
        status_val = STATUS_EN_PDF.get(res.status, res.status) if is_en else ar(STATUS_AR_PDF.get(res.status, res.status))

        import re as _re
        def _auto_ar(text):
            s = str(text) if text else '—'
            return ar(s) if _re.search(r'[\u0600-\u06ff]', s) else s

        if is_en:
            # LTR: [label | value] — Arabic values still reshaped for correct rendering
            rows_data = [
                ['Booking No.',    _auto_ar(res.booking_number or '—')],
                ['Title',          _auto_ar(res.title or '—')],
                ['Requested By',   _auto_ar(res.user.full_name if res.user else '—')],
                ['Venue',          _auto_ar(res.venue.name if res.venue else '—')],
                ['Location',       _auto_ar(res.venue.location.name if res.venue and res.venue.location else '—')],
                ['Start Time',     res.start_time.strftime('%Y-%m-%d  %H:%M') if res.start_time else '—'],
                ['End Time',       res.end_time.strftime('%Y-%m-%d  %H:%M') if res.end_time else '—'],
                ['Status',         status_val],
                ['Notes',          _auto_ar(res.requester_notes or '—')],
            ]
            if res.requested_employee_id and res.requested_employee:
                rows_data.append(['Requested Employee', _auto_ar(res.requested_employee.full_name or res.requested_employee.username)])
            col_widths   = [5*cm, 11*cm]
            label_col    = 0   # label on left
            val_col      = 1
            col_align    = 'LEFT'
        else:
            # RTL: swap to [value | label] so label appears on right side of page
            rows_data = [
                [str(res.booking_number or '—'),                                                         ar('رقم الحجز')],
                [ar(str(res.title or '—')),                                                               ar('العنوان')],
                [ar(str(res.user.full_name if res.user else '—')),                                        ar('مقدم الطلب')],
                [ar(str(res.venue.name if res.venue else '—')),                                           ar('القاعة')],
                [ar(str(res.venue.location.name if res.venue and res.venue.location else '—')),           ar('الموقع')],
                [res.start_time.strftime('%Y-%m-%d  %H:%M') if res.start_time else '—',                   ar('تاريخ البدء')],
                [res.end_time.strftime('%Y-%m-%d  %H:%M') if res.end_time else '—',                       ar('تاريخ الانتهاء')],
                [status_val,                                                                               ar('الحالة')],
                [ar(str(res.requester_notes or '—')),                                                     ar('ملاحظات')],
            ]
            if res.requested_employee_id and res.requested_employee:
                emp_name = res.requested_employee.full_name or res.requested_employee.username
                rows_data.append([ar(str(emp_name)), ar('الموظف المطلوب')])
            col_widths   = [11*cm, 5*cm]  # value wide on left, label narrow on right
            label_col    = 1   # label on right
            val_col      = 0
            col_align    = 'RIGHT'

        t = Table(rows_data, colWidths=col_widths)
        t.setStyle(TableStyle([
            ('FONTNAME',    (0,0),          (-1,-1),        AF),
            ('FONTNAME',    (label_col,0),  (label_col,-1), AFB),
            ('FONTSIZE',    (0,0),          (-1,-1),        10),
            ('ALIGN',       (0,0),          (-1,-1),        col_align),
            ('VALIGN',      (0,0),          (-1,-1),        'MIDDLE'),
            ('BACKGROUND',  (label_col,0),  (label_col,-1), light),
            ('TEXTCOLOR',   (label_col,0),  (label_col,-1), teal),
            ('GRID',        (0,0),          (-1,-1),        0.5, colors.HexColor('#dde8ea')),
            ('PADDING',     (0,0),          (-1,-1),        8),
            ('ROWBACKGROUNDS', (0,0),       (-1,-1),        [colors.white, light]),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.5*cm))

        # ── Footer ──
        footer_text = mcfg.get('report_header_footer') or mcfg.get('report_header_title', 'STAP — Student Tracking & Appointments')
        story.append(HRFlowable(width='100%', thickness=0.5,
                                color=colors.HexColor('#9ab3b8'), spaceAfter=4))
        story.append(Paragraph(ar(footer_text),
            ParagraphStyle('f', fontName=AF, fontSize=8, alignment=TA_CENTER, textColor=colors.grey)))

        doc.build(story)
        buf.seek(0)
        return Response(buf.read(), mimetype='application/pdf',
            headers={'Content-Disposition': f'attachment;filename=reservation_{res.booking_number}.pdf'})
    except ImportError:
        flash_msg('reportlab غير مثبت', 'danger')
        return redirect(url_for('reservations.detail', res_id=res_id))


# ── export single reservation Excel ──────────────────────────────────────────
@reservations_bp.route('/<int:res_id>/excel')
@login_required
def export_single_excel(res_id):
    import io
    db    = get_db()
    perms = get_permissions()
    res   = db.query(Reservation).get(res_id)
    if not res: abort(404)
    if res.user_id != current_user.id and not perms.is_admin_or_manager(): abort(403)

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        is_en = session.get('lang', 'ar') == 'en'
        h_align = 'left' if is_en else 'right'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Reservation'
        ws.sheet_view.rightToLeft = not is_en

        teal_fill = PatternFill('solid', fgColor='0C67EC')
        light_fill = PatternFill('solid', fgColor='EEF4F5')
        teal_font  = Font(color='FFFFFF', bold=True, size=11)
        bold_font  = Font(bold=True, size=10)
        thin_border = Border(
            left=Side(style='thin',color='9ab3b8'),
            right=Side(style='thin',color='9ab3b8'),
            top=Side(style='thin',color='9ab3b8'),
            bottom=Side(style='thin',color='9ab3b8'))

        status_val = {'pending':'Pending','approved':'Approved','rejected':'Rejected','cancelled':'Cancelled','completed':'Completed'}.get(res.status, res.status) if is_en else STATUS_AR.get(res.status, res.status)

        if is_en:
            rows = [
                ('Booking No.',        res.booking_number),
                ('Title',              res.title),
                ('Requested By',       res.user.full_name if res.user else '—'),
                ('Venue',              res.venue.name if res.venue else '—'),
                ('Location',           res.venue.location.name if res.venue and res.venue.location else '—'),
                ('Start Time',         str(res.start_time) if res.start_time else '—'),
                ('End Time',           str(res.end_time) if res.end_time else '—'),
                ('Status',             status_val),
                ('Notes',              res.requester_notes or '—'),
            ]
            title_val = f'Reservation Details — {res.booking_number}'
        else:
            rows = [
                ('رقم الحجز',    res.booking_number),
                ('العنوان',       res.title),
                ('المستخدم',     res.user.full_name if res.user else '—'),
                ('القاعة',        res.venue.name if res.venue else '—'),
                ('الموقع',        res.venue.location.name if res.venue and res.venue.location else '—'),
                ('تاريخ البدء',  str(res.start_time) if res.start_time else '—'),
                ('تاريخ الانتهاء', str(res.end_time) if res.end_time else '—'),
                ('الحالة',        status_val),
                ('ملاحظات',       res.requester_notes or '—'),
            ]
            title_val = f'تفاصيل الحجز — {res.booking_number}'

        # Add requested employee row if set
        if res.requested_employee_id and res.requested_employee:
            emp_name = res.requested_employee.full_name or res.requested_employee.username
            rows.append(('Requested Employee' if is_en else 'الموظف المطلوب', emp_name))

        # Title row
        ws.merge_cells('A1:B1')
        tc = ws['A1']
        tc.value = title_val
        tc.fill = teal_fill; tc.font = teal_font
        tc.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        for ri, (label, value) in enumerate(rows, 2):
            ca = ws.cell(ri, 1, label)
            cb = ws.cell(ri, 2, value)
            ca.font = bold_font
            ca.fill = light_fill
            ca.alignment = Alignment(horizontal=h_align, vertical='center')
            cb.alignment = Alignment(horizontal=h_align, vertical='center')
            for cell in (ca, cb):
                cell.border = thin_border
            ws.row_dimensions[ri].height = 20

        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 42

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        fname = f'reservation_{res.booking_number}.xlsx'
        return Response(buf.read(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment;filename={fname}'})
    except ImportError:
        flash_msg('openpyxl غير مثبت', 'danger')
        return redirect(url_for('reservations.detail', res_id=res_id))


# ── export PDF ────────────────────────────────────────────────────────────────
@reservations_bp.route('/export-pdf')
@login_required
def export_pdf():
    import io, json, os
    db    = get_db()
    perms = get_permissions()
    q     = db.query(Reservation)
    if not perms.is_admin_or_manager():
        q = q.filter_by(user_id=current_user.id)

    status   = request.args.get('status','')
    venue_id = request.args.get('venue_id','')
    if status:   q = q.filter(Reservation.status == status)
    if venue_id and venue_id.isdigit():
        q = q.filter(Reservation.venue_id == int(venue_id))
    items = q.order_by(Reservation.created_at.desc()).all()

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
        from datetime import date as dt_date
        from utils.pdf_helper import register_arabic_font, ar, arabic_font

        register_arabic_font()
        AF  = arabic_font(bold=False)
        AFB = arabic_font(bold=True)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                rightMargin=1.5*cm, leftMargin=1.5*cm,
                                topMargin=1.5*cm, bottomMargin=1.5*cm)
        from reportlab.platypus import Image as RLImage
        styles = getSampleStyleSheet()
        story  = []
        is_en  = session.get('lang', 'ar') == 'en'

        # Header from maintenance config
        mcfg = {}
        try:
            p = os.path.join(os.path.dirname(__file__), '..', 'maintenance_config.json')
            if os.path.exists(p): mcfg = json.loads(open(p).read())
        except: pass

        # ── Logo ──────────────────────────────────────────────────────────────
        import base64 as _b64
        _logo = mcfg.get('header_img_b64') or mcfg.get('logo_b64')
        if _logo:
            try:
                raw = _logo.split(',',1)[1] if ',' in _logo else _logo
                img_buf = io.BytesIO(_b64.b64decode(raw))
                li = RLImage(img_buf, width=2.5*cm, height=2.5*cm, kind='proportional')
                li.hAlign = 'CENTER'
                story.append(li)
                story.append(Spacer(1, 0.15*cm))
            except: pass

        center_s = ParagraphStyle('ch', fontName=AFB, fontSize=13, textColor=colors.HexColor('#0C67EC'), alignment=TA_CENTER, spaceAfter=2)
        sub_s    = ParagraphStyle('cs', fontName=AF,  fontSize=10, textColor=colors.HexColor('#3D8EF5'), alignment=TA_CENTER, spaceAfter=4)

        if mcfg.get('report_header_title'):
            story.append(Paragraph(ar(mcfg['report_header_title']), center_s))
        if mcfg.get('report_header_subtitle'):
            story.append(Paragraph(ar(mcfg['report_header_subtitle']), sub_s))
        if mcfg.get('report_header_title') or mcfg.get('report_header_subtitle'):
            story.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#0C67EC'), spaceAfter=6))

        report_title = f'Reservations Report — STAP  |  {dt_date.today()}' if is_en else f'تقرير الحجوزات — STAP  |  {dt_date.today()}'
        story.append(Paragraph(ar(report_title),
            ParagraphStyle('t', fontName=AFB, fontSize=12, textColor=colors.HexColor('#0C67EC'), spaceAfter=10, alignment=TA_CENTER)))

        STATUS_LABEL = {'pending':'Pending','approved':'Approved','rejected':'Rejected','cancelled':'Cancelled','completed':'Completed'} if is_en else \
                       {'pending':'معلق','approved':'موافق عليه','rejected':'مرفوض','cancelled':'ملغي','completed':'مكتمل'}

        if is_en:
            headers    = ['Ref', 'Title', 'User', 'Venue', 'Start', 'End', 'Status']
            col_widths = [3.5*cm, 5*cm, 3.5*cm, 3.5*cm, 4*cm, 2*cm, 2.5*cm]
            row_align  = 'LEFT'
        else:
            # RTL: reverse column order so reading flows right→left
            headers    = [ar('الحالة'), ar('الانتهاء'), ar('البدء'), ar('القاعة'), ar('المستخدم'), ar('العنوان'), ar('الرقم')]
            col_widths = [2.5*cm, 2*cm, 4*cm, 3.5*cm, 3.5*cm, 5*cm, 3.5*cm]
            row_align  = 'RIGHT'

        data = [headers]
        for r in items:
            if is_en:
                data.append([
                    r.booking_number or '',
                    (r.title or '')[:30],
                    r.user.full_name if r.user else '',
                    r.venue.name if r.venue else '',
                    r.start_time.strftime('%Y-%m-%d %H:%M') if r.start_time else '',
                    r.end_time.strftime('%H:%M') if r.end_time else '',
                    STATUS_LABEL.get(r.status, r.status),
                ])
            else:
                # reversed order for RTL
                data.append([
                    ar(STATUS_LABEL.get(r.status, r.status)),
                    r.end_time.strftime('%H:%M') if r.end_time else '',
                    r.start_time.strftime('%Y-%m-%d %H:%M') if r.start_time else '',
                    ar(r.venue.name if r.venue else ''),
                    ar(r.user.full_name if r.user else ''),
                    ar((r.title or '')[:30]),
                    r.booking_number or '',
                ])

        t = Table(data, repeatRows=1, colWidths=col_widths)
        t.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (-1,0),  colors.HexColor('#0C67EC')),
            ('TEXTCOLOR',     (0,0), (-1,0),  colors.white),
            ('FONTNAME',      (0,0), (-1,-1), AF),
            ('FONTNAME',      (0,0), (-1,0),  AFB),
            ('FONTSIZE',      (0,0), (-1,-1), 8),
            ('ALIGN',         (0,0), (-1,-1), row_align),
            ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
            ('GRID',          (0,0), (-1,-1), 0.4, colors.HexColor('#dee2e6')),
            ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.white, colors.HexColor('#EEF4F5')]),
            ('PADDING',       (0,0), (-1,-1), 5),
            ('TOPPADDING',    (0,0), (-1,0),  8),
        ]))
        story.append(t)
        doc.build(story)
        buf.seek(0)
        fname = f'reservations_{dt_date.today()}.pdf'
        return Response(buf.read(), mimetype='application/pdf',
                        headers={'Content-Disposition': f'attachment;filename={fname}'})
    except ImportError:
        flash_msg('reportlab غير مثبت', 'danger')
        return redirect(url_for('reservations.index'))


# ── send invitations — مطابق لـ v54 SendInvitationsWindow ────────────────────
@reservations_bp.route('/<int:res_id>/invite', methods=['GET','POST'])
@login_required
def invite(res_id):
    db    = get_db()
    perms = get_permissions()
    res   = db.query(Reservation).get(res_id)
    if not res: abort(404)
    if res.user_id != current_user.id and not perms.is_admin_or_manager():
        abort(403)

    contacts = db.query(Contact).filter_by(created_by=current_user.id).all()
    from models.database import ContactGroup
    groups = db.query(ContactGroup).filter_by(is_active=True).order_by(ContactGroup.name).all()

    if request.method == 'POST':
        selected_ids = request.form.getlist('contact_ids', type=int)
        # If a group was selected, add its contacts
        group_id = request.form.get('group_id', type=int)
        if group_id:
            from models.database import ContactGroup
            grp = db.query(ContactGroup).get(group_id)
            if grp:
                group_contact_ids = [c.id for c in grp.contacts]
                selected_ids = list(set(selected_ids + group_contact_ids))
        # Read message — prefer HTML from Quill, fallback to plain
        editor_mode  = request.form.get('editor_mode', 'plain')
        message_body = request.form.get('message', '').strip()      # hidden HTML field
        if not message_body or editor_mode == 'plain':
            message_body = request.form.get('message_plain', '').strip()
        if not message_body:
            message_body = request.form.get('message', '').strip()  # legacy
        if not selected_ids:
            flash_msg('اختر مدعويين على الأقل', 'warning')
            return redirect(url_for('reservations.invite', res_id=res_id))

        success = 0
        for cid in selected_ids:
            contact = db.query(Contact).get(cid)
            if not contact: continue
            try:
                personal_msg = message_body.replace('[NAME]', contact.first_name or '')
                from utils.email_helper import send_invitation
                send_invitation(contact, res, personal_msg)

                bc = db.query(BookingContact).filter_by(
                    booking_id=res_id, contact_id=cid).first()
                if not bc:
                    bc = BookingContact(booking_id=res_id, contact_id=cid,
                                       invitation_sent=True, sent_at=datetime.now())
                    db.add(bc)
                else:
                    bc.invitation_sent = True
                    bc.sent_at = datetime.now()
                success += 1
            except Exception as e:
                flash_msg(f'فشل إرسال إلى {contact.email}: {e}', 'warning')

        db.commit()
        _lang = session.get('lang', 'ar')
        syslog('SEND_INVITATIONS', f'{res.booking_number} — {success} دعوة')
        flash_msg(f'✅ {success} {"invitations sent" if _lang=="en" else "دعوة تم إرسالها"}', 'success')
        return redirect(url_for('reservations.detail', res_id=res_id))

    _lang = session.get('lang', 'ar')
    venue_name = res.venue.name if res.venue else '—'
    start_date = res.start_time.strftime('%Y-%m-%d') if res.start_time else ''
    start_time = res.start_time.strftime('%H:%M') if res.start_time else ''

    if _lang == 'en':
        default_msg = f"""Dear [NAME],

We are pleased to invite you to attend "{res.title}" at:
🏢 Venue: {venue_name}
📅 Date: {start_date}
⏰ Time: {start_time}

We look forward to seeing you,
STAP Team"""
    else:
        default_msg = f"""عزيزي/عزيزتي [NAME]،

يسرنا دعوتكم لحضور "{res.title}" في:
🏢 القاعة: {venue_name}
📅 التاريخ: {start_date}
⏰ الوقت: {start_time}

نتطلع لرؤيتكم،
فريق STAP"""

    return render_template('reservations/invite.html',
        res=res, contacts=contacts, groups=groups, default_msg=default_msg)


# ── notifications API — upcoming bookings ─────────────────────────────────────
@reservations_bp.route('/api/upcoming')
@login_required
def api_upcoming():
    db       = get_db()
    now      = datetime.now()
    tomorrow = now + timedelta(days=1)
    upcoming = db.query(Reservation).filter(
        Reservation.user_id    == current_user.id,
        Reservation.start_time >= now,
        Reservation.start_time <= tomorrow,
        Reservation.status     == 'approved',
    ).all()
    return jsonify([{
        'id':       r.id,
        'title':    r.title,
        'venue':    r.venue.name if r.venue else '—',
        'start':    r.start_time.strftime('%Y-%m-%d %H:%M') if r.start_time else '',
        'booking_number': r.booking_number,
    } for r in upcoming])


# ── availability AJAX ─────────────────────────────────────────────────────────
@reservations_bp.route('/api/availability')
@login_required
def availability():
    db       = get_db()
    venue_id = request.args.get('venue_id', type=int)
    date_str = request.args.get('date','')
    if not venue_id or not date_str:
        return jsonify([])
    try:
        d = datetime.strptime(date_str,'%Y-%m-%d').date()
    except:
        return jsonify([])
    rsvs = db.query(Reservation).filter(
        Reservation.venue_id == venue_id,
        Reservation.status.in_(['pending','approved']),
    ).all()
    return jsonify([{
        'id': r.id, 'title': r.title,
        'start': r.start_time.strftime('%H:%M') if r.start_time else '',
        'end':   r.end_time.strftime('%H:%M')   if r.end_time   else '',
        'status': r.status,
    } for r in rsvs if r.start_time and r.start_time.date() == d])


# ── Download attachment ───────────────────────────────────────────────────────
@reservations_bp.route('/attachment/<int:att_id>')
@login_required
def download_attachment(att_id):
    import base64 as b64lib
    db  = get_db()
    att = db.query(Attachment).get(att_id)
    if not att: abort(404)
    res = db.query(Reservation).get(att.reservation_id)
    if not res: abort(404)
    perms = get_permissions()
    if res.user_id != current_user.id and not perms.is_admin_or_manager(): abort(403)
    data = b64lib.b64decode(att.filedata)
    return Response(data, mimetype=att.mimetype,
                    headers={'Content-Disposition': f'attachment; filename="{att.filename}"'})


# ── Reactivate — إعادة تفعيل حجز ملغى أو مرفوض ───────────────────────────────
@reservations_bp.route('/<int:res_id>/reactivate', methods=['POST'])
@login_required
def reactivate(res_id):
    db    = get_db()
    perms = get_permissions()
    res   = db.query(Reservation).get(res_id)
    if not res: abort(404)
    if not perms.is_admin_or_manager(): abort(403)
    if res.status not in ('cancelled', 'rejected'):
        flash_msg('يمكن إعادة تفعيل الحجوزات الملغية أو المرفوضة فقط', 'warning')
        return redirect(url_for('reservations.detail', res_id=res_id))
    res.status = 'pending'
    db.commit()
    syslog('REACTIVATE', f'إعادة تفعيل حجز #{res.booking_number}')
    flash_msg('تم إعادة تفعيل الحجز وإرجاعه للمراجعة', 'success')
    return redirect(url_for('reservations.detail', res_id=res_id))


# ══════════════════════════════════════════════════════════════════════════════
# RESERVATION COMMENTS & ACTIVITY LOG
# ══════════════════════════════════════════════════════════════════════════════

def _log_activity(db, res_id, action, description, user_id=None):
    """Add entry to reservation activity log"""
    from models.database import ReservationLog
    try:
        db.add(ReservationLog(
            reservation_id=res_id,
            user_id=user_id or (current_user.id if current_user.is_authenticated else None),
            action=action,
            description=description
        ))
        db.commit()
    except Exception:
        pass


@reservations_bp.route('/<int:res_id>/comments', methods=['POST'])
@login_required
def add_comment(res_id):
    db    = get_db()
    perms = get_permissions()
    res   = db.query(Reservation).get(res_id)
    if not res: abort(404)

    # Permission check
    is_owner   = res.user_id == current_user.id
    is_manager = perms.is_admin_or_manager()
    can_comment_perm = perms.can_comment('reservations')

    if not (is_owner or is_manager or can_comment_perm):
        from utils.flash_helper import flash_msg
        flash_msg('ليس لديك صلاحية إضافة تعليقات', 'danger')
        return redirect(url_for('reservations.detail', res_id=res_id))

    content     = request.form.get('content', '').strip()
    is_internal = request.form.get('is_internal') == '1' and is_manager

    if not content:
        from utils.flash_helper import flash_msg
        flash_msg('التعليق لا يمكن أن يكون فارغاً', 'warning')
        return redirect(url_for('reservations.detail', res_id=res_id))

    from models.database import ReservationComment
    comment = ReservationComment(
        reservation_id=res_id,
        user_id=current_user.id,
        content=content,
        is_internal=is_internal
    )
    db.add(comment)
    db.commit()

    # Log the activity
    actor = current_user.full_name or current_user.username
    _log_activity(db, res_id, 'commented',
                  f'{"[داخلي] " if is_internal else ""}{actor}: {content[:80]}')

    # Notify reservation owner if commenter is not the owner
    if res.user_id != current_user.id and res.user:
        from utils.email_helper import push_notification
        lang = getattr(res.user, 'language', 'ar')
        push_notification(db, res.user_id,
            f'💬 تعليق جديد على حجزك {res.booking_number}',
            f'💬 New comment on booking {res.booking_number}',
            f'{actor}: {content[:60]}',
            f'{actor}: {content[:60]}',
            f'/reservations/{res_id}', lang)
        # Email notification
        try:
            from utils.email_helper import _send, _html_wrapper, _info_row, _user_lang
            ulang = _user_lang(res.user)
            if ulang == 'en':
                subj = f'💬 New comment on booking {res.booking_number}'
                html_content = f"""
<h2 style="color:#0C67EC;margin:0 0 12px">Hello {res.user.full_name},</h2>
<p style="color:#4a5568">{actor} added a comment on your booking:</p>
<div style="background:#f4f9ff;border-left:4px solid #0C67EC;border-radius:8px;padding:14px;margin:16px 0;color:#2d3748">
  {content}
</div>
{_info_row('📋 Booking', res.booking_number)}
{_info_row('📌 Title', res.title, '#FAFCFF')}"""
            else:
                subj = f'💬 تعليق جديد على حجزك {res.booking_number}'
                html_content = f"""
<h2 style="color:#0C67EC;margin:0 0 12px">مرحباً {res.user.full_name}،</h2>
<p style="color:#4a5568">أضاف {actor} تعليقاً على حجزك:</p>
<div style="background:#f4f9ff;border-right:4px solid #0C67EC;border-radius:8px;padding:14px;margin:16px 0;color:#2d3748">
  {content}
</div>
{_info_row('📋 رقم الحجز', res.booking_number)}
{_info_row('📌 العنوان', res.title, '#FAFCFF')}"""
            _send(res.user.email, res.user.full_name, subj,
                  _html_wrapper(html_content, subj, ulang),
                  sync=True, email_type='notification')
        except Exception:
            pass

    # Also notify admins/managers if comment from regular user
    if not is_manager and res.user_id == current_user.id:
        from models.database import User as _User, RolePermission as _RP
        try:
            managers = db.query(_User).filter(
                _User.is_active == True,
                _User.id != current_user.id
            ).join(_User.role_ref).filter(
                _User.role_ref.has(name='مدير النظام') |
                _User.role_ref.has(name='مشرف') |
                _User.role_ref.has(name='Manager') |
                _User.role_ref.has(name='Supervisor')
            ).all()
            for mgr in managers:
                lang = getattr(mgr, 'language', 'ar')
                push_notification(db, mgr.id,
                    f'💬 تعليق جديد من {actor} على حجز {res.booking_number}',
                    f'💬 New comment from {actor} on booking {res.booking_number}',
                    content[:60], content[:60],
                    f'/reservations/{res_id}', lang)
        except Exception:
            pass

    from utils.flash_helper import flash_msg
    flash_msg('تم إضافة التعليق بنجاح', 'success')
    return redirect(url_for('reservations.detail', res_id=res_id) + '#comments')


@reservations_bp.route('/<int:res_id>/comments/<int:cid>/delete', methods=['POST'])
@login_required
def delete_comment(res_id, cid):
    db    = get_db()
    perms = get_permissions()
    from models.database import ReservationComment
    comment = db.query(ReservationComment).get(cid)
    if not comment or comment.reservation_id != res_id: abort(404)
    # Only owner of comment or admin can delete
    if comment.user_id != current_user.id and not perms.is_admin_or_manager():
        abort(403)
    db.delete(comment)
    db.commit()
    from utils.flash_helper import flash_msg
    flash_msg('تم حذف التعليق', 'success')
    return redirect(url_for('reservations.detail', res_id=res_id) + '#comments')
