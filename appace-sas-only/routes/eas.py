"""
EAS - Employee Attendance System
نظام حضور وغياب الموظفين
"""
import json
import io
import csv
from datetime import datetime, date
from flask import (
    Blueprint, render_template, request, redirect, url_for,
    jsonify, abort, flash, Response
)
from sqlalchemy.orm import joinedload
from models.database import (
    SASConfig, EASGroup, EASEmployee, EASShift,
    EASShiftAssignment, EASRecord
)
from utils.helpers import admin_required, get_db
from utils.i18n import get_lang, t

eas_bp = Blueprint('eas', __name__, url_prefix='/eas')

def _get_config():
    db = get_db()
    cfg = db.query(SASConfig).first()
    if not cfg:
        abort(404)
    return cfg

def _t(ar, en):
    return en if get_lang() == 'en' else ar

# ── Dashboard ──────────────────────────────────────
@eas_bp.route('/')
@admin_required
def index():
    db = get_db()
    cfg = _get_config()
    groups = db.query(EASGroup).filter(EASGroup.config_id == cfg.id).all()
    return render_template('eas/index.html', config=cfg, groups=groups)

# ── Groups ─────────────────────────────────────────
@eas_bp.route('/groups')
@admin_required
def groups():
    db = get_db()
    cfg = _get_config()
    groups = db.query(EASGroup).filter(EASGroup.config_id == cfg.id).all()
    return render_template('eas/groups.html', config=cfg, groups=groups)

@eas_bp.route('/groups/add', methods=['POST'])
@admin_required
def group_add():
    db = get_db()
    cfg = _get_config()
    name    = request.form.get('name', '').strip()
    name_ar = request.form.get('name_ar', '').strip()
    work_days = ','.join(request.form.getlist('work_days'))
    shift_start    = request.form.get('shift_start', '07:30')
    shift_end      = request.form.get('shift_end', '15:00')
    late_tolerance = request.form.get('late_tolerance', 10, type=int)

    if not name:
        flash(_t('يرجى إدخال اسم المجموعة', 'Please enter group name'), 'danger')
        return redirect(url_for('eas.groups'))

    group = EASGroup(
        config_id=cfg.id, name=name, name_ar=name_ar,
        work_days=work_days, shift_start=shift_start,
        shift_end=shift_end, late_tolerance=late_tolerance
    )
    db.add(group)
    try:
        db.commit()
        flash(_t('تمت إضافة المجموعة', 'Group added'), 'success')
    except Exception:
        db.rollback()
        flash(_t('حدث خطأ', 'Error'), 'danger')
    return redirect(url_for('eas.groups'))

@eas_bp.route('/groups/<int:gid>/edit', methods=['POST'])
@admin_required
def group_edit(gid):
    """FIX: Edit group name / settings"""
    db = get_db()
    group = db.get(EASGroup, gid)
    if not group:
        abort(404)
    name    = request.form.get('name', '').strip()
    name_ar = request.form.get('name_ar', '').strip()
    if not name:
        flash(_t('يرجى إدخال اسم المجموعة', 'Please enter group name'), 'danger')
        return redirect(url_for('eas.groups'))
    group.name    = name
    group.name_ar = name_ar
    group.shift_start    = request.form.get('shift_start', group.shift_start)
    group.shift_end      = request.form.get('shift_end', group.shift_end)
    group.late_tolerance = request.form.get('late_tolerance', group.late_tolerance, type=int)
    group.work_days      = ','.join(request.form.getlist('work_days')) or group.work_days
    try:
        db.commit()
        flash(_t('تم تعديل المجموعة', 'Group updated'), 'success')
    except Exception:
        db.rollback()
        flash(_t('حدث خطأ', 'Error'), 'danger')
    return redirect(url_for('eas.groups'))

@eas_bp.route('/groups/<int:gid>/delete', methods=['POST'])
@admin_required
def group_delete(gid):
    db = get_db()
    group = db.get(EASGroup, gid)
    if not group: abort(404)
    db.delete(group)
    try:
        db.commit()
        flash(_t('تم الحذف', 'Deleted'), 'success')
    except Exception:
        db.rollback()
    return redirect(url_for('eas.groups'))

# ── Departments (derived from employee.department field) ───
@eas_bp.route('/groups/<int:gid>/departments')
@admin_required
def departments(gid):
    """FIX: List all unique departments in this group"""
    db = get_db()
    group = db.get(EASGroup, gid)
    if not group: abort(404)
    # Get distinct departments
    rows = db.query(EASEmployee.department).filter(
        EASEmployee.group_id == gid,
        EASEmployee.department.isnot(None),
        EASEmployee.department != '',
    ).distinct().all()
    dept_names = sorted([r[0] for r in rows if r[0]])
    return render_template('eas/departments.html', group=group, dept_names=dept_names)

@eas_bp.route('/groups/<int:gid>/departments/rename', methods=['POST'])
@admin_required
def department_rename(gid):
    """FIX: Rename a department across all employees in this group"""
    db = get_db()
    group = db.get(EASGroup, gid)
    if not group: abort(404)
    old_name = request.form.get('old_name', '').strip()
    new_name = request.form.get('new_name', '').strip()
    if not old_name or not new_name:
        flash(_t('يرجى إدخال الاسم القديم والجديد', 'Please enter old and new name'), 'danger')
        return redirect(url_for('eas.departments', gid=gid))
    updated = db.query(EASEmployee).filter(
        EASEmployee.group_id == gid,
        EASEmployee.department == old_name
    ).update({'department': new_name}, synchronize_session='fetch')
    try:
        db.commit()
        flash(_t(f'تم تغيير اسم القسم لـ {updated} موظف', f'Renamed department for {updated} employees'), 'success')
    except Exception:
        db.rollback()
        flash(_t('حدث خطأ', 'Error'), 'danger')
    return redirect(url_for('eas.departments', gid=gid))

# ── Employees ──────────────────────────────────────
@eas_bp.route('/groups/<int:gid>/employees')
@admin_required
def employees(gid):
    db = get_db()
    group = db.get(EASGroup, gid)
    if not group: abort(404)
    # FIX: support filter by department
    dept_filter = request.args.get('department', '').strip()
    q = db.query(EASEmployee).filter(EASEmployee.group_id == gid)
    if dept_filter:
        q = q.filter(EASEmployee.department == dept_filter)
    emps = q.order_by(EASEmployee.name).all()
    # FIX: get distinct departments for the filter dropdown
    dept_rows = db.query(EASEmployee.department).filter(
        EASEmployee.group_id == gid,
        EASEmployee.department.isnot(None),
        EASEmployee.department != '',
    ).distinct().all()
    dept_names = sorted([r[0] for r in dept_rows if r[0]])
    return render_template('eas/employees.html', group=group, employees=emps,
                           dept_names=dept_names, dept_filter=dept_filter)

@eas_bp.route('/groups/<int:gid>/employees/add', methods=['POST'])
@admin_required
def employee_add(gid):
    db = get_db()
    group = db.get(EASGroup, gid)
    if not group: abort(404)
    name  = request.form.get('name', '').strip()
    if not name:
        flash(_t('يرجى إدخال الاسم', 'Please enter name'), 'danger')
        return redirect(url_for('eas.employees', gid=gid))
    emp = EASEmployee(
        group_id=gid,
        name=name,
        name_en=request.form.get('name_en', '').strip(),
        email=request.form.get('email', '').strip(),
        phone=request.form.get('phone', '').strip(),
        employee_id=request.form.get('employee_id', '').strip(),
        department=request.form.get('department', '').strip(),  # FIX: save department
    )
    db.add(emp)
    try:
        db.commit()
        flash(_t('تمت الإضافة', 'Added'), 'success')
    except Exception:
        db.rollback()
    return redirect(url_for('eas.employees', gid=gid))

@eas_bp.route('/employees/import/<int:gid>', methods=['POST'])
@admin_required
def employees_import(gid):
    db = get_db()
    group = db.get(EASGroup, gid)
    if not group: abort(404)
    file = request.files.get('file')
    if not file:
        flash(_t('يرجى رفع ملف', 'Please upload a file'), 'danger')
        return redirect(url_for('eas.employees', gid=gid))
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
        added = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            name = str(row_dict.get('name') or row_dict.get('الاسم') or '').strip()
            if not name: continue
            emp = EASEmployee(
                group_id=gid,
                name=name,
                name_en=str(row_dict.get('name_en') or '').strip(),
                email=str(row_dict.get('email') or row_dict.get('البريد') or '').strip(),
                phone=str(row_dict.get('phone') or row_dict.get('الهاتف') or '').strip(),
                employee_id=str(row_dict.get('employee_id') or row_dict.get('رقم الموظف') or '').strip(),
                department=str(row_dict.get('department') or row_dict.get('القسم') or '').strip(),  # FIX
            )
            db.add(emp)
            added += 1
        db.commit()
        flash(_t(f'تم استيراد {added} موظف', f'Imported {added} employees'), 'success')
    except Exception as e:
        db.rollback()
        flash(_t(f'خطأ: {e}', f'Error: {e}'), 'danger')
    return redirect(url_for('eas.employees', gid=gid))

# ── Upload Excel attendance ────────────────────────
@eas_bp.route('/upload', methods=['GET', 'POST'])
@admin_required
def upload():
    db = get_db()
    cfg = _get_config()
    groups = db.query(EASGroup).filter(EASGroup.config_id == cfg.id).all()
    if request.method == 'GET':
        return render_template('eas/upload.html', config=cfg, groups=groups)

    file = request.files.get('file')
    group_id = request.form.get('group_id', type=int)
    if not file or not group_id:
        flash(_t('يرجى اختيار الملف والمجموعة', 'Please select file and group'), 'danger')
        return redirect(url_for('eas.upload'))

    group = db.get(EASGroup, group_id)
    if not group: abort(404)

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        # case-insensitive header → column index map (skip empty columns)
        raw_headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
        col = {h.lower(): i for i, h in enumerate(raw_headers) if h}

        def _col_idx(names, default):
            for n in names:
                if n.lower() in col:
                    return col[n.lower()]
            return default

        idx_name       = _col_idx(['name','الاسم','employee name','employee_name'], 0)
        idx_date       = _col_idx(['date','التاريخ','record_date','recorddate'], 1)
        idx_check_in   = _col_idx(['check_in','checkin','check in','وقت الدخول','دخول','time_in'], 2)
        idx_check_out  = _col_idx(['check_out','checkout','check out','وقت الخروج','خروج','time_out'], 3)
        idx_department = _col_idx(['department','dept','القسم','الإدارة','الادارة'], -1)

        def _parse_date(v):
            """Convert any date value to YYYY-MM-DD string."""
            if v is None: return ''
            # openpyxl already parsed it as datetime
            if hasattr(v, 'strftime'):
                return v.strftime('%Y-%m-%d')
            s = str(v).strip()
            if not s or s == 'None': return ''
            # Excel serial integer (e.g. 46091)
            if s.isdigit():
                from datetime import date as _d, timedelta as _td
                return (_d(1899, 12, 30) + _td(days=int(s))).strftime('%Y-%m-%d')
            # Common string formats — flexible dd/mm or mm/dd or yyyy-mm-dd
            for fmt in ('%Y-%m-%d','%d/%m/%Y','%m/%d/%Y','%Y/%m/%d',
                        '%d-%m-%Y','%m-%d-%Y','%d.%m.%Y','%Y.%m.%d'):
                try:
                    return datetime.strptime(s[:10], fmt).strftime('%Y-%m-%d')
                except ValueError:
                    continue
            return s[:10]

        added = 0
        errors = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                def cell(idx):
                    if idx < 0 or idx >= len(row): return ''
                    v = row[idx]
                    if v is None: return ''
                    if hasattr(v, 'strftime'):
                        # time-only cell
                        return v.strftime('%H:%M')
                    return str(v).strip()

                name       = cell(idx_name)
                rec_date   = _parse_date(row[idx_date] if idx_date < len(row) else None)
                check_in   = cell(idx_check_in)[:5] if cell(idx_check_in) else ''
                check_out  = cell(idx_check_out)[:5] if cell(idx_check_out) else ''
                department = cell(idx_department) if idx_department >= 0 else ''

                if not name or not rec_date:
                    continue

                # FIX: find or create employee, also update department
                emp = db.query(EASEmployee).filter(
                    EASEmployee.group_id == group_id,
                    EASEmployee.name == name
                ).first()
                if not emp:
                    emp = EASEmployee(group_id=group_id, name=name, department=department)
                    db.add(emp)
                    db.flush()
                elif department and not emp.department:
                    # FIX: backfill department if missing
                    emp.department = department

                # calculate status
                status = 'present'
                late_min = 0
                if check_in and group.shift_start:
                    try:
                        ci = datetime.strptime(check_in[:5], '%H:%M')
                        sh = datetime.strptime(group.shift_start, '%H:%M')
                        diff = int((ci - sh).total_seconds() / 60)
                        if diff > group.late_tolerance:
                            status = 'late'
                            late_min = diff
                    except:
                        pass

                if not check_in:
                    status = 'absent'

                rec = EASRecord(
                    employee_id=emp.id,
                    record_date=rec_date,
                    check_in=check_in[:5] if check_in else None,
                    check_out=check_out[:5] if check_out else None,
                    status=status,
                    late_minutes=late_min,
                    source='excel'
                )
                db.add(rec)
                added += 1
            except Exception as e:
                errors.append(str(e))

        db.commit()
        msg = _t(f'تم رفع {added} سجل', f'Uploaded {added} records')
        if errors:
            msg += f' | {len(errors)} errors'
        flash(msg, 'success')
        return redirect(url_for('eas.report', group_id=group_id))
    except Exception as e:
        db.rollback()
        flash(_t(f'خطأ في الملف: {e}', f'File error: {e}'), 'danger')
        return redirect(url_for('eas.upload'))

# ── Employee Search API (autocomplete) ───────────────
@eas_bp.route('/api/employees/search')
@admin_required
def employees_search():
    db = get_db()
    q = request.args.get('q', '').strip()
    group_id = request.args.get('group_id', type=int)
    if not q or len(q) < 2:
        return jsonify([])
    query = db.query(EASEmployee).filter(EASEmployee.name.ilike(f'%{q}%'))
    if group_id:
        query = query.filter(EASEmployee.group_id == group_id)
    emps = query.limit(10).all()
    return jsonify([{'id': e.id, 'name': e.name, 'name_en': e.name_en or '',
                     'department': e.department or ''} for e in emps])

# ── Report ─────────────────────────────────────────
@eas_bp.route('/report')
@admin_required
def report():
    db = get_db()
    cfg = _get_config()
    group_id   = request.args.get('group_id', type=int)
    date_from  = request.args.get('from', '')
    date_to    = request.args.get('to', '')
    # FIX: department filter
    dept_filter = request.args.get('department', '').strip()
    groups     = db.query(EASGroup).filter(EASGroup.config_id == cfg.id).all()

    use_shift   = request.args.get('use_shift', '')
    shift_names = request.args.get('shift_names', '')
    shift_from  = request.args.get('shift_from', '')
    shift_to    = request.args.get('shift_to', '')

    shift_name_list = [n.strip() for n in shift_names.splitlines() if n.strip()]

    records = []
    group = None
    shift_employee_ids = set()
    dept_names = []

    if group_id:
        group = db.get(EASGroup, group_id)
        q = db.query(EASRecord).join(EASEmployee).filter(EASEmployee.group_id == group_id)

        # FIX: filter by department
        if dept_filter:
            q = q.filter(EASEmployee.department == dept_filter)

        if date_from: q = q.filter(EASRecord.record_date >= date_from)
        if date_to:   q = q.filter(EASRecord.record_date <= date_to)
        records = q.order_by(EASRecord.record_date.desc(), EASEmployee.name).all()

        # فلتر المناوبة بالأسماء
        if use_shift and shift_name_list:
            shift_emps = db.query(EASEmployee).filter(
                EASEmployee.group_id == group_id,
                EASEmployee.name.in_(shift_name_list)
            ).all()
            shift_employee_ids = {e.id for e in shift_emps}

        # FIX: get distinct departments for dropdown
        dept_rows = db.query(EASEmployee.department).filter(
            EASEmployee.group_id == group_id,
            EASEmployee.department.isnot(None),
            EASEmployee.department != '',
        ).distinct().all()
        dept_names = sorted([r[0] for r in dept_rows if r[0]])

    return render_template('eas/report.html',
        config=cfg, groups=groups, group=group,
        records=records, shift_employees=shift_employee_ids,
        date_from=date_from, date_to=date_to,
        use_shift=use_shift, shift_names=shift_names,
        shift_from=shift_from, shift_to=shift_to,
        dept_filter=dept_filter, dept_names=dept_names,  # FIX: pass to template
    )
