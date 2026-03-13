"""
مسارات الحجوزات — مطابق حرفياً لـ v54
"""
import csv, io, hashlib
from datetime import datetime, timedelta
from utils.flash_helper import flash_msg
from flask import (Blueprint, render_template, redirect, url_for,
                   request, flash, jsonify, abort, Response, session)
from flask_login import login_required, current_user
from sqlalchemy import or_
from models.database import (Reservation, Venue, Location, BlockedPeriod,
                              Approval, SystemLog, Contact, BookingContact,
                              ChecklistItem, Rating, Attachment)
from utils.helpers import get_db, get_permissions, syslog, paginate

reservations_bp = Blueprint('reservations', __name__, url_prefix='/reservations')

STATUS_AR = {
    'pending':   'معلق',
    'approved':  'موافق عليه',
    'rejected':  'مرفوض',
    'cancelled': 'ملغي',
    'completed': 'مكتمل',
}
STATUS_EN = {
    'pending':   'Pending',
    'approved':  'Approved',
    'rejected':  'Rejected',
    'cancelled': 'Cancelled',
    'completed': 'Completed',
}
STATUS_CLS = {
    'pending':   'warning',
    'approved':  'success',
    'rejected':  'danger',
    'cancelled': 'secondary',
    'completed': 'info',
}

def _book_num(db):
    c = db.query(Reservation).count() + 1
    return f"ARS-{datetime.now().strftime('%Y%m%d%H%M%S')}-{c:04d}"

def _conflict(db, venue_id, start, end, exclude_id=None):
    q = db.query(Reservation).filter(
        Reservation.venue_id == venue_id,
        Reservation.status.in_(['pending','approved']),
        Reservation.start_time < end,
        Reservation.end_time   > start,
    )
    if exclude_id:
        q = q.filter(Reservation.id != exclude_id)
    c = q.first()
    if c:
        return f"{c.booking_number} ({c.start_time.strftime('%H:%M')}-{c.end_time.strftime('%H:%M')})"
    blk = db.query(BlockedPeriod).filter(
        BlockedPeriod.venue_id == venue_id,
        BlockedPeriod.start_time < end,
        BlockedPeriod.end_time   > start,
    ).first()
    if blk:
        return f"فترة محظورة: {blk.reason or 'صيانة'}"
    return None


# ── index ─────────────────────────────────────────────────────────────────────
@reservations_bp.route('/')
@login_required
def index():
    db    = get_db()
    perms = get_permissions()
    page  = request.args.get('page', 1, type=int)

    q = db.query(Reservation)
    if not perms.is_admin_or_manager():
        q = q.filter_by(user_id=current_user.id)

    status   = request.args.get('status', '')
    venue_id = request.args.get('venue_id', '')
    search   = request.args.get('q', '')

    if status:   q = q.filter(Reservation.status == status)
    if venue_id and venue_id.isdigit():
        q = q.filter(Reservation.venue_id == int(venue_id))
    if search:
        q = q.filter(or_(
            Reservation.title.ilike(f'%{search}%'),
            Reservation.booking_number.ilike(f'%{search}%')
        ))

    q = q.order_by(Reservation.created_at.desc())
    items, total, total_pages = paginate(q, page, 20)

    # إحصاءات مثل v54 update_reservation_stats
    base = db.query(Reservation)
    if not perms.is_admin_or_manager():
        base = base.filter_by(user_id=current_user.id)
    stats = {
        'total':     base.count(),
        'pending':   base.filter_by(status='pending').count(),
        'approved':  base.filter_by(status='approved').count(),
        'rejected':  base.filter_by(status='rejected').count(),
        'cancelled': base.filter_by(status='cancelled').count(),
    }

    venues = db.query(Venue).filter_by(is_active=True).all()

    # إشعارات الحجوزات القريبة (24 ساعة) — مثل check_upcoming_bookings
    now      = datetime.now()
    tomorrow = now + timedelta(days=1)
    upcoming = db.query(Reservation).filter(
        Reservation.user_id   == current_user.id,
        Reservation.start_time >= now,
        Reservation.start_time <= tomorrow,
        Reservation.status     == 'approved',
    ).all()

    return render_template('reservations/index.html',
        reservations=items, total=total, page=page, total_pages=total_pages,
        stats=stats, venues=venues,
        status_ar=STATUS_AR, status_en=STATUS_EN, status_cls=STATUS_CLS,
        upcoming=upcoming,
        filters={'status': status, 'venue_id': venue_id, 'q': search})


# ── new ───────────────────────────────────────────────────────────────────────
@reservations_bp.route('/new', methods=['GET','POST'])
@login_required
def new():
    db    = get_db()
    perms = get_permissions()
    if not perms.can('reservations_add'):
        flash_msg('ليس لديك صلاحية إنشاء حجز', 'danger')
        return redirect(url_for('reservations.index'))

    venues = perms.get_allowed_venues() if hasattr(perms, 'get_allowed_venues') else \
             db.query(Venue).filter_by(is_active=True).all()

    if request.method == 'POST':
        title     = request.form.get('title','').strip()
        venue_id  = request.form.get('venue_id','')
        start_str = request.form.get('start_time','')
        end_str   = request.form.get('end_time','')
        notes     = request.form.get('notes','').strip()

        errors = []
        if not title:    errors.append(_('عنوان الحجز مطلوب') if False else ('Title is required' if session.get('lang')=='en' else 'عنوان الحجز مطلوب'))
        if not venue_id: errors.append('Please select a venue' if session.get('lang')=='en' else 'يرجى اختيار القاعة')

        start_dt = end_dt = None
        try:
            start_dt = datetime.fromisoformat(start_str)
            end_dt   = datetime.fromisoformat(end_str)
            if start_dt < datetime.now():
                errors.append('⚠️ Cannot create a booking in the past' if session.get('lang')=='en' else '⚠️ لا يمكن إنشاء حجز بتاريخ في الماضي')
            if end_dt <= start_dt:
                errors.append('End time must be after start time' if session.get('lang')=='en' else 'وقت الانتهاء يجب أن يكون بعد وقت البداية')
        except:
            errors.append('Invalid time format' if session.get('lang')=='en' else 'صيغة الوقت غير صحيحة')

        if start_dt and end_dt and venue_id and venue_id.isdigit():
            c = _conflict(db, int(venue_id), start_dt, end_dt)
            if c: errors.append(f'تعارض: {c}')

        if errors:
            for e in errors: flash(e, 'danger')
            from models.database import Contact as _C
            _contacts = get_db().query(_C).filter_by(user_id=current_user.id).order_by(_C.name).all()
            return render_template('reservations/new.html',
                                   venues=venues, form=request.form, now=datetime.now(), contacts=_contacts)

        venue = db.query(Venue).get(int(venue_id))
        res   = Reservation(
            booking_number  = _book_num(db),
            title           = title,
            user_id         = current_user.id,
            venue_id        = int(venue_id),
            start_time      = start_dt,
            end_time        = end_dt,
            requester_notes = notes,
            status          = 'pending' if (venue and venue.requires_approval) else 'approved',
        )
        db.add(res)
        db.commit()
        syslog('NEW_RESERVATION', f'{res.booking_number} — {title}')

        # حفظ المرفقات
        import base64
        for f in request.files.getlist('attachments'):
            if f and f.filename:
                try:
                    data = base64.b64encode(f.read()).decode('utf-8')
                    att  = Attachment(reservation_id=res.id, filename=f.filename,
                                      mimetype=f.content_type or 'application/octet-stream',
                                      filedata=data, uploaded_by=current_user.id)
                    db.add(att)
                except: pass
        db.commit()

        # Save selected contacts
        contact_ids = request.form.getlist('contact_ids')
        if contact_ids:
            from models.database import BookingContact
            for cid in contact_ids:
                if cid.isdigit():
                    bc = BookingContact(reservation_id=res.id, contact_id=int(cid))
                    db.add(bc)
            db.commit()

        # إرسال بريد إشعار — مثل NewReservationWindow.save
        try:
            from utils.email_helper import send_booking_request
            send_booking_request(current_user, res)
        except: pass

        flash_msg(f'✅ تم تقديم طلب الحجز — رقم الحجز: {res.booking_number}', 'success')
        return redirect(url_for('reservations.detail', res_id=res.id))

    db2 = get_db()
    contacts = db2.query(__import__('models.database', fromlist=['Contact']).Contact).filter_by(user_id=current_user.id).order_by(__import__('models.database', fromlist=['Contact']).Contact.name).all()
    return render_template('reservations/new.html', venues=venues, form={}, now=datetime.now(), contacts=contacts)


# ── detail ────────────────────────────────────────────────────────────────────
@reservations_bp.route('/<int:res_id>')
@login_required
def detail(res_id):
    db    = get_db()
    perms = get_permissions()
    res   = db.query(Reservation).get(res_id)
    if not res: abort(404)
    if res.user_id != current_user.id and not perms.is_admin_or_manager():
        abort(403)

    checklist_items = db.query(ChecklistItem).filter_by(
        reservation_id=res_id).order_by(ChecklistItem.order_index).all()
    rating = db.query(Rating).filter_by(
        reservation_id=res_id, user_id=current_user.id).first()
    contacts = db.query(BookingContact).filter_by(booking_id=res_id).all()

    attachments = db.query(Attachment).filter_by(reservation_id=res_id).all()
    return render_template('reservations/detail.html',
        res=res, perms=perms,
        status_ar=STATUS_AR, status_en=STATUS_EN, status_cls=STATUS_CLS,
        checklist_items=checklist_items, rating=rating,
        booking_contacts=contacts,
        attachments=attachments)


# ── edit pending — مطابق لـ v54 edit_pending_reservation ─────────────────────
@reservations_bp.route('/<int:res_id>/edit', methods=['GET','POST'])
@login_required
def edit(res_id):
    db    = get_db()
    perms = get_permissions()
    res   = db.query(Reservation).get(res_id)
    if not res: abort(404)

    # يمكن تعديل pending/approved فقط (مثل v54)
    if res.status not in ('pending','approved'):
        _STATUS_EN = {'pending':'Pending','approved':'Approved','rejected':'Rejected','cancelled':'Cancelled','completed':'Completed'}
        if session.get('lang','ar') == 'en':
            flash('Cannot edit this booking — current status: ' + _STATUS_EN.get(res.status, res.status), 'warning')
        else:
            flash('لا يمكن تعديل هذا الحجز — الحالة الحالية: ' + STATUS_AR.get(res.status,''), 'warning')
        return redirect(url_for('reservations.detail', res_id=res_id))

    if res.user_id != current_user.id and not perms.is_admin_or_manager():
        abort(403)

    venues = db.query(Venue).filter_by(is_active=True).all()

    if request.method == 'POST':
        title     = request.form.get('title','').strip()
        venue_id  = request.form.get('venue_id','')
        start_str = request.form.get('start_time','')
        end_str   = request.form.get('end_time','')
        notes     = request.form.get('notes','').strip()

        errors = []
        start_dt = end_dt = None
        try:
            start_dt = datetime.fromisoformat(start_str)
            end_dt   = datetime.fromisoformat(end_str)
            if end_dt <= start_dt: errors.append('End time must be after start time' if session.get('lang')=='en' else 'وقت الانتهاء يجب أن يكون بعد وقت البداية')
        except:
            errors.append('Invalid time format' if session.get('lang')=='en' else 'صيغة الوقت غير صحيحة')

        if start_dt and end_dt and venue_id and venue_id.isdigit():
            c = _conflict(db, int(venue_id), start_dt, end_dt, exclude_id=res_id)
            if c: errors.append(f'تعارض: {c}')

        if errors:
            for e in errors: flash(e, 'danger')
            return render_template('reservations/edit.html',
                                   res=res, venues=venues)

        # إذا كان موافقاً عليه يرجع لمعلق (مثل v54)
        if res.status == 'approved':
            res.status = 'pending'

        res.title          = title
        res.venue_id       = int(venue_id) if venue_id.isdigit() else res.venue_id
        res.start_time     = start_dt
        res.end_time       = end_dt
        res.requester_notes= notes
        db.commit()
        syslog('EDIT_RESERVATION', f'{res.booking_number} — تم التعديل')
        flash_msg('✅ تم حفظ التعديلات — تم إرجاع الحجز لحالة معلق إن كان موافقاً عليه', 'success')
        return redirect(url_for('reservations.detail', res_id=res_id))

    return render_template('reservations/edit.html', res=res, venues=venues)


# ── approve / reject / cancel — مطابق لـ v54 ─────────────────────────────────
@reservations_bp.route('/<int:res_id>/approve', methods=['POST'])
@login_required
def approve(res_id):
    db    = get_db()
    perms = get_permissions()
    if not perms.can_action('reservations_approve','approve'): abort(403)
    res = db.query(Reservation).get(res_id)
    if not res: abort(404)
    if res.status != 'pending':
        flash_msg('يمكن الموافقة فقط على الحجوزات المعلقة', 'warning')
        return redirect(url_for('reservations.detail', res_id=res_id))

    res.status        = 'approved'
    res.approver_id   = current_user.id
    res.approval_date = datetime.now()
    res.approver_notes= request.form.get('note','')
    appr = Approval(reservation_id=res.id, approver_id=current_user.id,
                    status='approved', comments=request.form.get('note',''))
    db.add(appr)
    db.commit()
    syslog('APPROVE_RESERVATION', f'{res.booking_number}')

    # بريد موافقة مثل v54
    try:
        from utils.email_helper import send_booking_approved
        send_booking_approved(res)
    except: pass

    flash_msg('✅ تمت الموافقة على الحجز', 'success')
    return redirect(url_for('reservations.detail', res_id=res_id))


@reservations_bp.route('/<int:res_id>/reject', methods=['POST'])
@login_required
def reject(res_id):
    db    = get_db()
    perms = get_permissions()
    if not perms.can_action('reservations_approve','approve'): abort(403)
    res = db.query(Reservation).get(res_id)
    if not res: abort(404)
    if res.status != 'pending':
        flash_msg('يمكن رفض الحجوزات المعلقة فقط', 'warning')
        return redirect(url_for('reservations.detail', res_id=res_id))

    reason = request.form.get('reason','')
    res.status         = 'rejected'
    res.approver_id    = current_user.id
    res.approver_notes = reason
    res.approval_date  = datetime.now()
    appr = Approval(reservation_id=res.id, approver_id=current_user.id,
                    status='rejected', comments=reason)
    db.add(appr)
    db.commit()
    syslog('REJECT_RESERVATION', f'{res.booking_number} — {reason}')

    try:
        from utils.email_helper import send_booking_rejected
        send_booking_rejected(res, reason)
    except: pass

    flash_msg('تم رفض الحجز', 'warning')
    return redirect(url_for('reservations.detail', res_id=res_id))


@reservations_bp.route('/<int:res_id>/cancel', methods=['POST'])
@login_required
def cancel(res_id):
    db    = get_db()
    perms = get_permissions()
    res   = db.query(Reservation).get(res_id)
    if not res: abort(404)
    if res.user_id != current_user.id and not perms.is_admin_or_manager():
        abort(403)
    if res.status in ('cancelled','rejected'):
        flash_msg('الحجز ملغي أو مرفوض بالفعل', 'warning')
        return redirect(url_for('reservations.detail', res_id=res_id))

    res.status = 'cancelled'
    db.commit()
    syslog('CANCEL_RESERVATION', f'{res.booking_number}')
    flash_msg('تم إلغاء الحجز', 'success')
    return redirect(url_for('reservations.index'))


# ── export CSV — مطابق لـ v54 export_reservations_excel (CSV للويب) ──────────
@reservations_bp.route('/export-csv')
@reservations_bp.route('/export-excel')
@login_required
def export_csv():
    db    = get_db()
    perms = get_permissions()
    q     = db.query(Reservation)
    if not perms.is_admin_or_manager():
        q = q.filter_by(user_id=current_user.id)

    status   = request.args.get('status','')
    venue_id = request.args.get('venue_id','')
    if status:   q = q.filter(Reservation.status == status)
    if venue_id and venue_id.isdigit():
        q = q.filter(Reservation.venue_id == int(venue_id))

    # Use openpyxl for proper Excel file
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Reservations'
        headers = ['رقم الحجز','العنوان','المستخدم','القاعة','تاريخ البدء','تاريخ الانتهاء','الحالة','ملاحظات']
        hfill = PatternFill('solid', fgColor='1A555C')
        hfont = Font(color='FFFFFF', bold=True)
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(1, ci, h)
            cell.fill = hfill; cell.font = hfont
            cell.alignment = Alignment(horizontal='center')
        for ri, r in enumerate(q.all(), 2):
            row = [
                r.booking_number, r.title,
                r.user.full_name if r.user else '',
                r.venue.name if r.venue else '',
                r.start_time.strftime('%Y-%m-%d %H:%M') if r.start_time else '',
                r.end_time.strftime('%Y-%m-%d %H:%M')   if r.end_time   else '',
                STATUS_AR.get(r.status, r.status),
                r.requester_notes or '',
            ]
            for ci, val in enumerate(row, 1):
                ws.cell(ri, ci, val).alignment = Alignment(horizontal='center')
            if ri % 2 == 0:
                for ci in range(1, 9):
                    ws.cell(ri, ci).fill = PatternFill('solid', fgColor='EEF4F5')
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)
        ws.freeze_panes = 'A2'
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        fname = f'reservations_{datetime.now().strftime("%Y%m%d")}.xlsx'
        return Response(buf.read(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={fname}'})
    except ImportError:
        # Fallback to CSV
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Ref','Title','User','Venue','Start','End','Status','Notes'])
        for r in q.all():
            writer.writerow([r.booking_number, r.title,
                r.user.full_name if r.user else '',
                r.venue.name if r.venue else '',
                r.start_time.strftime('%Y-%m-%d %H:%M') if r.start_time else '',
                r.end_time.strftime('%Y-%m-%d %H:%M')   if r.end_time   else '',
                r.status, r.requester_notes or ''])
        content = b'\xef\xbb\xbf' + output.getvalue().encode('utf-8')
        fname   = f'reservations_{datetime.now().strftime("%Y%m%d")}.csv'
        return Response(content, mimetype='text/csv',
                        headers={'Content-Disposition': f'attachment; filename={fname}'})


# ── export single reservation PDF ────────────────────────────────────────────
@reservations_bp.route('/<int:res_id>/pdf')
@login_required
def export_single_pdf(res_id):
    import io, json, os
    db    = get_db()
    perms = get_permissions()
    res   = db.query(Reservation).get(res_id)
    if not res: abort(404)
    if res.user_id != current_user.id and not perms.is_admin_or_manager(): abort(403)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer, HRFlowable, Image)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
        from datetime import date as dt_date
        from utils.pdf_helper import register_arabic_font, ar, arabic_font
        import base64

        # Register Arabic font
        register_arabic_font()
        AF  = arabic_font(bold=False)
        AFB = arabic_font(bold=True)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2.5*cm)

        # Load maintenance config for header/footer
        mcfg = {}
        try:
            p = os.path.join(os.path.dirname(__file__), '..', 'maintenance_config.json')
            if os.path.exists(p): mcfg = json.loads(open(p).read())
        except: pass

        story = []

        # ── Letterhead Header ──
        center_style = ParagraphStyle('c', fontName=AFB, fontSize=14, alignment=TA_CENTER,
                                      textColor=colors.HexColor('#1A555C'), spaceAfter=3)
        sub_style    = ParagraphStyle('s', fontName=AF,  fontSize=11, alignment=TA_CENTER,
                                      textColor=colors.HexColor('#2E8B8F'), spaceAfter=2)
        small_style  = ParagraphStyle('sm', fontName=AF, fontSize=8, alignment=TA_CENTER,
                                      textColor=colors.grey, spaceAfter=6)

        # Header image (dedicated header_img_b64 or fallback to logo_b64)
        _himg_b64 = mcfg.get('header_img_b64') or mcfg.get('logo_b64')
        _himg_pos = mcfg.get('header_img_position', 'center').upper()
        _halign_map = {'LEFT': 'LEFT', 'CENTER': 'CENTER', 'RIGHT': 'RIGHT',
                       'left': 'LEFT', 'center': 'CENTER', 'right': 'RIGHT'}
        _halign = _halign_map.get(_himg_pos, 'CENTER')
        if _himg_b64:
            try:
                raw = _himg_b64
                if ',' in raw: raw = raw.split(',', 1)[1]
                img_data = base64.b64decode(raw)
                img_buf  = io.BytesIO(img_data)
                logo_img = Image(img_buf, width=3*cm, height=3*cm, kind='proportional')
                logo_img.hAlign = _halign
                story.append(logo_img)
                story.append(Spacer(1, 0.2*cm))
            except Exception: pass

        if mcfg.get('report_header_title'):
            story.append(Paragraph(ar(mcfg['report_header_title']), center_style))
        if mcfg.get('report_header_subtitle'):
            story.append(Paragraph(ar(mcfg['report_header_subtitle']), sub_style))
        if mcfg.get('report_header_extra'):
            story.append(Paragraph(ar(mcfg['report_header_extra']), small_style))
        if mcfg.get('report_header_title') or mcfg.get('report_header_subtitle'):
            story.append(HRFlowable(width='100%', thickness=1.5,
                                    color=colors.HexColor('#1A555C'), spaceAfter=10))

        # ── Title ──
        story.append(Paragraph(ar(f'تفاصيل الحجز — {res.booking_number}'),
            ParagraphStyle('t', fontName=AFB, fontSize=13, textColor=colors.HexColor('#1A555C'),
                           spaceAfter=4, alignment=TA_CENTER)))
        story.append(Paragraph(ar(f'تاريخ الطباعة: {dt_date.today()}'), small_style))
        story.append(Spacer(1, 0.4*cm))

        # ── Data table ──
        teal  = colors.HexColor('#1A555C')
        light = colors.HexColor('#EEF4F5')
        # Status English labels for PDF (avoids Arabic reshaping issues)
        STATUS_EN_PDF = {'pending':'Pending','approved':'Approved','rejected':'Rejected',
                         'cancelled':'Cancelled','completed':'Completed'}
        data = [
            [ar('رقم الحجز'),       str(res.booking_number or '—')],
            [ar('العنوان'),          ar(str(res.title or '—'))],
            [ar('المستخدم'),        ar(str(res.user.full_name if res.user else '—'))],
            [ar('القاعة'),           ar(str(res.venue.name if res.venue else '—'))],
            [ar('الموقع'),           ar(str(res.venue.location.name if res.venue and res.venue.location else '—'))],
            [ar('تاريخ البدء'),     res.start_time.strftime('%Y-%m-%d  %H:%M') if res.start_time else '—'],
            [ar('تاريخ الانتهاء'), res.end_time.strftime('%Y-%m-%d  %H:%M') if res.end_time else '—'],
            [ar('الحالة'),           STATUS_EN_PDF.get(res.status, res.status)],
            [ar('ملاحظات'),          ar(str(res.requester_notes or '—'))],
        ]
        t = Table(data, colWidths=[5*cm, 11*cm])
        t.setStyle(TableStyle([
            ('FONTNAME',   (0,0), (-1,-1), AF),
            ('FONTNAME',   (0,0), (0,-1),  AFB),
            ('FONTSIZE',   (0,0), (-1,-1), 10),
            ('ALIGN',      (0,0), (-1,-1), 'RIGHT'),
            ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
            ('BACKGROUND', (0,0), (0,-1), light),
            ('TEXTCOLOR',  (0,0), (0,-1), teal),
            ('GRID',       (0,0), (-1,-1), 0.5, colors.HexColor('#dde8ea')),
            ('PADDING',    (0,0), (-1,-1), 8),
            ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.white, light]),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.5*cm))

        # ── Footer ──
        footer_text = mcfg.get('report_header_footer') or mcfg.get('report_header_title', 'ARS — Applied Reservation System')
        story.append(HRFlowable(width='100%', thickness=0.5,
                                color=colors.HexColor('#9ab3b8'), spaceAfter=4))
        story.append(Paragraph(ar(footer_text),
            ParagraphStyle('f', fontName=AF, fontSize=8, alignment=TA_CENTER, textColor=colors.grey)))

        doc.build(story)
        buf.seek(0)
        return Response(buf.read(), mimetype='application/pdf',
            headers={'Content-Disposition': f'attachment;filename=reservation_{res.booking_number}.pdf'})
    except ImportError:
        flash_msg('reportlab غير مثبت', 'danger')
        return redirect(url_for('reservations.detail', res_id=res_id))


# ── export single reservation Excel ──────────────────────────────────────────
@reservations_bp.route('/<int:res_id>/excel')
@login_required
def export_single_excel(res_id):
    import io
    db    = get_db()
    perms = get_permissions()
    res   = db.query(Reservation).get(res_id)
    if not res: abort(404)
    if res.user_id != current_user.id and not perms.is_admin_or_manager(): abort(403)

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Reservation'
        ws.sheet_view.rightToLeft = True

        teal_fill = PatternFill('solid', fgColor='1A555C')
        light_fill = PatternFill('solid', fgColor='EEF4F5')
        teal_font  = Font(color='FFFFFF', bold=True, size=11)
        bold_font  = Font(bold=True, size=10)
        thin_border = Border(
            left=Side(style='thin',color='9ab3b8'),
            right=Side(style='thin',color='9ab3b8'),
            top=Side(style='thin',color='9ab3b8'),
            bottom=Side(style='thin',color='9ab3b8'))

        rows = [
            ('رقم الحجز',    res.booking_number),
            ('العنوان',       res.title),
            ('المستخدم',     res.user.full_name if res.user else '—'),
            ('القاعة',        res.venue.name if res.venue else '—'),
            ('الموقع',        res.venue.location.name if res.venue and res.venue.location else '—'),
            ('تاريخ البدء',  str(res.start_time) if res.start_time else '—'),
            ('تاريخ الانتهاء', str(res.end_time) if res.end_time else '—'),
            ('الحالة',        STATUS_AR.get(res.status, res.status)),
            ('ملاحظات',       res.requester_notes or '—'),
        ]

        # Title row
        ws.merge_cells('A1:B1')
        tc = ws['A1']
        tc.value = f'تفاصيل الحجز — {res.booking_number}'
        tc.fill = teal_fill; tc.font = teal_font
        tc.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        for ri, (label, value) in enumerate(rows, 2):
            ca = ws.cell(ri, 1, label)
            cb = ws.cell(ri, 2, value)
            ca.font = bold_font
            ca.fill = light_fill
            ca.alignment = Alignment(horizontal='right', vertical='center')
            cb.alignment = Alignment(horizontal='right', vertical='center')
            for cell in (ca, cb):
                cell.border = thin_border
            ws.row_dimensions[ri].height = 20

        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 42

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        fname = f'reservation_{res.booking_number}.xlsx'
        return Response(buf.read(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment;filename={fname}'})
    except ImportError:
        flash_msg('openpyxl غير مثبت', 'danger')
        return redirect(url_for('reservations.detail', res_id=res_id))


# ── export PDF ────────────────────────────────────────────────────────────────
@reservations_bp.route('/export-pdf')
@login_required
def export_pdf():
    import io, json, os
    db    = get_db()
    perms = get_permissions()
    q     = db.query(Reservation)
    if not perms.is_admin_or_manager():
        q = q.filter_by(user_id=current_user.id)

    status   = request.args.get('status','')
    venue_id = request.args.get('venue_id','')
    if status:   q = q.filter(Reservation.status == status)
    if venue_id and venue_id.isdigit():
        q = q.filter(Reservation.venue_id == int(venue_id))
    items = q.order_by(Reservation.created_at.desc()).all()

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from datetime import date as dt_date
        from utils.pdf_helper import register_arabic_font, ar, arabic_font

        register_arabic_font()
        AF  = arabic_font(bold=False)
        AFB = arabic_font(bold=True)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                rightMargin=1.5*cm, leftMargin=1.5*cm,
                                topMargin=1.5*cm, bottomMargin=1.5*cm)
        styles = getSampleStyleSheet()
        story  = []

        # Header from maintenance config
        mcfg = {}
        try:
            p = os.path.join(os.path.dirname(__file__), '..', 'maintenance_config.json')
            if os.path.exists(p): mcfg = json.loads(open(p).read())
        except: pass

        if mcfg.get('report_header_title'):
            story.append(Paragraph(ar(mcfg['report_header_title']),
                ParagraphStyle('h', fontName=AFB, fontSize=14, textColor=colors.HexColor('#1A555C'), spaceAfter=2)))
        if mcfg.get('report_header_subtitle'):
            story.append(Paragraph(ar(mcfg['report_header_subtitle']),
                ParagraphStyle('s', fontName=AF, fontSize=11, textColor=colors.HexColor('#2E8B8F'), spaceAfter=4)))
        if mcfg.get('report_header_title'):
            story.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#1A555C'), spaceAfter=6))

        story.append(Paragraph(ar(f'تقرير الحجوزات — ARS  |  {dt_date.today()}'),
            ParagraphStyle('t', fontName=AFB, fontSize=12, textColor=colors.HexColor('#1A555C'), spaceAfter=10)))

        headers = [ar('الرقم'), ar('العنوان'), ar('المستخدم'), ar('القاعة'), ar('البدء'), ar('الانتهاء'), ar('الحالة')]
        data = [headers]
        for r in items:
            data.append([
                r.booking_number or '',
                ar((r.title or '')[:30]),
                ar(r.user.full_name if r.user else ''),
                ar(r.venue.name if r.venue else ''),
                r.start_time.strftime('%Y-%m-%d %H:%M') if r.start_time else '',
                r.end_time.strftime('%H:%M') if r.end_time else '',
                {'pending':'Pending','approved':'Approved','rejected':'Rejected','cancelled':'Cancelled','completed':'Completed'}.get(r.status, r.status),
            ])

        col_widths = [3.5*cm, 5*cm, 3.5*cm, 3.5*cm, 4*cm, 2*cm, 2.5*cm]
        t = Table(data, repeatRows=1, colWidths=col_widths)
        t.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (-1,0), colors.HexColor('#1A555C')),
            ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
            ('FONTNAME',      (0,0), (-1,-1), AF),
            ('FONTNAME',      (0,0), (-1,0),  AFB),
            ('FONTSIZE',      (0,0), (-1,-1), 8),
            ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
            ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
            ('GRID',          (0,0), (-1,-1), 0.4, colors.HexColor('#dee2e6')),
            ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.white, colors.HexColor('#EEF4F5')]),
            ('PADDING',       (0,0), (-1,-1), 5),
            ('TOPPADDING',    (0,0), (-1,0), 8),
        ]))
        story.append(t)
        doc.build(story)
        buf.seek(0)
        fname = f'reservations_{dt_date.today()}.pdf'
        return Response(buf.read(), mimetype='application/pdf',
                        headers={'Content-Disposition': f'attachment;filename={fname}'})
    except ImportError:
        flash_msg('reportlab غير مثبت', 'danger')
        return redirect(url_for('reservations.index'))


# ── send invitations — مطابق لـ v54 SendInvitationsWindow ────────────────────
@reservations_bp.route('/<int:res_id>/invite', methods=['GET','POST'])
@login_required
def invite(res_id):
    db    = get_db()
    perms = get_permissions()
    res   = db.query(Reservation).get(res_id)
    if not res: abort(404)
    if res.user_id != current_user.id and not perms.is_admin_or_manager():
        abort(403)

    contacts = db.query(Contact).filter_by(created_by=current_user.id).all()

    if request.method == 'POST':
        selected_ids = request.form.getlist('contact_ids', type=int)
        message_body = request.form.get('message','').strip()
        if not selected_ids:
            flash_msg('اختر مدعويين على الأقل', 'warning')
            return redirect(url_for('reservations.invite', res_id=res_id))

        success = 0
        for cid in selected_ids:
            contact = db.query(Contact).get(cid)
            if not contact: continue
            try:
                personal_msg = message_body.replace('[NAME]', contact.first_name or '')
                from utils.email_helper import send_invitation
                send_invitation(contact, res, personal_msg)

                bc = db.query(BookingContact).filter_by(
                    booking_id=res_id, contact_id=cid).first()
                if not bc:
                    bc = BookingContact(booking_id=res_id, contact_id=cid,
                                       invitation_sent=True, sent_at=datetime.now())
                    db.add(bc)
                else:
                    bc.invitation_sent = True
                    bc.sent_at = datetime.now()
                success += 1
            except Exception as e:
                flash_msg(f'فشل إرسال إلى {contact.email}: {e}', 'warning')

        db.commit()
        syslog('SEND_INVITATIONS', f'{res.booking_number} — {success} دعوة')
        flash_msg(f'✅ تم إرسال {success} دعوة', 'success')
        return redirect(url_for('reservations.detail', res_id=res_id))

    default_msg = f"""عزيزي/عزيزتي [NAME]،

يسرنا دعوتكم لحضور "{res.title}" في:
🏢 القاعة: {res.venue.name if res.venue else '—'}
📅 التاريخ: {res.start_time.strftime('%Y-%m-%d') if res.start_time else ''}
⏰ الوقت: {res.start_time.strftime('%H:%M') if res.start_time else ''}

نتطلع لرؤيتكم،
فريق ARS"""

    return render_template('reservations/invite.html',
        res=res, contacts=contacts, default_msg=default_msg)


# ── notifications API — upcoming bookings ─────────────────────────────────────
@reservations_bp.route('/api/upcoming')
@login_required
def api_upcoming():
    db       = get_db()
    now      = datetime.now()
    tomorrow = now + timedelta(days=1)
    upcoming = db.query(Reservation).filter(
        Reservation.user_id    == current_user.id,
        Reservation.start_time >= now,
        Reservation.start_time <= tomorrow,
        Reservation.status     == 'approved',
    ).all()
    return jsonify([{
        'id':       r.id,
        'title':    r.title,
        'venue':    r.venue.name if r.venue else '—',
        'start':    r.start_time.strftime('%Y-%m-%d %H:%M') if r.start_time else '',
        'booking_number': r.booking_number,
    } for r in upcoming])


# ── availability AJAX ─────────────────────────────────────────────────────────
@reservations_bp.route('/api/availability')
@login_required
def availability():
    db       = get_db()
    venue_id = request.args.get('venue_id', type=int)
    date_str = request.args.get('date','')
    if not venue_id or not date_str:
        return jsonify([])
    try:
        d = datetime.strptime(date_str,'%Y-%m-%d').date()
    except:
        return jsonify([])
    rsvs = db.query(Reservation).filter(
        Reservation.venue_id == venue_id,
        Reservation.status.in_(['pending','approved']),
    ).all()
    return jsonify([{
        'id': r.id, 'title': r.title,
        'start': r.start_time.strftime('%H:%M') if r.start_time else '',
        'end':   r.end_time.strftime('%H:%M')   if r.end_time   else '',
        'status': r.status,
    } for r in rsvs if r.start_time and r.start_time.date() == d])


# ── Download attachment ───────────────────────────────────────────────────────
@reservations_bp.route('/attachment/<int:att_id>')
@login_required
def download_attachment(att_id):
    import base64 as b64lib
    db  = get_db()
    att = db.query(Attachment).get(att_id)
    if not att: abort(404)
    res = db.query(Reservation).get(att.reservation_id)
    if not res: abort(404)
    perms = get_permissions()
    if res.user_id != current_user.id and not perms.is_admin_or_manager(): abort(403)
    data = b64lib.b64decode(att.filedata)
    return Response(data, mimetype=att.mimetype,
                    headers={'Content-Disposition': f'attachment; filename="{att.filename}"'})


# ── Reactivate — إعادة تفعيل حجز ملغى أو مرفوض ───────────────────────────────
@reservations_bp.route('/<int:res_id>/reactivate', methods=['POST'])
@login_required
def reactivate(res_id):
    db    = get_db()
    perms = get_permissions()
    res   = db.query(Reservation).get(res_id)
    if not res: abort(404)
    if not perms.is_admin_or_manager(): abort(403)
    if res.status not in ('cancelled', 'rejected'):
        flash_msg('يمكن إعادة تفعيل الحجوزات الملغية أو المرفوضة فقط', 'warning')
        return redirect(url_for('reservations.detail', res_id=res_id))
    res.status = 'pending'
    db.commit()
    syslog('REACTIVATE', f'إعادة تفعيل حجز #{res.booking_number}')
    flash_msg('تم إعادة تفعيل الحجز وإرجاعه للمراجعة', 'success')
    return redirect(url_for('reservations.detail', res_id=res_id))

