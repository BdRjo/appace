"""
التقارير والمقارنة — مطابق حرفياً لـ v54 ReportsWindow + ComparisonWindow
"""
import io, csv
from datetime import datetime, timedelta, date
from utils.flash_helper import flash_msg
from flask import Blueprint, render_template, request, Response, abort
from flask_login import login_required, current_user
from sqlalchemy import func
from models.database import (Reservation, User, Venue, Location,
                              BlockedPeriod, Rating)
from utils.helpers import get_db, get_permissions

reports_bp = Blueprint('reports', __name__, url_prefix='/reports')


def _date_filter(q, f, start=None, end_d=None):
    """مطابق لـ v54 ReportsWindow.get_date_filter"""
    today = date.today()
    if f == 'today':
        return q.filter(func.date(Reservation.start_time) == today)
    elif f == 'week':
        ws = today - timedelta(days=today.weekday())
        return q.filter(Reservation.start_time >= datetime.combine(ws, datetime.min.time()))
    elif f == 'month':
        ms = today.replace(day=1)
        return q.filter(Reservation.start_time >= datetime.combine(ms, datetime.min.time()))
    elif f == 'custom' and start and end_d:
        try:
            s = datetime.strptime(start, '%Y-%m-%d')
            e = datetime.strptime(end_d, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            return q.filter(Reservation.start_time >= s, Reservation.start_time <= e)
        except: pass
    return q


# ── Main reports ──────────────────────────────────────────────────────────────
@reports_bp.route('/')
@login_required
def index():
    db    = get_db()
    perms = get_permissions()
    if not perms.can('reports_view'): abort(403)

    # مطابق لـ v54 _rt_defs report types
    rtype = request.args.get('type', 'my_reservations')
    f     = request.args.get('filter', 'all')
    start = request.args.get('start', '')
    end_d = request.args.get('end', '')
    page  = request.args.get('page', 1, type=int)
    per_p = 25

    rows=[]; columns=[]; stats={}; total_count=0

    if rtype == 'my_reservations':
        q = db.query(Reservation).filter_by(user_id=current_user.id)
        q = _date_filter(q, f, start, end_d)
        # مطابق لـ v54 update_stats_bar
        stats = {
            'total':     q.count(),
            'approved':  q.filter_by(status='approved').count(),
            'pending':   q.filter_by(status='pending').count(),
            'rejected':  q.filter_by(status='rejected').count(),
            'cancelled': q.filter_by(status='cancelled').count(),
        }
        total_count = q.count()
        items = q.order_by(Reservation.created_at.desc()).offset((page-1)*per_p).limit(per_p).all()
        columns = ['رقم الحجز','العنوان','القاعة','تاريخ البدء','الانتهاء','الحالة']
        rows = [[r.booking_number, r.title,
                 r.venue.name if r.venue else '—',
                 r.start_time.strftime('%Y-%m-%d %H:%M') if r.start_time else '',
                 r.end_time.strftime('%H:%M') if r.end_time else '',
                 r.status] for r in items]

    elif rtype == 'all_reservations' and perms.is_admin_or_manager():
        q = db.query(Reservation)
        q = _date_filter(q, f, start, end_d)
        stats = {
            'total':     q.count(),
            'approved':  q.filter_by(status='approved').count(),
            'pending':   q.filter_by(status='pending').count(),
            'rejected':  q.filter_by(status='rejected').count(),
            'cancelled': q.filter_by(status='cancelled').count(),
        }
        total_count = q.count()
        items = q.order_by(Reservation.created_at.desc()).offset((page-1)*per_p).limit(per_p).all()
        columns = ['رقم الحجز','العنوان','المستخدم','القاعة','تاريخ البدء','الحالة']
        rows = [[r.booking_number, r.title,
                 r.user.full_name if r.user else '—',
                 r.venue.name if r.venue else '—',
                 r.start_time.strftime('%Y-%m-%d %H:%M') if r.start_time else '',
                 r.status] for r in items]

    elif rtype == 'venues':
        venues = db.query(Venue).order_by(Venue.name).all()
        stats  = {'total_venues': len(venues),
                  'active_venues': sum(1 for v in venues if v.is_active)}
        total_count = len(venues)
        columns = ['القاعة','الموقع','إجمالي الحجوزات','حجوزاتي','فترات محظورة','الحالة']
        rows = [[
            v.name,
            v.location.name if v.location else '—',
            len(v.reservations),
            len([r for r in v.reservations if r.user_id == current_user.id]),
            db.query(BlockedPeriod).filter_by(venue_id=v.id).count(),
            'نشط' if v.is_active else 'غير نشط',
        ] for v in venues]

    elif rtype == 'locations':
        locs = db.query(Location).order_by(Location.name).all()
        stats = {'total_locations': len(locs),
                 'active_locations': sum(1 for l in locs if l.is_active)}
        total_count = len(locs)
        columns = ['الموقع','المدينة','المنطقة','القاعات','إجمالي الحجوزات']
        rows = [[
            l.name, l.city or '—', l.area or '—',
            len(l.venues),
            sum(len(v.reservations) for v in l.venues),
        ] for l in locs]

    elif rtype == 'users' and perms.can('users_view'):
        users = db.query(User).order_by(User.full_name).all()
        stats = {'total_users': len(users),
                 'active_users':    sum(1 for u in users if u.is_active),
                 'verified_users':  sum(1 for u in users if u.is_verified)}
        total_count = len(users)
        columns = ['الاسم الكامل','اسم المستخدم','البريد','الدور','الحجوزات','الحالة']
        rows = [[
            u.full_name, u.username, u.email or '—',
            u.role.name if u.role else '—',
            len(u.reservations),
            '✓' if u.is_active else '✗',
        ] for u in users]

    total_pages = max(1, (total_count + per_p - 1) // per_p)

    # بيانات الرسم البياني الشهري (12 شهر) — مثل v54 show_bar_chart
    now = datetime.now()
    monthly = []
    for i in range(11, -1, -1):
        d   = (now - timedelta(days=30*i)).replace(day=1, hour=0, minute=0, second=0)
        nd  = (d + timedelta(days=32)).replace(day=1)
        cnt = db.query(Reservation).filter(
            Reservation.created_at >= d, Reservation.created_at < nd).count()
        monthly.append({'month': d.strftime('%m/%Y'), 'count': cnt})

    # أعلى 5 قاعات
    top_venues = (db.query(Venue.name, func.count(Reservation.id).label('cnt'))
                  .join(Reservation, Reservation.venue_id == Venue.id, isouter=True)
                  .group_by(Venue.id)
                  .order_by(func.count(Reservation.id).desc())
                  .limit(5).all())

    return render_template('reports/index.html',
        rows=rows, columns=columns, stats=stats,
        rtype=rtype, filter=f, start=start, end=end_d,
        page=page, total_pages=total_pages, total_count=total_count,
        monthly=monthly, top_venues=top_venues, perms=perms)


# ── export CSV — مطابق لـ v54 export_csv ─────────────────────────────────────
@reports_bp.route('/export-csv')
@login_required
def export_csv():
    db    = get_db()
    perms = get_permissions()
    if not perms.can('reports_export'): abort(403)

    rtype = request.args.get('type','my_reservations')
    f     = request.args.get('filter','all')
    start = request.args.get('start','')
    end_d = request.args.get('end','')

    out    = io.StringIO()
    writer = csv.writer(out)

    if rtype == 'my_reservations':
        writer.writerow(['رقم الحجز','العنوان','القاعة','تاريخ البدء','الانتهاء','الحالة'])
        q = _date_filter(db.query(Reservation).filter_by(user_id=current_user.id), f, start, end_d)
        for r in q.all():
            writer.writerow([r.booking_number, r.title,
                r.venue.name if r.venue else '',
                r.start_time.strftime('%Y-%m-%d %H:%M') if r.start_time else '',
                r.end_time.strftime('%H:%M') if r.end_time else '', r.status])

    elif rtype == 'all_reservations' and perms.is_admin_or_manager():
        writer.writerow(['رقم الحجز','العنوان','المستخدم','القاعة','تاريخ البدء','الحالة'])
        q = _date_filter(db.query(Reservation), f, start, end_d)
        for r in q.all():
            writer.writerow([r.booking_number, r.title,
                r.user.full_name if r.user else '',
                r.venue.name if r.venue else '',
                r.start_time.strftime('%Y-%m-%d %H:%M') if r.start_time else '', r.status])

    elif rtype == 'venues':
        writer.writerow(['القاعة','الموقع','إجمالي الحجوزات','فترات محظورة','الحالة'])
        for v in db.query(Venue).all():
            writer.writerow([v.name, v.location.name if v.location else '',
                len(v.reservations),
                db.query(BlockedPeriod).filter_by(venue_id=v.id).count(),
                'نشط' if v.is_active else 'غير نشط'])

    elif rtype == 'locations':
        writer.writerow(['الموقع','المدينة','المنطقة','القاعات','إجمالي الحجوزات'])
        for l in db.query(Location).all():
            writer.writerow([l.name, l.city or '', l.area or '',
                len(l.venues), sum(len(v.reservations) for v in l.venues)])

    elif rtype == 'users' and perms.can('users_view'):
        writer.writerow(['الاسم الكامل','اسم المستخدم','البريد','الدور','الحجوزات','الحالة'])
        for u in db.query(User).all():
            writer.writerow([u.full_name, u.username, u.email or '',
                u.role.name if u.role else '', len(u.reservations),
                'نشط' if u.is_active else 'غير نشط'])

    content = b'\xef\xbb\xbf' + out.getvalue().encode('utf-8')
    fname   = f'report_{rtype}_{datetime.now().strftime("%Y%m%d")}.csv'
    return Response(content, mimetype='text/csv',
                    headers={'Content-Disposition': f'attachment; filename={fname}'})


# ── comparison — مطابق حرفياً لـ v54 ComparisonWindow ───────────────────────
@reports_bp.route('/comparison')
@login_required
def comparison():
    db    = get_db()
    perms = get_permissions()
    if not perms.is_admin_or_manager(): abort(403)

    mode  = request.args.get('mode', 'venue')
    a     = request.args.get('a', '')
    b     = request.args.get('b', '')
    fa    = request.args.get('fa', '')
    ta    = request.args.get('ta', '')
    fb    = request.args.get('fb', '')
    tb    = request.args.get('tb', '')

    venues    = db.query(Venue).filter_by(is_active=True).order_by(Venue.name).all()
    locations = db.query(Location).filter_by(is_active=True).order_by(Location.name).all()
    users_all = db.query(User).filter_by(is_active=True).order_by(User.full_name).all()

    result = {}

    # مطابق لـ v54 _gather_venue_data
    if mode == 'venue' and a and b:
        def _vs(name):
            v = db.query(Venue).filter_by(name=name).first()
            if not v: return {}
            res = v.reservations
            return {
                'إجمالي الحجوزات': len(res),
                'حجوزاتي':  len([r for r in res if r.user_id == current_user.id]),
                'موافقة':   len([r for r in res if r.status == 'approved']),
                'معلقة':    len([r for r in res if r.status == 'pending']),
                'مرفوضة':  len([r for r in res if r.status == 'rejected']),
                'فترات محظورة': db.query(BlockedPeriod).filter_by(venue_id=v.id).count(),
            }
        result = {'label_a': a, 'label_b': b, 'title': 'مقارنة القاعات',
                  'data_a': _vs(a), 'data_b': _vs(b)}

    # مطابق لـ v54 _gather_location_data
    elif mode == 'location' and a and b:
        def _ls(name):
            loc = db.query(Location).filter_by(name=name).first()
            if not loc: return {}
            res = sum(len(v.reservations) for v in loc.venues)
            app = sum(len([r for r in v.reservations if r.status=='approved']) for v in loc.venues)
            blk = sum(db.query(BlockedPeriod).filter_by(venue_id=v.id).count() for v in loc.venues)
            return {'القاعات': len(loc.venues), 'إجمالي الحجوزات': res,
                    'موافقة': app, 'فترات محظورة': blk}
        result = {'label_a': a, 'label_b': b, 'title': 'مقارنة المواقع',
                  'data_a': _ls(a), 'data_b': _ls(b)}

    # مطابق لـ v54 _gather_period_data
    elif mode == 'period' and fa and ta and fb and tb:
        def _ps(d_from, d_to):
            try:
                s = datetime.strptime(d_from, '%Y-%m-%d')
                e = datetime.strptime(d_to, '%Y-%m-%d').replace(hour=23, minute=59)
                q = db.query(Reservation).filter(
                    Reservation.start_time >= s, Reservation.start_time <= e)
                return {'إجمالي': q.count(),
                        'موافقة': q.filter_by(status='approved').count(),
                        'معلقة':  q.filter_by(status='pending').count(),
                        'مرفوضة':q.filter_by(status='rejected').count(),
                        'ملغاة': q.filter_by(status='cancelled').count()}
            except: return {}
        result = {'label_a': f'{fa} → {ta}', 'label_b': f'{fb} → {tb}',
                  'title': 'مقارنة الفترات',
                  'data_a': _ps(fa, ta), 'data_b': _ps(fb, tb)}

    # مطابق لـ v54 _gather_user_data
    elif mode == 'user' and a and b:
        def _us(uname):
            u = db.query(User).filter_by(username=uname).first()
            if not u: return {}
            res = u.reservations
            return {'إجمالي': len(res),
                    'موافقة': len([r for r in res if r.status=='approved']),
                    'معلقة':  len([r for r in res if r.status=='pending']),
                    'مرفوضة':len([r for r in res if r.status=='rejected']),
                    'ملغاة': len([r for r in res if r.status=='cancelled'])}
        result = {'label_a': a, 'label_b': b, 'title': 'مقارنة المستخدمين',
                  'data_a': _us(a), 'data_b': _us(b)}

    # مطابق لـ v54 _gather_status_data
    elif mode == 'status':
        q = db.query(Reservation)
        data = {'موافقة': q.filter_by(status='approved').count(),
                'معلقة':  q.filter_by(status='pending').count(),
                'مرفوضة':q.filter_by(status='rejected').count(),
                'ملغاة': q.filter_by(status='cancelled').count(),
                'مكتملة':q.filter_by(status='completed').count()}
        result = {'label_a': 'توزيع الحالات', 'label_b': '',
                  'title': 'توزيع حالات الحجوزات',
                  'data_a': data, 'data_b': {}, 'single': True}

    return render_template('reports/comparison.html',
        mode=mode, result=result, a=a, b=b,
        fa=fa, ta=ta, fb=fb, tb=tb,
        venues=venues, locations=locations, users_all=users_all)



# ── _get_report_data helper ───────────────────────────────────────────────────
def _get_report_data(db, perms, rtype, filter_val, start, end_d):
    """Returns (list_of_dicts, title_string) for PDF/Excel export."""
    rows = []
    title = 'Report'
    if rtype == 'my_reservations':
        title = 'My Reservations'
        q = _date_filter(db.query(Reservation).filter_by(user_id=current_user.id), filter_val, start, end_d)
        for r in q.all():
            rows.append({'Ref': r.booking_number, 'Title': r.title,
                'Venue': r.venue.name if r.venue else '',
                'Start': r.start_time.strftime('%Y-%m-%d %H:%M') if r.start_time else '',
                'End':   r.end_time.strftime('%H:%M') if r.end_time else '',
                'Status': r.status})
    elif rtype == 'all_reservations' and perms.is_admin_or_manager():
        title = 'All Reservations'
        q = _date_filter(db.query(Reservation), filter_val, start, end_d)
        for r in q.all():
            rows.append({'Ref': r.booking_number, 'Title': r.title,
                'User': r.user.full_name if r.user else '',
                'Venue': r.venue.name if r.venue else '',
                'Start': r.start_time.strftime('%Y-%m-%d %H:%M') if r.start_time else '',
                'Status': r.status})
    elif rtype == 'venues':
        title = 'Venues Report'
        for v in db.query(Venue).all():
            rows.append({'Venue': v.name,
                'Location': v.location.name if v.location else '',
                'Total Reservations': len(v.reservations),
                'Blocked': db.query(BlockedPeriod).filter_by(venue_id=v.id).count(),
                'Status': 'Active' if v.is_active else 'Inactive'})
    elif rtype == 'locations':
        title = 'Locations Report'
        for l in db.query(Location).all():
            rows.append({'Location': l.name, 'City': l.city or '', 'Area': l.area or '',
                'Venues': len(l.venues),
                'Total Reservations': sum(len(v.reservations) for v in l.venues)})
    elif rtype == 'users' and perms.can('users_view'):
        title = 'Users Report'
        for u in db.query(User).all():
            rows.append({'Full Name': u.full_name, 'Username': u.username,
                'Email': u.email or '', 'Role': u.role.name if u.role else '',
                'Reservations': len(u.reservations),
                'Status': 'Active' if u.is_active else 'Inactive'})
    return rows, title


# ── Export PDF ────────────────────────────────────────────────────────────────
@reports_bp.route('/export/pdf')
@login_required
def export_pdf():
    from flask import Response
    import io
    db    = get_db()
    perms = get_permissions()
    rtype  = request.args.get('type', 'my_reservations')
    filter = request.args.get('filter', 'all')
    start  = request.args.get('start', '')
    end    = request.args.get('end', '')
    rows, title_txt = _get_report_data(db, perms, rtype, filter, start, end)

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.lib.units import cm
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        import os, json

        buf = io.BytesIO()
        # Use landscape for wide tables
        page_size = landscape(A4) if rows and len(rows[0]) > 5 else A4
        doc = SimpleDocTemplate(buf, pagesize=page_size,
                                rightMargin=1.5*cm, leftMargin=1.5*cm,
                                topMargin=2*cm, bottomMargin=2*cm)

        # Use ARS Arabic font helper
        from utils.pdf_helper import register_arabic_font, ar, arabic_font as _af
        register_arabic_font()
        arabic_font = _af(bold=False)
        arabic_font_bold = _af(bold=True)

        styles = getSampleStyleSheet()
        story  = []

        # Load report header from maintenance config
        mcfg = {}
        try:
            cfg_path = os.path.join(os.path.dirname(__file__), '..', 'maintenance_config.json')
            if os.path.exists(cfg_path):
                mcfg = json.loads(open(cfg_path).read())
        except: pass

        # Header
        if mcfg.get('report_header_title'):
            h_style = ParagraphStyle('h', fontName=arabic_font_bold, fontSize=14,
                                     textColor=colors.HexColor('#1A555C'), alignment=TA_CENTER, spaceAfter=2)
            story.append(Paragraph(ar(mcfg['report_header_title']), h_style))
        if mcfg.get('report_header_subtitle'):
            s_style = ParagraphStyle('s', fontName=arabic_font, fontSize=11,
                                     textColor=colors.HexColor('#2E8B8F'), alignment=TA_CENTER, spaceAfter=2)
            story.append(Paragraph(ar(mcfg['report_header_subtitle']), s_style))
        if mcfg.get('report_header_title') or mcfg.get('report_header_subtitle'):
            story.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#1A555C'), spaceAfter=8))

        title_style = ParagraphStyle('title', fontName=arabic_font, fontSize=13,
                                     textColor=colors.HexColor('#1A555C'), spaceAfter=6)
        story.append(Paragraph(f'ARS — {title_txt}', title_style))
        from datetime import date
        date_style = ParagraphStyle('date', fontName=arabic_font, fontSize=9,
                                    textColor=colors.grey, spaceAfter=12)
        story.append(Paragraph(f'Report Date: {date.today()}', date_style))
        story.append(Spacer(1, 0.3*cm))

        if rows:
            headers = list(rows[0].keys())
            col_w = (page_size[0] - 3*cm) / len(headers)
            def _cell(v):
                s = str(v) if v else ''
                import re
                return ar(s) if re.search(r'[\u0600-\u06ff]', s) else s
            data = [headers] + [[_cell(r.get(h,'')) for h in headers] for r in rows]
            t = Table(data, repeatRows=1, colWidths=[col_w]*len(headers))
            t.setStyle(TableStyle([
                ('BACKGROUND',    (0,0), (-1,0), colors.HexColor('#1A555C')),
                ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
                ('FONTNAME',      (0,0), (-1,-1), arabic_font),
                ('FONTSIZE',      (0,0), (-1,0), 9),
                ('FONTSIZE',      (0,1), (-1,-1), 8),
                ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
                ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
                ('GRID',          (0,0), (-1,-1), 0.4, colors.HexColor('#dee2e6')),
                ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.white, colors.HexColor('#EEF4F5')]),
                ('PADDING',       (0,0), (-1,-1), 5),
                ('TOPPADDING',    (0,0), (-1,0), 8),
                ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ]))
            story.append(t)
        else:
            story.append(Paragraph('No data available.', styles['Normal']))

        doc.build(story)
        buf.seek(0)
        return Response(buf.read(), mimetype='application/pdf',
                        headers={'Content-Disposition': f'attachment;filename=ARS_Report_{date.today()}.pdf'})
    except ImportError:
        flash_msg('يرجى تثبيت reportlab: pip install reportlab', 'danger')
        return redirect(url_for('reports.index'))


# ── Export Excel ──────────────────────────────────────────────────────────────
@reports_bp.route('/export/excel')
@login_required
def export_excel():
    from flask import Response
    import io
    db    = get_db()
    perms = get_permissions()
    rtype  = request.args.get('type', 'my_reservations')
    filter = request.args.get('filter', 'all')
    start  = request.args.get('start', '')
    end    = request.args.get('end', '')
    rows, title_txt = _get_report_data(db, perms, rtype, filter, start, end)

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.chart import BarChart, Reference
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title_txt[:30]
        header_fill = PatternFill('solid', fgColor='1A555C')
        header_font = Font(color='FFFFFF', bold=True)
        if rows:
            headers = list(rows[0].keys())
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(1, ci, h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(
                    bottom=Side(style='medium', color='FFFFFF'),
                    right=Side(style='thin', color='2E8B8F'))
            ws.row_dimensions[1].height = 28
            for ri, row in enumerate(rows, 2):
                for ci, h in enumerate(headers, 1):
                    val = row.get(h, '')
                    cell = ws.cell(ri, ci, val)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if ri % 2 == 0:
                        cell.fill = PatternFill('solid', fgColor='EEF4F5')
                ws.row_dimensions[ri].height = 20
            # Auto-fit column widths
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except: pass
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)
            # Freeze top row
            ws.freeze_panes = 'A2' 
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return Response(buf.read(),
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition': 'attachment;filename=ARS_Report.xlsx'})
    except ImportError:
        flash_msg('يرجى تثبيت openpyxl: pip install openpyxl', 'danger')
        return redirect(url_for('reports.index'))
